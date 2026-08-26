"""Read task, trajectory, and immutable tree-run data for the web application."""

from __future__ import annotations

import json
import re
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import quote

from openpyxl import load_workbook
from PIL import Image

from .trajectories_tree.tree_builder import parse_action


BACKEND_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BACKEND_DIR.parent
WORKSPACE_DIR = PROJECT_ROOT / "backend_workspace"
TRAJECTORY_ROOT = WORKSPACE_DIR / "rollout_trajectories"
ANNOTATED_XLSX = WORKSPACE_DIR / "annotated_trajectories.xlsx"
TREE_RUNS_DIR = WORKSPACE_DIR / "trajectory_tree_runs"
TREE_JOBS_DIR = WORKSPACE_DIR / "trajectory_tree_jobs"
QUALITY_JOBS_DIR = WORKSPACE_DIR / "trajectory_quality_jobs"
QUALITY_RESULTS_DIR = WORKSPACE_DIR / "trajectory_quality_results"

GOAL_RE = re.compile(
    r"\*\*原始目标\*\*\s*[:：]\s*(.*?)(?=\r?\n\s*\r?\n|\Z)",
    re.DOTALL,
)
STEP_RE = re.compile(r"step(\d+)", re.IGNORECASE)
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
BBOX_WRITE_LOCK = threading.Lock()


@dataclass(frozen=True)
class TaskMetadata:
    task_id: str
    goal: str
    warning: str
    first_trajectory: str


def _trajectory_number(task_id: str, path: Path) -> int | None:
    match = re.fullmatch(re.escape(task_id) + r"-(\d+)", path.name)
    return int(match.group(1)) if match else None


def first_trajectory_dir(task_dir: Path) -> Path | None:
    candidates = [
        (number, child)
        for child in task_dir.iterdir()
        if child.is_dir()
        and (number := _trajectory_number(task_dir.name, child)) is not None
    ]
    return min(candidates, key=lambda item: item[0])[1] if candidates else None


def _message_text(content: Any) -> str:
    if isinstance(content, str):
        return content
    if isinstance(content, list):
        return "\n".join(
            str(item.get("text", ""))
            for item in content
            if isinstance(item, dict) and item.get("type") == "text"
        )
    return ""


def extract_original_goal(request_path: Path) -> str:
    payload = json.loads(request_path.read_text(encoding="utf-8-sig"))
    for message in payload.get("messages", []):
        if not isinstance(message, dict) or message.get("role") != "user":
            continue
        match = GOAL_RE.search(_message_text(message.get("content")))
        if match:
            return " ".join(match.group(1).strip().splitlines())
    return ""


def discover_tasks(trajectory_root: Path = TRAJECTORY_ROOT) -> dict[str, TaskMetadata]:
    if not trajectory_root.is_dir():
        return {}
    tasks: dict[str, TaskMetadata] = {}
    for task_dir in sorted(
        (path for path in trajectory_root.iterdir() if path.is_dir()),
        key=lambda path: path.name.casefold(),
    ):
        first = first_trajectory_dir(task_dir)
        if first is None:
            continue
        request_path = first / "turn001_orch_model_request.json"
        warning = ""
        goal = ""
        if request_path.is_file():
            try:
                goal = extract_original_goal(request_path)
            except (OSError, ValueError, json.JSONDecodeError) as exc:
                warning = f"无法读取原始目标：{exc}"
        else:
            warning = f"缺少 {request_path.name}"
        if not goal:
            goal = task_dir.name
            warning = warning or "未找到 **原始目标**，已回退为任务 ID"
        tasks[task_dir.name] = TaskMetadata(
            task_id=task_dir.name,
            goal=goal,
            warning=warning,
            first_trajectory=first.name,
        )
    return tasks


def task_id_from_resource(value: str) -> str:
    normalized = value.replace("\\", "/").strip("/")
    return normalized.split("/", 1)[0] if normalized else ""


def _step_number(image: str, fallback: int) -> int:
    match = STEP_RE.search(Path(image).name)
    return int(match.group(1)) if match else fallback


def asset_url(relative_path: str) -> str:
    normalized = relative_path.replace("\\", "/").strip("/")
    return "/api/assets/" + quote(normalized, safe="/")


def load_annotated_trajectories(
    xlsx_path: Path = ANNOTATED_XLSX,
) -> dict[str, list[dict[str, Any]]]:
    """Return task -> trajectories, preserving every annotated workbook row."""
    if not xlsx_path.is_file():
        return {}
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = {
            str(cell.value).strip(): cell.column
            for cell in sheet[1]
            if cell.value is not None
        }
        required = {"image", "xml", "action", "summary", "actions_box"}
        missing = sorted(required - headers.keys())
        if missing:
            raise ValueError(f"Excel 缺少必要列：{', '.join(missing)}")
        trajectory_column = headers.get("文件夹名", 1)
        grouped: dict[str, dict[str, list[dict[str, Any]]]] = {}
        for row_index in range(2, sheet.max_row + 1):
            trajectory = str(sheet.cell(row_index, trajectory_column).value or "").strip()
            image = str(sheet.cell(row_index, headers["image"]).value or "").strip()
            if not trajectory or not image:
                continue
            task_id = task_id_from_resource(image)
            if not task_id:
                continue
            action_text = str(sheet.cell(row_index, headers["action"]).value or "")
            steps = grouped.setdefault(task_id, {}).setdefault(trajectory, [])
            steps.append(
                {
                    "step": _step_number(image, len(steps) + 1),
                    "excel_row": row_index,
                    "image": image,
                    "image_url": asset_url(image),
                    "xml": str(sheet.cell(row_index, headers["xml"]).value or "").strip(),
                    "action_text": action_text,
                    "action": parse_action(action_text),
                    "action_summary": str(
                        sheet.cell(row_index, headers["summary"]).value or ""
                    ),
                    "actions_box": str(
                        sheet.cell(row_index, headers["actions_box"]).value or ""
                    ),
                }
            )
    finally:
        workbook.close()

    result: dict[str, list[dict[str, Any]]] = {}
    for task_id, trajectories in grouped.items():
        result[task_id] = []
        for trajectory, steps in sorted(trajectories.items(), key=lambda item: item[0]):
            steps.sort(key=lambda step: (step["step"], step["excel_row"]))
            result[task_id].append(
                {"trajectory_id": trajectory, "step_count": len(steps), "steps": steps}
            )
    return result


def trajectory_summaries(
    task_id: str,
    xlsx_path: Path = ANNOTATED_XLSX,
) -> list[dict[str, Any]]:
    """Return only trajectory names and counts for the task list UI."""
    return load_trajectory_index(xlsx_path).get(task_id, [])


def load_trajectory_index(
    xlsx_path: Path = ANNOTATED_XLSX,
) -> dict[str, list[dict[str, Any]]]:
    """Scan only identifiers/counts, avoiding action parsing until a trajectory opens."""
    if not xlsx_path.is_file():
        return {}
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = {
            str(value).strip(): index
            for index, value in enumerate(header_values)
            if value is not None
        }
        image_column = headers.get("image")
        trajectory_column = headers.get("文件夹名", 0)
        if image_column is None:
            raise ValueError("Excel 缺少必要列：image")
        counts: dict[str, dict[str, int]] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            trajectory = str(row[trajectory_column] or "").strip()
            image = str(row[image_column] or "").strip()
            task_id = task_id_from_resource(image)
            if trajectory and task_id:
                task_counts = counts.setdefault(task_id, {})
                task_counts[trajectory] = task_counts.get(trajectory, 0) + 1
    finally:
        workbook.close()
    return {
        task_id: [
            {"trajectory_id": trajectory, "step_count": count}
            for trajectory, count in sorted(trajectories.items())
        ]
        for task_id, trajectories in counts.items()
    }


def load_annotated_trajectory(
    task_id: str,
    trajectory_id: str,
    xlsx_path: Path = ANNOTATED_XLSX,
) -> dict[str, Any] | None:
    """Load one trajectory on demand after its name is selected in the UI."""
    if not xlsx_path.is_file():
        return None
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    steps: list[dict[str, Any]] = []
    try:
        sheet = workbook.active
        header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = {
            str(value).strip(): index
            for index, value in enumerate(header_values)
            if value is not None
        }
        required = {"image", "xml", "action", "summary", "actions_box"}
        missing = sorted(required - headers.keys())
        if missing:
            raise ValueError(f"Excel 缺少必要列：{', '.join(missing)}")
        trajectory_column = headers.get("文件夹名", 0)
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            current_trajectory = str(row[trajectory_column] or "").strip()
            image = str(row[headers["image"]] or "").strip()
            if current_trajectory != trajectory_id or task_id_from_resource(image) != task_id:
                continue
            action_text = str(row[headers["action"]] or "")
            steps.append(
                {
                    "step": _step_number(image, len(steps) + 1),
                    "excel_row": row_index,
                    "image": image,
                    "image_url": asset_url(image),
                    "xml": str(row[headers["xml"]] or "").strip(),
                    "action_text": action_text,
                    "action": parse_action(action_text),
                    "action_summary": str(row[headers["summary"]] or ""),
                    "actions_box": str(row[headers["actions_box"]] or ""),
                }
            )
    finally:
        workbook.close()
    if not steps:
        return None
    steps.sort(key=lambda step: (step["step"], step["excel_row"]))
    return {"trajectory_id": trajectory_id, "step_count": len(steps), "steps": steps}


def _format_manual_actions_box(
    action: dict[str, Any], bbox: tuple[int, int, int, int]
) -> str:
    kind = str(action.get("action", "")).lower()
    tagged = f"<bbox>[{','.join(str(value) for value in bbox)}]</bbox>"
    if kind == "click":
        return f"click(bbox={tagged})"
    if kind == "long_press":
        return f"long_press(bbox={tagged})"
    if kind == "swipe":
        start = action.get("start_coordinate")
        end = action.get("end_coordinate")
        if not isinstance(start, list) or len(start) < 2 or not isinstance(end, list) or len(end) < 2:
            raise ValueError("swipe 缺少起止坐标")
        dx = float(end[0]) - float(start[0])
        dy = float(end[1]) - float(start[1])
        direction = (
            ("left" if dx < 0 else "right")
            if abs(dx) > abs(dy)
            else ("up" if dy < 0 else "down")
        )
        return f"swipe_screen(bbox={tagged}, direction={direction})"
    raise ValueError(f"动作 {kind or 'unknown'} 不支持 bbox")


def update_action_bbox(
    task_id: str,
    trajectory_id: str,
    step: int,
    excel_row: int,
    bbox: tuple[int, int, int, int],
    *,
    xlsx_path: Path = ANNOTATED_XLSX,
    trajectory_root: Path = TRAJECTORY_ROOT,
) -> str:
    """Validate and atomically persist a manually redrawn action bbox."""
    if not xlsx_path.is_file():
        raise FileNotFoundError("标注轨迹 Excel 不存在")
    if excel_row < 2:
        raise ValueError("无效的 Excel 行号")
    x1, y1, x2, y2 = (int(value) for value in bbox)
    if x1 < 0 or y1 < 0 or x2 <= x1 or y2 <= y1:
        raise ValueError("bbox 必须是有效的 [x1,y1,x2,y2]")

    with BBOX_WRITE_LOCK:
        workbook = load_workbook(xlsx_path)
        temporary_path = xlsx_path.with_name(f".{xlsx_path.stem}.bbox-edit.tmp{xlsx_path.suffix}")
        try:
            sheet = workbook.active
            headers = {
                str(cell.value).strip(): cell.column
                for cell in sheet[1]
                if cell.value is not None
            }
            required = {"文件夹名", "image", "action", "actions_box"}
            missing = sorted(required - headers.keys())
            if missing:
                raise ValueError(f"Excel 缺少必要列：{', '.join(missing)}")
            if excel_row > sheet.max_row:
                raise ValueError("Excel 行号超出范围")

            row_trajectory = str(sheet.cell(excel_row, headers["文件夹名"]).value or "").strip()
            image_value = str(sheet.cell(excel_row, headers["image"]).value or "").strip()
            if row_trajectory != trajectory_id:
                raise ValueError("Excel 行与轨迹不匹配")
            if task_id_from_resource(image_value) != task_id:
                raise ValueError("Excel 行与任务不匹配")
            if _step_number(image_value, -1) != step:
                raise ValueError("Excel 行与 step 不匹配")

            image_path = resolve_image_asset(image_value, trajectory_root)
            with Image.open(image_path) as image:
                width, height = image.size
            if x2 > width or y2 > height:
                raise ValueError(f"bbox 超出截图范围 {width}x{height}")

            action_text = str(sheet.cell(excel_row, headers["action"]).value or "")
            actions_box = _format_manual_actions_box(parse_action(action_text), (x1, y1, x2, y2))
            sheet.cell(excel_row, headers["actions_box"]).value = actions_box
            if temporary_path.exists():
                temporary_path.unlink()
            workbook.save(temporary_path)
            workbook.close()
            temporary_path.replace(xlsx_path)
            return actions_box
        finally:
            workbook.close()
            if temporary_path.exists():
                temporary_path.unlink()


def task_summaries(
    trajectory_root: Path = TRAJECTORY_ROOT,
    xlsx_path: Path = ANNOTATED_XLSX,
) -> list[dict[str, Any]]:
    metadata = discover_tasks(trajectory_root)
    annotated = load_trajectory_index(xlsx_path)
    values: list[dict[str, Any]] = []
    for task_id, item in metadata.items():
        trajectories = annotated.get(task_id, [])
        values.append(
            {
                "task_id": task_id,
                "goal": item.goal,
                "warning": item.warning,
                "first_trajectory": item.first_trajectory,
                "trajectory_count": len(trajectories),
                "step_count": sum(value["step_count"] for value in trajectories),
                "annotated": bool(trajectories),
            }
        )
    return values


def resolve_image_asset(relative_path: str, root: Path = TRAJECTORY_ROOT) -> Path:
    normalized = relative_path.replace("\\", "/").lstrip("/")
    candidate = (root / Path(normalized)).resolve()
    resolved_root = root.resolve()
    try:
        candidate.relative_to(resolved_root)
    except ValueError as exc:
        raise ValueError("资源路径超出轨迹目录") from exc
    if candidate.suffix.lower() not in IMAGE_EXTENSIONS:
        raise ValueError("仅允许访问轨迹图片")
    if not candidate.is_file():
        raise FileNotFoundError(relative_path)
    return candidate


def list_tree_runs(runs_dir: Path = TREE_RUNS_DIR) -> list[dict[str, Any]]:
    if not runs_dir.is_dir():
        return []
    runs: list[dict[str, Any]] = []
    for directory in runs_dir.iterdir():
        manifest_path = directory / "manifest.json"
        if directory.is_dir() and not directory.name.startswith(".") and manifest_path.is_file():
            try:
                manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            except (OSError, ValueError, json.JSONDecodeError):
                continue
            if isinstance(manifest, dict) and manifest.get("run_id") == directory.name:
                runs.append(manifest)
    return sorted(runs, key=lambda item: str(item.get("completed_at", "")), reverse=True)


def find_tree_run(run_id: str, runs_dir: Path = TREE_RUNS_DIR) -> dict[str, Any] | None:
    return next((item for item in list_tree_runs(runs_dir) if item.get("run_id") == run_id), None)
