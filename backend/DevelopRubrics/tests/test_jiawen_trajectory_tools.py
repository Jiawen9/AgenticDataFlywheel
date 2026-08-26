from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from trajectory_tools.gui_trajectory_excel import collect_rollouts, discover_trajectory_directories, export_trajectory_workbook


def _write_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False), encoding="utf-8")


def _trajectory(root: Path, task: str, trajectory: str, *, screenshot_suffix: str = "input_stability") -> Path:
    directory = root / task / trajectory
    _write_json(
        directory / "_trajectory_for_evaluate.json",
        {
            "task": "打开应用并完成任务",
            "session_id": "raw-session",
            "actions_flat": [{"global_step": 1, "action": {"action": "click", "coordinate": [10, 20]}}],
        },
    )
    _write_json(directory / "step001_vla_model_request.json", {"openai_request": {"messages": []}})
    _write_json(directory / "step001_vla_model_response.json", {"content": "<summary>点击目标</summary>"})
    (directory / f"step001_vla_{screenshot_suffix}.jpg").write_bytes(b"image")
    return directory


def test_recursive_discovery_ignores_prefetch(tmp_path: Path) -> None:
    expected = _trajectory(tmp_path, "TASK-1", "TASK-1-1")
    _write_json(tmp_path / "TASK-1" / "TASK-1-1" / "_prefetch_staging" / "x_trajectory_for_evaluate.json", {})
    assert discover_trajectory_directories(tmp_path) == [expected]


def test_input_screenshot_fallback_and_executed_action(tmp_path: Path) -> None:
    _trajectory(tmp_path, "TASK-1", "TASK-1-1", screenshot_suffix="input")
    tasks, trajectories = collect_rollouts(tmp_path)
    assert list(tasks) == ["TASK-1"]
    assert trajectories[0].trajectory_id == "TASK-1-1"
    step = trajectories[0].steps[0]
    assert json.loads(step.action) == {"action": "click", "coordinate": [10, 20]}
    assert step.action_input["screenshot"] == "TASK-1\\TASK-1-1\\step001_vla_input.jpg"
    assert "fallback" in step.source_warning


def test_atomic_workbook_has_normalized_sheets(tmp_path: Path) -> None:
    _trajectory(tmp_path / "source", "TASK-1", "TASK-1-1", screenshot_suffix="done")
    output = tmp_path / "output.xlsx"
    assert export_trajectory_workbook(tmp_path / "source", output) == (1, 1, 1)
    workbook = load_workbook(output, read_only=True)
    try:
        assert workbook.sheetnames == ["Tasks", "Trajectories", "Steps"]
        assert workbook["Steps"].max_row == 2
    finally:
        workbook.close()
