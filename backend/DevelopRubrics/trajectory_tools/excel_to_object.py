"""Load the normalized GUI workbook into AdaRubric model objects."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _rows(sheet: Any) -> list[dict[str, Any]]:
    iterator = sheet.iter_rows(values_only=True)
    headers = next(iterator, None)
    if not headers:
        return []
    names = [str(value) for value in headers]
    return [dict(zip(names, row)) for row in iterator if any(value is not None for value in row)]


def load_objects(workbook_path: Path) -> tuple[dict[str, Any], list[Any]]:
    from adarubric import TaskDescription, Trajectory, TrajectoryStep

    workbook_path = Path(workbook_path).resolve()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"trajectory workbook not found: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        required = {"Tasks", "Trajectories", "Steps"}
        missing = required.difference(workbook.sheetnames)
        if missing:
            raise ValueError(f"workbook is missing required sheets: {sorted(missing)}")

        tasks: dict[str, Any] = {}
        for row in _rows(workbook["Tasks"]):
            task = TaskDescription(
                task_id=str(row["task_id"]),
                instruction=str(row["instruction"]),
                domain=str(row["domain"]),
                complexity=str(row["complexity"]),
                context=json.loads(str(row["context_json"])),
                expected_tools=json.loads(str(row["expected_tools_json"])),
            )
            tasks[task.task_id] = task

        steps_by_trajectory: dict[str, list[Any]] = {}
        for row in _rows(workbook["Steps"]):
            trajectory_id = str(row["trajectory_id"])
            step = TrajectoryStep(
                step_id=int(row["step_id"]),
                action=str(row["action"]),
                action_input=json.loads(str(row["action_input_json"])),
                observation=str(row["observation"]),
            )
            steps_by_trajectory.setdefault(trajectory_id, []).append(step)

        trajectories: list[Any] = []
        for row in _rows(workbook["Trajectories"]):
            trajectory_id = str(row["trajectory_id"])
            task_id = str(row["task_id"])
            if task_id not in tasks:
                raise ValueError(f"trajectory {trajectory_id!r} references unknown task {task_id!r}")
            trajectories.append(
                Trajectory(
                    trajectory_id=trajectory_id,
                    task_id=task_id,
                    steps=sorted(steps_by_trajectory.get(trajectory_id, []), key=lambda item: item.step_id),
                    final_answer=str(row["final_answer"]) if row["final_answer"] is not None else None,
                    metadata=json.loads(str(row["metadata_json"])),
                )
            )
        return tasks, trajectories
    finally:
        workbook.close()
