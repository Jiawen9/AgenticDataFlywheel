"""Convert rollout artifacts into a normalized, AdaRubric-ready workbook."""

from __future__ import annotations

import base64
import hashlib
import json
import mimetypes
import os
import re
import threading
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Protocol

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


DOMAIN = "Mobile GUI Agent"
SYSTEM_PROMPT = "You are a GUI agent. Complete the requested mobile GUI task from screenshots."
EXPECTED_TOOLS = [
    '{"action":"click","coordinate":[x,y]}',
    '{"action":"long_press","coordinate":[x,y]}',
    '{"action":"type","text":""}',
    '{"action":"open","text":"app_name"}',
    '{"action":"swipe","start_coordinate":[x1,y1],"end_coordinate":[x2,y2]}',
    '{"action":"system_button","button":"back|home|menu|enter"}',
    '{"action":"wait"}',
    '{"action":"terminate","status":"success|failure"}',
]


class Summarizer(Protocol):
    model: str

    def summarize_screenshot(self, image_path: Path, task: str, summary: str, action: str) -> str: ...
    def summarize_trajectory(self, task: str, trajectory_id: str, steps: list["StepRecord"]) -> str: ...


@dataclass
class StepRecord:
    trajectory_id: str
    task_id: str
    step_id: int
    action: str
    action_input: dict[str, Any]
    observation: str
    request_file: str
    response_file: str
    screenshot_path: str
    source_warning: str


@dataclass
class TrajectoryRecord:
    trajectory_id: str
    task_id: str
    source_directory: str
    final_answer: str
    metadata: dict[str, Any]
    steps: list[StepRecord]


@dataclass
class TaskRecord:
    task_id: str
    task_text: str


def _read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def _relative(path: Path, root: Path) -> str:
    return str(path.resolve().relative_to(root.resolve()))


def discover_trajectory_directories(root: Path) -> list[Path]:
    root = root.resolve()
    found: set[Path] = set()
    for marker in root.rglob("*_trajectory_for_evaluate.json"):
        directory = marker.parent
        relative_parts = directory.relative_to(root).parts
        if any(part.startswith("_") for part in relative_parts):
            continue
        found.add(directory)
    return sorted(found, key=lambda item: tuple(part.lower() for part in item.relative_to(root).parts))


def _screenshot(directory: Path, step_id: int) -> tuple[Path | None, str]:
    prefix = f"step{step_id:03d}_vla"
    candidates = [
        directory / f"{prefix}_input_stability.jpg",
        directory / f"{prefix}_input.jpg",
        directory / f"{prefix}_done.jpg",
    ]
    for index, path in enumerate(candidates):
        if path.is_file():
            warning = "" if index == 0 else f"screenshot fallback: {path.name}"
            return path, warning
    return None, ""


def _extract_summary(response: dict[str, Any]) -> str:
    content = str(response.get("content", ""))
    match = re.search(r"<summary>\s*(.*?)\s*</summary>", content, re.DOTALL | re.IGNORECASE)
    if match:
        return match.group(1).strip()
    thought = re.search(r"<thought>\s*(.*?)\s*</thought>", content, re.DOTALL | re.IGNORECASE)
    return thought.group(1).strip() if thought else ""


def collect_rollouts(root: Path, summarizer: Summarizer | None = None) -> tuple[dict[str, TaskRecord], list[TrajectoryRecord]]:
    root = root.resolve()
    tasks: dict[str, TaskRecord] = {}
    trajectories: list[TrajectoryRecord] = []
    for directory in discover_trajectory_directories(root):
        marker = next(iter(sorted(directory.glob("*_trajectory_for_evaluate.json"))))
        evaluate = _read_json(marker)
        task_text = str(evaluate.get("task", "")).strip()
        if not task_text:
            raise ValueError(f"missing task text: {marker}")
        try:
            task_id = directory.relative_to(root).parts[0]
        except IndexError as exc:
            raise ValueError(f"trajectory directory must be under a task directory: {directory}") from exc
        trajectory_id = directory.name
        tasks.setdefault(task_id, TaskRecord(task_id=task_id, task_text=task_text))

        actions = [item for item in evaluate.get("actions_flat", []) if isinstance(item, dict)]
        actions.sort(key=lambda item: int(item["global_step"]))
        if not actions:
            raise ValueError(f"no actions_flat entries: {marker}")
        steps: list[StepRecord] = []
        warnings: list[str] = []
        for item in actions:
            step_id = int(item["global_step"])
            prefix = f"step{step_id:03d}_vla"
            request_path = directory / f"{prefix}_model_request.json"
            response_path = directory / f"{prefix}_model_response.json"
            if not request_path.is_file() or not response_path.is_file():
                raise FileNotFoundError(f"missing request/response for {trajectory_id} step {step_id}")
            image_path, warning = _screenshot(directory, step_id)
            if image_path is None:
                raise FileNotFoundError(f"missing screenshot for {trajectory_id} step {step_id}")
            response = _read_json(response_path)
            summary = _extract_summary(response)
            action_obj = item.get("action")
            if not isinstance(action_obj, dict):
                raise ValueError(f"invalid executed action for {trajectory_id} step {step_id}")
            action = json.dumps(action_obj, ensure_ascii=False, separators=(",", ":"))
            screenshot_relative = _relative(image_path, root)
            if summarizer is None:
                observation = f"未调用视觉模型；当前截图：{screenshot_relative}"
            else:
                observation = summarizer.summarize_screenshot(image_path, task_text, summary, action)
            if warning:
                warnings.append(f"step {step_id}: {warning}")
            steps.append(
                StepRecord(
                    trajectory_id=trajectory_id,
                    task_id=task_id,
                    step_id=step_id,
                    action=action,
                    action_input={"summary": summary, "screenshot": screenshot_relative},
                    observation=observation,
                    request_file=_relative(request_path, root),
                    response_file=_relative(response_path, root),
                    screenshot_path=screenshot_relative,
                    source_warning=warning,
                )
            )
        final_answer = (
            summarizer.summarize_trajectory(task_text, trajectory_id, steps)
            if summarizer is not None
            else f"轨迹执行了 {len(steps)} 步；最后观察：{steps[-1].observation}"
        )
        trajectories.append(
            TrajectoryRecord(
                trajectory_id=trajectory_id,
                task_id=task_id,
                source_directory=str(directory.relative_to(root)),
                final_answer=final_answer,
                metadata={
                    "source_session_id": str(evaluate.get("session_id", "")),
                    "source_directory": str(directory.relative_to(root)),
                    "source_warnings": warnings,
                    "observation_model": summarizer.model if summarizer else "not_called",
                },
                steps=steps,
            )
        )
    if not trajectories:
        raise ValueError(f"no formal trajectories found under {root}")
    return tasks, trajectories


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _style(sheet: Any) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(output: Path, tasks: dict[str, TaskRecord], trajectories: list[TrajectoryRecord]) -> None:
    workbook = Workbook()
    tasks_sheet = workbook.active
    tasks_sheet.title = "Tasks"
    trajectories_sheet = workbook.create_sheet("Trajectories")
    steps_sheet = workbook.create_sheet("Steps")
    tasks_sheet.append(["task_id", "task_text", "instruction", "domain", "complexity", "context_json", "expected_tools_json"])
    for task in tasks.values():
        tasks_sheet.append([task.task_id, task.task_text, f"{SYSTEM_PROMPT}\n\nTask:\n{task.task_text}", DOMAIN, "complex", _json({"task_description": task.task_text}), _json(EXPECTED_TOOLS)])
    trajectories_sheet.append(["trajectory_id", "task_id", "source_directory", "step_count", "final_answer", "metadata_json"])
    for trajectory in trajectories:
        trajectories_sheet.append([trajectory.trajectory_id, trajectory.task_id, trajectory.source_directory, len(trajectory.steps), trajectory.final_answer, _json(trajectory.metadata)])
    steps_sheet.append(["trajectory_id", "task_id", "step_id", "action", "action_input_json", "observation", "request_file", "response_file", "screenshot_path", "source_warning"])
    for trajectory in trajectories:
        for step in trajectory.steps:
            steps_sheet.append([step.trajectory_id, step.task_id, step.step_id, step.action, _json(step.action_input), step.observation, step.request_file, step.response_file, step.screenshot_path, step.source_warning])
    for sheet in (tasks_sheet, trajectories_sheet, steps_sheet):
        _style(sheet)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


class QwenSummarizer:
    def __init__(self, model: str, base_url: str, api_key: str, cache_path: Path) -> None:
        from openai import OpenAI

        self.model = model
        self.client = OpenAI(api_key=api_key, base_url=base_url, max_retries=2)
        self.cache_path = cache_path
        self.cache = json.loads(cache_path.read_text(encoding="utf-8")) if cache_path.is_file() else {}
        self._cache_lock = threading.RLock()

    def _save_unlocked(self) -> None:
        self.cache_path.parent.mkdir(parents=True, exist_ok=True)
        temporary = self.cache_path.with_suffix(self.cache_path.suffix + ".tmp")
        temporary.write_text(json.dumps(self.cache, ensure_ascii=False, indent=2), encoding="utf-8")
        for attempt in range(6):
            try:
                os.replace(temporary, self.cache_path)
                return
            except PermissionError:
                if attempt == 5:
                    raise
                time.sleep(0.1 * (attempt + 1))

    def _save(self) -> None:
        with self._cache_lock:
            self._save_unlocked()

    def _complete(self, key: str, messages: list[dict[str, Any]], max_tokens: int) -> str:
        with self._cache_lock:
            cached = self.cache.get(key)
        if cached is not None:
            return str(cached)
        response = self.client.chat.completions.create(
            model=self.model,
            messages=messages,
            temperature=0.0,
            max_tokens=max_tokens,
            extra_body={"chat_template_kwargs": {"enable_thinking": False}},
        )
        text = (response.choices[0].message.content or "").strip()
        if not text:
            raise ValueError("Qwen returned an empty summary")
        with self._cache_lock:
            cached = self.cache.get(key)
            if cached is not None:
                return str(cached)
            self.cache[key] = text
            self._save_unlocked()
            return text

    def summarize_screenshot(self, image_path: Path, task: str, summary: str, action: str) -> str:
        image = image_path.read_bytes()
        digest = hashlib.sha256(image).hexdigest()
        key = "observation:" + hashlib.sha256(f"{self.model}\n{digest}\n{task}\n{summary}\n{action}".encode()).hexdigest()
        mime = mimetypes.guess_type(image_path.name)[0] or "image/jpeg"
        data_url = f"data:{mime};base64," + base64.b64encode(image).decode("ascii")
        return self._complete(key, [
            {"role": "system", "content": "你是移动端GUI截图观察器。只描述截图中实际可见的页面、文字、控件、弹窗、加载或错误状态，不推测动作成功，不输出JSON。使用具体的中文。"},
            {"role": "user", "content": [
                {"type": "text", "text": f"总任务：{task}\n动作摘要：{summary}\n执行动作：{action}\n请描述当前截图。"},
                {"type": "image_url", "image_url": {"url": data_url}},
            ]},
        ], 900)

    def summarize_trajectory(self, task: str, trajectory_id: str, steps: list[StepRecord]) -> str:
        evidence = [{"step_id": item.step_id, "action": item.action, "observation": item.observation} for item in steps]
        payload = json.dumps(evidence, ensure_ascii=False, separators=(",", ":"))
        key = "trajectory-evidence-v2:" + hashlib.sha256(f"{self.model}\n{task}\n{trajectory_id}\n{payload}".encode()).hexdigest()
        return self._complete(key, [
            {"role": "system", "content": "根据 Task、逐步 Action 和 Observation，用一段简洁中文总结与任务相关的关键可见状态变化以及最终可见页面。只写可观察事实；不得根据 terminate 声明判断成功、失败、完成或未完成，不评价动作，不规划下一步。描述最终可见状态后立即结束。"},
            {"role": "user", "content": f"任务：{task}\n轨迹证据：{payload}"},
        ], 1200)


def export_trajectory_workbook(root: Path, output: Path, summarizer: Summarizer | None = None) -> tuple[int, int, int]:
    tasks, trajectories = collect_rollouts(root, summarizer)
    temporary = output.with_name(f".{output.stem}.tmp{output.suffix}")
    write_workbook(temporary, tasks, trajectories)
    temporary.replace(output)
    return len(tasks), len(trajectories), sum(len(item.steps) for item in trajectories)
