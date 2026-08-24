"""Export GUIAgent VLA trajectory steps to an Excel workbook."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


STEP_RESPONSE_RE = re.compile(r"^step(\d+)_vla_model_response\.json$", re.IGNORECASE)
TOOL_CALL_TAG_RE = re.compile(r"<tool_call>\s*", re.IGNORECASE)
SUMMARY_RE = re.compile(r"<summary>\s*(.*?)(?:</summary>|$)", re.DOTALL | re.IGNORECASE)
IGNORED_DIRECTORY_NAMES = {"_prefetch_staging"}


def json_object_after_tool_call(content: str, start: int) -> tuple[Any, int]:
    """Decode the first JSON value after a tool-call tag, closing tag optional."""
    decoder = json.JSONDecoder()
    remainder = content[start:].lstrip()
    value, end = decoder.raw_decode(remainder)
    return value, start + (len(content[start:]) - len(remainder)) + end


def extract_response_fields(response_path: Path) -> tuple[str, str]:
    """Return tool-call payloads and summary text from a VLA response."""
    with response_path.open("r", encoding="utf-8-sig") as file:
        response = json.load(file)

    content = response.get("content", "")
    if not isinstance(content, str):
        raise ValueError("the 'content' field is not a string")

    actions: list[Any] = []
    position = 0
    while match := TOOL_CALL_TAG_RE.search(content, position):
        try:
            action, position = json_object_after_tool_call(content, match.end())
            actions.append(action)
        except json.JSONDecodeError as exc:
            raise ValueError("no valid JSON object found after <tool_call>") from exc

    if not actions:
        raise ValueError("no <tool_call> followed by JSON found in 'content'")

    value: Any = actions[0] if len(actions) == 1 else actions
    action_text = json.dumps(value, ensure_ascii=False, separators=(",", ":"))
    summary_match = SUMMARY_RE.search(content)
    summary = summary_match.group(1).strip() if summary_match else ""
    return action_text, summary


def discover_case_dirs(run_dir: Path) -> list[Path]:
    """Return finalized trajectory directories, excluding nested runtime candidates."""
    case_dirs: set[Path] = set()
    for evaluation_path in run_dir.rglob("_trajectory_for_evaluate.json"):
        case_dir = evaluation_path.parent
        try:
            relative_parts = case_dir.relative_to(run_dir).parts
        except ValueError:
            continue
        if any(part in IGNORED_DIRECTORY_NAMES for part in relative_parts):
            continue
        if any(
            path.is_file() and STEP_RESPONSE_RE.match(path.name)
            for path in case_dir.iterdir()
        ):
            case_dirs.add(case_dir)
    return sorted(
        case_dirs,
        key=lambda path: path.relative_to(run_dir).as_posix().casefold(),
    )


def collect_rows(run_dir: Path) -> tuple[list[list[str]], list[str]]:
    rows: list[list[str]] = []
    warnings: list[str] = []
    run_dir = run_dir.expanduser().resolve()

    for case_dir in discover_case_dirs(run_dir):
        responses: list[tuple[int, Path]] = []
        for path in case_dir.iterdir():
            match = STEP_RESPONSE_RE.match(path.name)
            if match:
                responses.append((int(match.group(1)), path))

        for step_number, response_path in sorted(responses, key=lambda item: item[0]):
            prefix = f"step{step_number:03d}"
            image_path = case_dir / f"{prefix}_vla_input.jpg"
            input_xml_path = case_dir / f"{prefix}_vla_input_ui.xml"
            previous_prefix = f"step{step_number - 1:03d}"
            fallback_xml_path = case_dir / f"{previous_prefix}_vla_done_ui.xml"
            if input_xml_path.is_file():
                xml_value = str(input_xml_path.resolve().relative_to(run_dir))
            elif step_number > 1 and fallback_xml_path.is_file():
                xml_value = str(fallback_xml_path.resolve().relative_to(run_dir))
                warnings.append(
                    f"{case_dir.name}/{prefix}: input UI XML missing; "
                    f"using {fallback_xml_path.name}"
                )
            else:
                xml_value = "无"
                warnings.append(f"{case_dir.name}/{prefix}: no matching UI XML")

            if not image_path.is_file():
                warnings.append(f"{case_dir.name}/{prefix}: skipped because image is missing: {image_path}")
                continue
            if xml_value == "无":
                warnings.append(f"{case_dir.name}/{prefix}: skipped because UI XML is missing")
                continue

            try:
                action, summary = extract_response_fields(response_path)
            except (OSError, json.JSONDecodeError, ValueError) as exc:
                action = ""
                summary = ""
                warnings.append(f"{case_dir.name}/{prefix}: {exc}")

            rows.append([
                case_dir.name,
                str(image_path.resolve().relative_to(run_dir)),
                xml_value,
                action,
                summary,
            ])

    return rows, warnings


def write_xlsx(rows: list[list[str]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "VLA trajectories"
    sheet.append(
        ["文件夹名", "image", "xml", "action", "summary"]
    )
    for row in rows:
        sheet.append(row)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [24, 90, 90, 70, 70]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export VLA trajectory artifacts to xlsx.")
    parser.add_argument("run_dir", type=Path, help=r"Run directory, e.g. .runs\20260711_104625")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Output xlsx path (default: <run_dir>/vla_trajectories.xlsx)",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    run_dir = args.run_dir.expanduser().resolve()
    if not run_dir.is_dir():
        print(f"Error: run directory does not exist: {run_dir}", file=sys.stderr)
        return 2

    output_path = (args.output or run_dir / "vla_trajectories.xlsx").expanduser().resolve()
    rows, warnings = collect_rows(run_dir)
    if not rows:
        print(f"Error: no step*_vla_model_response.json files found under {run_dir}", file=sys.stderr)
        return 1

    write_xlsx(rows, output_path)
    print(f"Exported {len(rows)} steps to: {output_path}")
    if warnings:
        print(f"Warnings ({len(warnings)}):", file=sys.stderr)
        for warning in warnings:
            print(f"- {warning}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
