from __future__ import annotations

import json
import tempfile
import threading
import time
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook
from PIL import Image

from backend.trajectory_data import (
    discover_tasks,
    load_annotated_trajectory,
    load_annotated_trajectories,
    resolve_image_asset,
    task_summaries,
    trajectory_summaries,
    update_action_bbox,
)
from backend.tree_build_jobs import TreeBuildJobManager
from backend.quality_data import quality_manifest, quality_task
from backend.quality_jobs import QualityJobManager


def write_request(path: Path, goal: str) -> None:
    path.write_text(
        json.dumps(
            {
                "messages": [
                    {"role": "system", "content": "system"},
                    {
                        "role": "user",
                        "content": (
                            "请完成以下 GUI 操作任务：\n"
                            f"**原始目标**: {goal}\n\n"
                            "看截图与 Current State，用 tool 推进。"
                        ),
                    },
                ]
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )


def write_workbook(path: Path, rows: list[list[str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["文件夹名", "image", "xml", "action", "summary", "actions_box"])
    for row in rows:
        sheet.append(row)
    workbook.save(path)


class TrajectoryDataTests(unittest.TestCase):
    def test_discovers_numeric_first_trajectory_and_extracts_goal(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            task = root / "TASK-001"
            first = task / "TASK-001-1"
            tenth = task / "TASK-001-10"
            first.mkdir(parents=True)
            tenth.mkdir()
            write_request(first / "turn001_orch_model_request.json", "打开应用并完成目标。")
            write_request(tenth / "turn001_orch_model_request.json", "错误的第十条目标")

            result = discover_tasks(root)["TASK-001"]

            self.assertEqual(result.first_trajectory, "TASK-001-1")
            self.assertEqual(result.goal, "打开应用并完成目标。")
            self.assertEqual(result.warning, "")

    def test_missing_goal_falls_back_to_task_id(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            trajectory = root / "TASK-002" / "TASK-002-1"
            trajectory.mkdir(parents=True)
            (trajectory / "turn001_orch_model_request.json").write_text(
                json.dumps({"messages": []}), encoding="utf-8"
            )

            result = discover_tasks(root)["TASK-002"]

            self.assertEqual(result.goal, "TASK-002")
            self.assertTrue(result.warning)

    def test_workbook_is_grouped_by_task_resource_prefix(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "annotated.xlsx"
            write_workbook(
                workbook_path,
                [
                    [
                        "TASK-A-1",
                        r"TASK-A\TASK-A-1\step001_vla_input.jpg",
                        r"TASK-A\TASK-A-1\step001_vla_input_ui.xml",
                        '{"action":"click","coordinate":[500,500]}',
                        "点击按钮",
                        "click(bbox=<bbox>[1,2,30,40]</bbox>)",
                    ],
                    [
                        "TASK-B-1",
                        r"TASK-B\TASK-B-1\step001_vla_input.jpg",
                        r"TASK-B\TASK-B-1\step001_vla_input_ui.xml",
                        '{"action":"wait"}',
                        "等待",
                        "",
                    ],
                ],
            )

            grouped = load_annotated_trajectories(workbook_path)

            self.assertEqual(set(grouped), {"TASK-A", "TASK-B"})
            step = grouped["TASK-A"][0]["steps"][0]
            self.assertEqual(step["action"]["action"], "click")
            self.assertTrue(step["image_url"].startswith("/api/assets/TASK-A/"))

            summaries = trajectory_summaries("TASK-A", workbook_path)
            self.assertEqual(summaries, [{"trajectory_id": "TASK-A-1", "step_count": 1}])
            self.assertNotIn("steps", summaries[0])
            detail = load_annotated_trajectory("TASK-A", "TASK-A-1", workbook_path)
            self.assertIsNotNone(detail)
            self.assertEqual(detail["steps"][0]["action_summary"], "点击按钮")

    def test_task_summary_marks_unprocessed_tasks(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            trajectory = root / "TASK-C" / "TASK-C-1"
            trajectory.mkdir(parents=True)
            write_request(trajectory / "turn001_orch_model_request.json", "目标")
            workbook_path = root / "empty.xlsx"
            write_workbook(workbook_path, [])

            summary = task_summaries(root, workbook_path)[0]

            self.assertFalse(summary["annotated"])
            self.assertEqual(summary["step_count"], 0)

    def test_asset_resolution_rejects_traversal_and_non_images(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            image = root / "task" / "screen.jpg"
            image.parent.mkdir()
            image.write_bytes(b"image")
            text_file = root / "task" / "secret.json"
            text_file.write_text("{}", encoding="utf-8")

            self.assertEqual(resolve_image_asset("task/screen.jpg", root), image.resolve())
            with self.assertRaisesRegex(ValueError, "超出"):
                resolve_image_asset("../outside.jpg", root)
            with self.assertRaisesRegex(ValueError, "仅允许"):
                resolve_image_asset("task/secret.json", root)

    def test_manual_bbox_update_is_validated_and_persisted(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            image = root / "TASK-A" / "TASK-A-1" / "step001_vla_input.jpg"
            image.parent.mkdir(parents=True)
            Image.new("RGB", (200, 300), "white").save(image)
            workbook_path = root / "annotated.xlsx"
            write_workbook(
                workbook_path,
                [[
                    "TASK-A-1",
                    r"TASK-A\TASK-A-1\step001_vla_input.jpg",
                    r"TASK-A\TASK-A-1\step001_vla_input_ui.xml",
                    '{"action":"click","coordinate":[50,60]}',
                    "点击按钮",
                    "click(bbox=<bbox>[1,2,30,40]</bbox>)",
                ]],
            )

            value = update_action_bbox(
                "TASK-A", "TASK-A-1", 1, 2, (20, 30, 120, 180),
                xlsx_path=workbook_path,
                trajectory_root=root,
            )

            self.assertEqual(value, "click(bbox=<bbox>[20,30,120,180]</bbox>)")
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook.active["F2"].value, value)
            finally:
                workbook.close()
            with self.assertRaisesRegex(ValueError, "超出截图范围"):
                update_action_bbox(
                    "TASK-A", "TASK-A-1", 1, 2, (20, 30, 220, 180),
                    xlsx_path=workbook_path,
                    trajectory_root=root,
                )


class TreeBuildJobTests(unittest.TestCase):
    def test_successful_job_persists_progress_and_run_id(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            finished = threading.Event()

            def runner(task_ids, *, job_id, progress):
                self.assertEqual(task_ids, ["TASK-A"])
                self.assertTrue(job_id)
                progress({"total_steps": 4, "classified_steps": 2, "stage": "classifying"})
                progress({"total_steps": 4, "classified_steps": 4, "stage": "publishing"})
                finished.set()
                return "20260819_120000", {}

            manager = TreeBuildJobManager(root, runner)
            job = manager.submit(["TASK-A"])
            self.assertTrue(finished.wait(2))
            for _ in range(100):
                result = manager.get(job["job_id"])
                if result and result["status"] == "succeeded":
                    break
                time.sleep(0.01)
            manager.shutdown()

            self.assertIsNotNone(result)
            self.assertEqual(result["run_id"], "20260819_120000")
            self.assertEqual(result["percent"], 100)

    def test_runner_failure_does_not_report_a_run(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)

            def runner(_task_ids, *, job_id, progress):
                progress({"stage": "building"})
                raise RuntimeError(f"task {job_id} failed")

            manager = TreeBuildJobManager(root, runner)
            job = manager.submit(["TASK-A", "TASK-B"])
            for _ in range(100):
                result = manager.get(job["job_id"])
                if result and result["status"] == "failed":
                    break
                time.sleep(0.01)
            manager.shutdown()

            self.assertEqual(result["status"], "failed")
            self.assertIsNone(result["run_id"])
            self.assertIn("failed", result["error"])

    def test_startup_marks_incomplete_jobs_interrupted(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            path = root / "existing.json"
            path.write_text(
                json.dumps({"job_id": "existing", "status": "running", "stage": "building"}),
                encoding="utf-8",
            )

            manager = TreeBuildJobManager(root, lambda *_args, **_kwargs: ("", {}))
            result = manager.get("existing")
            manager.shutdown()

            self.assertEqual(result["status"], "interrupted")
            self.assertTrue(result["error"])


class QualityJobTests(unittest.TestCase):
    def test_list_jobs_restores_persisted_queue_in_creation_order(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            for job_id, created_at, status in (
                ("old", "2026-08-27T10:00:00+08:00", "succeeded"),
                ("queued", "2026-08-27T12:00:00+08:00", "queued"),
                ("new", "2026-08-27T13:00:00+08:00", "failed"),
            ):
                (root / f"{job_id}.json").write_text(
                    json.dumps({"job_id": job_id, "created_at": created_at, "status": status}),
                    encoding="utf-8",
                )
            (root / "broken.json").write_text("not json", encoding="utf-8")

            manager = QualityJobManager(root, lambda *_args, **_kwargs: {"run_id": ""})
            jobs = manager.list_jobs()
            manager.shutdown()

            self.assertEqual([job["job_id"] for job in jobs], ["new", "queued", "old"])

    def test_successful_quality_job_persists_trajectory_progress(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            finished = threading.Event()

            def runner(run_id, task_ids, *, job_id, progress):
                self.assertEqual((run_id, task_ids), ("run-1", ["TASK-A"]))
                progress({"stage": "evaluating", "current_task": "TASK-A", "current_trajectory": "TASK-A-1", "completed_trajectories": 1, "total_trajectories": 2, "percent": 55})
                finished.set()
                return {"run_id": run_id}

            manager = QualityJobManager(Path(temp_dir), runner)
            job = manager.submit("run-1", ["TASK-A"])
            self.assertTrue(finished.wait(2))
            for _ in range(100):
                result = manager.get(job["job_id"])
                if result and result["status"] == "succeeded":
                    break
                time.sleep(0.01)
            manager.shutdown()
            self.assertEqual(result["status"], "succeeded")
            self.assertEqual(result["percent"], 100)
            self.assertEqual(result["current_trajectory"], "TASK-A-1")

    def test_quality_result_readers_handle_missing_and_published_files(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            self.assertEqual(quality_manifest("run-1", root)["tasks"], [])
            run = root / "run-1"
            run.mkdir()
            (run / "manifest.json").write_text(json.dumps({"run_id": "run-1", "tasks": [{"task_id": "TASK-A"}]}), encoding="utf-8")
            (run / "TASK-A.json").write_text(json.dumps({"task_id": "TASK-A", "evaluations": {}}), encoding="utf-8")
            self.assertEqual(quality_manifest("run-1", root)["run_id"], "run-1")
            self.assertEqual(quality_task("run-1", "TASK-A", root)["task_id"], "TASK-A")
            self.assertIsNone(quality_task("run-1", "../secret", root))


if __name__ == "__main__":
    unittest.main()
