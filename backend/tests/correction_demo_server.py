"""UI-only acceptance server with synthetic, in-memory correction sessions.

Build frontend, then run: python -m backend.tests.correction_demo_server --port 8792
No production app, trajectory files, model clients, or workspace data are loaded.
Restarting discards all demo edits. Export routing is covered by backend tests;
the demo export only provides four downloadable fixture sheets for UI checks.
"""
from __future__ import annotations

import argparse
import asyncio
from copy import deepcopy
from datetime import datetime
from io import BytesIO
import json
from pathlib import Path
import uuid

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
import uvicorn


def create_demo_app() -> FastAPI:
    dist = Path(__file__).resolve().parents[2] / "frontend" / "dist"
    app = FastAPI(title="Correction UI demo — synthetic data only")
    sessions = {}
    details = {}
    exports = {}
    events = []
    faults = {"next_detail_error": False, "next_save_error": False, "detail_delay_ms": 0}
    screenshot = '''<svg xmlns="http://www.w3.org/2000/svg" width="360" height="720" viewBox="0 0 360 720">
      <rect width="360" height="720" fill="#f1f5f9"/><rect x="16" y="52" width="328" height="54" rx="14" fill="white"/>
      <text x="28" y="32" font-family="sans-serif" font-size="14" fill="#64748b">9:41                 DEMO</text>
      <text x="34" y="85" font-family="sans-serif" font-size="18" fill="#334155">Search movies</text>
      <rect x="16" y="134" width="328" height="180" rx="14" fill="#0f766e"/>
      <text x="36" y="230" font-family="sans-serif" font-size="24" fill="white">Synthetic screenshot</text>
      <rect x="16" y="344" width="328" height="78" rx="14" fill="white"/><rect x="16" y="438" width="328" height="78" rx="14" fill="white"/>
      <text x="34" y="390" font-family="sans-serif" font-size="20" fill="#334155">Movie A</text>
      <text x="34" y="484" font-family="sans-serif" font-size="20" fill="#334155">Movie B</text>
      <text x="52" y="672" font-family="sans-serif" font-size="14" fill="#64748b">No real trajectory data is used</text></svg>'''

    def summary(group):
        value = {key: deepcopy(item) for key, item in group.items() if key != "rows"}
        value.update(active_row_count=sum(not row["deleted"] for row in group["rows"]),
                     edited_row_count=sum(row["edited"] for row in group["rows"]),
                     action_edit_count=sum(row["action_edited"] for row in group["rows"]))
        return value

    for batch, multi in [("demo-multi", True), ("demo-top1", False)]:
        sid = f"session-{batch}"
        selected = []
        groups = []
        for index, (case_id, trajectory_n) in enumerate(
            [("AT-YYSP-AQY-001", n) for n in range(1, 4 if multi else 2)] + [("AT-YYSP-TXSP-002", 1)]
        ):
            trajectory_id = f"{case_id}-{trajectory_n}"
            goal = "搜索一部电影并查看详情" if "AQY" in case_id else "搜索纪录片并加入收藏"
            actions = [{"action": "click", "coordinate": [400, 110]}, {"action": "type", "text": "测试电影"},
                       {"action": "swipe", "start_coordinate": [500, 800], "end_coordinate": [500, 300]},
                       {"action": "terminate", "status": "success"}]
            rows = []
            for step, action in enumerate(actions, 1):
                rows.append({"excel_row": index * 4 + step + 1, "step": step, "task": goal, "meta_task": trajectory_id,
                             "image": f"{trajectory_id}/step{step:03}.jpg", "image_url": "", "xml": "", "actions": json.dumps(action, ensure_ascii=False),
                             "action": action, "sop": f"原始 SOP {trajectory_n}-{step}", "summary": ["点击搜索框", "输入电影名称", "向下浏览结果", "完成任务"][step - 1],
                             "task_manual_result": "", "micro_manual": "", "macro_manual": "", "micro_pred": "", "macro_pred": "",
                             "Bad_Interval": "", "trajectory_quality_type": "", "actions_box": "", "deleted": False,
                             "edited": False, "action_edited": False, "sop_edited": False, "edit_status": ""})
            group = {"group_id": f"group_{index}", "task": goal, "meta_task": trajectory_id, "quality": "未知", "prefix": "",
                     "export": False, "row_count": 4, "rows": rows}
            details[(sid, group["group_id"])] = group
            groups.append(summary(group))
            selected.append({"task_id": case_id, "goal": goal, "trajectory_id": trajectory_id, "global_score": 4.8 - (trajectory_n - 1) * .7,
                             "passed_threshold": trajectory_n < 3, "trajectory_count": 10, "step_count": 4})
        selection = {"status": "ready", "tree_run_id": batch, "run_id": batch, "tree_completed_at": "2026-08-31 10:00", "quality_completed_at": "2026-08-31 10:20",
                     "total_task_count": 3, "reviewed_task_count": 2, "tasks": selected}
        sessions[sid] = {"session_id": sid, "tree_run_id": batch, "source_id": "demo-only", "source": None, "created_at": "2026-08-31 10:30", "updated_at": "2026-08-31 10:30",
                         "row_count": len(groups) * 4, "group_count": len(groups), "groups": groups, "selection": selection, "exports": []}

    @app.middleware("http")
    async def record(request, call_next):
        if request.url.path.startswith("/api/"):
            events.append({"method": request.method, "path": request.url.path})
        return await call_next(request)

    @app.get("/__demo__/events")
    async def event_log():
        return events

    @app.post("/__demo__/faults")
    async def configure_faults(value: dict):
        for key in faults:
            if key in value:
                faults[key] = value[key]
        return faults

    @app.get("/api/correction/batches")
    async def batches():
        return {"default_tree_run_id": "demo-multi", "batches": [
            {"tree_run_id": name, "tree_completed_at": "2026-08-31 10:00", "quality_completed_at": "2026-08-31 10:20", "total_task_count": 3,
             "reviewed_task_count": 2, "status": "ready", "is_default": name == "demo-multi"} for name in ("demo-multi", "demo-top1")]}

    @app.get("/api/correction/sessions")
    async def list_sessions():
        return {"sessions": list(sessions.values())}

    @app.get("/api/correction/sessions/{sid}")
    async def get_session(sid: str):
        if sid not in sessions:
            raise HTTPException(404)
        return {"session": sessions[sid]}

    @app.get("/api/correction/sessions/{sid}/tasks/{gid}")
    async def get_group(sid: str, gid: str):
        if faults["next_detail_error"]:
            faults["next_detail_error"] = False
            raise HTTPException(503, "模拟轨迹加载失败，可重试")
        result = deepcopy(details[(sid, gid)])
        await asyncio.sleep(min(5, max(0, int(faults["detail_delay_ms"])) / 1000))
        return {"group": {**result, **summary(result)}}

    @app.patch("/api/correction/sessions/{sid}/rows/{row_id}")
    async def patch_row(sid: str, row_id: int, patch: dict):
        if faults["next_save_error"]:
            faults["next_save_error"] = False
            raise HTTPException(503, "模拟保存失败，输入应保留")
        group = next(group for (session_id, _), group in details.items() if session_id == sid and any(row["excel_row"] == row_id for row in group["rows"]))
        row = next(row for row in group["rows"] if row["excel_row"] == row_id)
        for key in ("actions", "sop", "deleted"):
            if key in patch:
                row[key] = patch[key]
        row["action_edited"] |= "actions" in patch
        row["sop_edited"] |= "sop" in patch
        row["edited"] = row["action_edited"] or row["sop_edited"] or row["deleted"]
        row["action"] = json.loads(row["actions"])
        row["edit_status"] = "已修改" if row["edited"] else ""
        sessions[sid]["groups"] = [summary(details[(sid, item["group_id"])]) for item in sessions[sid]["groups"]]
        return {"group": summary(group), "row": row}

    @app.patch("/api/correction/sessions/{sid}/tasks/{gid}/export")
    async def export_flag(sid: str, gid: str, patch: dict):
        details[(sid, gid)]["export"] = bool(patch["export"])
        sessions[sid]["groups"] = [summary(details[(sid, item["group_id"])]) for item in sessions[sid]["groups"]]
        return {"group": summary(details[(sid, gid)])}

    @app.get("/api/correction/sessions/{sid}/assets/{path:path}")
    async def image(sid: str, path: str):
        return Response(screenshot, media_type="image/svg+xml")

    @app.post("/api/correction/sessions/{sid}/export")
    async def export(sid: str):
        book = Workbook()
        book.remove(book.active)
        names = ["SFT_人工精修", "RL_负向反思", "原生_完美通过", "原生_异常待处理"]
        for name in names:
            book.create_sheet(name).append(["task", "meta_task", "actions", "sop"])
        sheet = book[names[0]]
        for (session_id, _), group in details.items():
            if session_id == sid and group["export"]:
                for row in group["rows"]:
                    if not row["deleted"]:
                        sheet.append([row["task"], row["meta_task"], row["actions"], row["sop"]])
        output = BytesIO()
        book.save(output)
        eid = uuid.uuid4().hex
        filename = f"demo-{eid}.xlsx"
        exports[(sid, filename)] = output.getvalue()
        result = {"export_id": eid, "filename": filename, "created_at": datetime.now().isoformat(), "sheets": {name: book[name].max_row - 1 for name in names}}
        sessions[sid]["exports"].insert(0, result)
        return result

    @app.get("/api/correction/sessions/{sid}/exports/{filename}")
    async def download(sid: str, filename: str):
        data = exports.get((sid, filename))
        if data is None:
            raise HTTPException(404)
        return Response(data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    app.mount("/assets", StaticFiles(directory=dist / "assets"), name="assets")

    @app.get("/{path:path}")
    async def frontend(path: str):
        if path.startswith("api/"):
            raise HTTPException(404)
        return FileResponse(dist / "index.html")

    return app


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--port", type=int, default=8792)
    args = parser.parse_args()
    uvicorn.run(create_demo_app(), host="127.0.0.1", port=args.port)
