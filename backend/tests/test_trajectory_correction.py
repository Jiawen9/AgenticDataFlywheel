from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from backend.trajectory_correction.assets import resolve_asset
from backend.trajectory_correction.exporter import export_session_workbook
from backend.trajectory_correction import quality_selection
from backend.trajectory_correction import workbook as correction_workbook
from backend.trajectory_correction.workbook import load_snapshot


def write_correction_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "task",
        "meta_task",
        "image",
        "actions",
        "sop",
        "task_manual_result",
        "micro_manual",
        "macro_manual",
        "micro_pred",
        "macro_pred",
        "Bad_Interval",
    ])
    sheet.append(["TASK-A", "打开应用", r"TASK-A\TASK-A-1\step001.jpg", json.dumps({"action": "click", "coordinate": [10, 20]}), "原 SOP", 1, "1", "1", "", "", "Normal"])
    sheet.append(["TASK-A", "打开应用", r"TASK-A\TASK-A-1\step002.jpg", json.dumps({"action": "wait"}), "", 1, "1", "1", "", "", "Normal"])
    sheet.append(["TASK-A", "打开应用", r"TASK-A\TASK-A-1\step003.jpg", json.dumps({"action": "terminate", "status": "success"}), "", 1, "1", "1", "", "", "Normal"])
    workbook.save(path)


def write_annotated_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["文件夹名", "image", "xml", "action", "summary", "actions_box"])
    sheet.append([
        "TASK-A",
        r"collection\TASK-A\step001_vla_input.jpg",
        "<root />",
        json.dumps({"action": "click", "coordinate": [10, 20]}),
        "点击入口",
        "click(bbox=<bbox>[1,2,3,4]</bbox>)",
    ])
    sheet.append([
        "TASK-B",
        r"collection\TASK-B\step001_vla_input.jpg",
        "<root />",
        json.dumps({"action": "wait"}),
        "等待加载",
        "wait()",
    ])
    # Deliberately make TASK-A non-consecutive: grouping must use 文件夹名.
    sheet.append([
        "TASK-A",
        r"collection\TASK-A\step002_vla_input.jpg",
        "<root />",
        json.dumps({"action": "type", "text": "hello"}),
        "输入内容",
        "type(text=hello)",
    ])
    workbook.save(path)


def write_annotated_assets(root: Path) -> None:
    for task in ("TASK-A", "TASK-B"):
        task_dir = root / "collection" / task
        task_dir.mkdir(parents=True)
        (task_dir / "step001_vla_input.jpg").write_bytes(b"test-image")
    (root / "collection" / "TASK-A" / "step002_vla_input.jpg").write_bytes(b"test-image")
    (root / "collection" / "TASK-A" / "_trajectory_for_evaluate.json").write_text(
        json.dumps({"task": "Open the TASK-A application"}),
        encoding="utf-8",
    )


def write_ranked_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["文件夹名", "image", "xml", "action", "summary", "actions_box"])
    for index, trajectory_id in enumerate(("TASK-A-1", "TASK-A-2", "TASK-A-3"), 1):
        sheet.append([
            trajectory_id,
            f"TASK-A\\{trajectory_id}\\step001_vla_input.jpg",
            "<root />",
            json.dumps({"action": "click", "coordinate": [index, index]}),
            f"步骤 {index}",
            "click(bbox=<bbox>[1,2,3,4]</bbox>)",
        ])
    workbook.save(path)


def write_ranked_assets(root: Path) -> None:
    for trajectory_id in ("TASK-A-1", "TASK-A-2", "TASK-A-3"):
        trajectory = root / "TASK-A" / trajectory_id
        trajectory.mkdir(parents=True)
        (trajectory / "step001_vla_input.jpg").write_bytes(b"test-image")
        (trajectory / "_trajectory_for_evaluate.json").write_text(
            json.dumps({"task": "Open TASK-A"}),
            encoding="utf-8",
        )


def write_quality_run(root: Path, run_id: str, updated_at: str, source_hash: str, *, valid: bool = True) -> None:
    run = root / run_id
    run.mkdir(parents=True)
    (run / "manifest.json").write_text(
        json.dumps({
            "run_id": run_id,
            "updated_at": updated_at,
            "tasks": [{"task_id": "TASK-A", "goal": "Open TASK-A"}],
        }),
        encoding="utf-8",
    )
    evaluations = {
        "TASK-A-1": {"trajectory_id": "TASK-A-1", "global_score": 4.0, "passed_threshold": False},
        "TASK-A-2": {"trajectory_id": "TASK-A-2", "global_score": 4.0, "passed_threshold": True},
        "TASK-A-3": {"trajectory_id": "TASK-A-3", "global_score": 3.5, "passed_threshold": True},
    } if valid else {}
    (run / "TASK-A.json").write_text(
        json.dumps({"task_id": "TASK-A", "trajectory_count": 3, "evaluations": evaluations}),
        encoding="utf-8",
    )
    tree_dir = root / "_tree_runs" / run_id
    tree_dir.mkdir(parents=True, exist_ok=True)
    (tree_dir / "manifest.json").write_text(
        json.dumps({"source_xlsx": {"sha256": source_hash}}),
        encoding="utf-8",
    )


def write_tree_batch(
    tree_root: Path,
    run_id: str,
    completed_at: str,
    source_hash: str,
    task_ids: tuple[str, ...] = ("TASK-A", "TASK-B"),
    trajectory_count: int = 2,
) -> None:
    run = tree_root / run_id
    run.mkdir(parents=True)
    tasks = []
    for task_id in task_ids:
        (run / f"{task_id}.json").write_text(
            json.dumps({"task_id": task_id}), encoding="utf-8"
        )
        tasks.append(
            {
                "task_id": task_id,
                "goal": f"Goal {task_id}",
                "tree_file": f"{task_id}.json",
                "trajectory_count": trajectory_count,
            }
        )
    (run / "manifest.json").write_text(
        json.dumps(
            {
                "run_id": run_id,
                "completed_at": completed_at,
                "source_xlsx": {"sha256": source_hash},
                "tasks": tasks,
            }
        ),
        encoding="utf-8",
    )


def write_multi_task_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["文件夹名", "image", "xml", "action", "summary", "actions_box"])
    for task_id in ("TASK-A", "TASK-B"):
        for trajectory_id in (f"{task_id}-1", f"{task_id}-2"):
            sheet.append(
                [
                    trajectory_id,
                    f"{task_id}\\{trajectory_id}\\step001_vla_input.jpg",
                    "<root />",
                    json.dumps({"action": "click", "coordinate": [1, 1]}),
                    f"{task_id} step",
                    "click(bbox=<bbox>[1,2,3,4]</bbox>)",
                ]
            )
    workbook.save(path)


def write_multi_task_quality(
    quality_root: Path,
    run_id: str,
    updated_at: str,
    reviewed: dict[str, dict[str, float]],
) -> None:
    run = quality_root / run_id
    run.mkdir(parents=True)
    summaries = []
    for task_id, scores in reviewed.items():
        evaluations = {
            trajectory_id: {
                "trajectory_id": trajectory_id,
                "global_score": score,
                "passed_threshold": False,
            }
            for trajectory_id, score in scores.items()
        }
        (run / f"{task_id}.json").write_text(
            json.dumps(
                {
                    "run_id": run_id,
                    "task_id": task_id,
                    "trajectory_count": len(evaluations),
                    "evaluations": evaluations,
                }
            ),
            encoding="utf-8",
        )
        summaries.append(
            {
                "task_id": task_id,
                "completed_at": updated_at,
                "trajectory_count": len(evaluations),
            }
        )
    (run / "manifest.json").write_text(
        json.dumps({"run_id": run_id, "updated_at": updated_at, "tasks": summaries}),
        encoding="utf-8",
    )


class CorrectionWorkbookTests(unittest.TestCase):
    def test_snapshot_cache_reuses_snapshot_and_invalidates_when_workbook_changes(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "input.xlsx"
            write_correction_workbook(source)
            correction_workbook.clear_snapshot_cache()
            with patch.object(
                correction_workbook,
                "_load_snapshot_uncached",
                wraps=correction_workbook._load_snapshot_uncached,
            ) as loader:
                first = load_snapshot(source)
                second = load_snapshot(source)
                self.assertEqual(loader.call_count, 1)

                second["groups"][0]["rows"][0]["actions"] = "changed only in caller"
                third = load_snapshot(source)
                self.assertEqual(loader.call_count, 1)
                self.assertNotEqual(third["groups"][0]["rows"][0]["actions"], "changed only in caller")
                self.assertEqual(first["row_count"], third["row_count"])

                workbook = load_workbook(source)
                workbook.active["E2"] = "changed in source workbook"
                workbook.save(source)
                workbook.close()

                changed = load_snapshot(source)
                self.assertEqual(loader.call_count, 2)
                self.assertEqual(changed["groups"][0]["rows"][0]["sop"], "changed in source workbook")

            correction_workbook.clear_snapshot_cache()

    def test_groups_follow_human_script_quality_rules(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "input.xlsx"
            write_correction_workbook(path)
            snapshot = load_snapshot(path)

            self.assertEqual(len(snapshot["groups"]), 1)
            group = snapshot["groups"][0]
            self.assertEqual(group["quality"], "完成且过程正常")
            self.assertTrue(group["export"])
            self.assertEqual([row["step"] for row in group["rows"]], [1, 2, 3])

    def test_export_routes_first_two_action_edits_to_sft_and_rl(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "input.xlsx"
            write_correction_workbook(source)
            snapshot = load_snapshot(source)
            session = {
                "session_id": "session",
                "group_exports": {"group_0": True},
                "row_edits": {
                    "3": {
                        "actions": json.dumps({"action": "click", "coordinate": [100, 100]}),
                        "original_actions": snapshot["groups"][0]["rows"][1]["actions"],
                    },
                    "4": {
                        "actions": json.dumps({"action": "click", "coordinate": [200, 200]}),
                        "original_actions": snapshot["groups"][0]["rows"][2]["actions"],
                    },
                },
            }
            result = export_session_workbook(
                workbook_path=source,
                snapshot=snapshot,
                session=session,
                output_dir=root / "exports",
                export_id="abcdef1234567890",
            )

            self.assertEqual(result["sheets"]["SFT_人工精修"], 2)
            self.assertEqual(result["sheets"]["RL_负向反思"], 3)
            exported = load_workbook(root / "exports" / result["filename"], read_only=True, data_only=True)
            try:
                self.assertEqual(exported["SFT_人工精修"]["D3"].value, session["row_edits"]["3"]["actions"])
                self.assertEqual(exported["RL_负向反思"]["D3"].value, session["row_edits"]["3"]["original_actions"])
                self.assertEqual(exported["RL_负向反思"]["L4"].value, 1)
            finally:
                exported.close()

    def test_annotated_workbook_groups_by_folder_and_reads_rollout_goal(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "annotated.xlsx"
            assets = root / "rollout_trajectories"
            write_annotated_workbook(source)
            write_annotated_assets(assets)

            snapshot = load_snapshot(
                source,
                asset_root=assets,
                source_kind="annotated_workbook",
            )

            self.assertEqual(snapshot["headers"], ["文件夹名", "image", "xml", "action", "summary", "actions_box"])
            self.assertEqual(snapshot["row_count"], 3)
            self.assertEqual([group["meta_task"] for group in snapshot["groups"]], ["TASK-A", "TASK-B"])
            self.assertEqual(snapshot["groups"][0]["task"], "Open the TASK-A application")
            self.assertEqual(snapshot["groups"][1]["task"], "TASK-B")
            self.assertEqual(snapshot["groups"][0]["quality"], "未知")
            self.assertFalse(snapshot["groups"][0]["export"])
            self.assertEqual([row["excel_row"] for row in snapshot["groups"][0]["rows"]], [2, 4])
            self.assertEqual(snapshot["groups"][0]["rows"][0]["action"]["action"], "click")
            self.assertEqual(snapshot["groups"][0]["rows"][0]["summary"], "点击入口")
            self.assertEqual(
                snapshot["groups"][0]["rows"][0]["actions_box"],
                "click(bbox=<bbox>[1,2,3,4]</bbox>)",
            )
            self.assertEqual(
                resolve_asset(assets, r"collection\TASK-A\step001_vla_input.jpg"),
                assets / "collection" / "TASK-A" / "step001_vla_input.jpg",
            )

    def test_annotated_export_preserves_six_source_columns_and_action_name(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "annotated.xlsx"
            assets = root / "rollout_trajectories"
            write_annotated_workbook(source)
            write_annotated_assets(assets)
            snapshot = load_snapshot(source, asset_root=assets)
            new_action = json.dumps({"action": "type", "text": "updated"})
            session = {
                "session_id": "session",
                "group_exports": {"group_0": True, "group_1": False},
                "row_edits": {
                    "2": {
                        "actions": new_action,
                        "original_actions": snapshot["groups"][0]["rows"][0]["actions"],
                    },
                },
            }

            result = export_session_workbook(
                workbook_path=source,
                snapshot=snapshot,
                session=session,
                output_dir=root / "exports",
                export_id="abcdef1234567890",
            )

            exported = load_workbook(root / "exports" / result["filename"], read_only=True, data_only=True)
            try:
                sheet = exported["SFT_人工精修"]
                self.assertEqual(
                    [sheet.cell(1, column).value for column in range(1, 8)],
                    ["文件夹名", "image", "xml", "action", "summary", "actions_box", "is_refined"],
                )
                self.assertEqual(sheet["D2"].value, new_action)
                self.assertEqual(sheet["G2"].value, 1)
            finally:
                exported.close()

    def test_annotated_asset_traversal_is_rejected(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            image = root / "inside.jpg"
            image.write_bytes(b"test-image")
            with self.assertRaises(ValueError):
                resolve_asset(root, r"..\outside.jpg")

    def test_quality_selection_uses_latest_run_tie_order_and_ignores_pass_flag(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "annotated.xlsx"
            assets = root / "rollout_trajectories"
            quality_root = root / "quality_results"
            write_ranked_workbook(source)
            write_ranked_assets(assets)
            source_hash = quality_selection._source_sha256(source)
            write_quality_run(quality_root, "old", "2026-08-27T10:00:00+08:00", source_hash)
            write_quality_run(quality_root, "new", "2026-08-27T12:00:00+08:00", source_hash)
            snapshot = load_snapshot(source, asset_root=assets)

            runs = quality_selection._completed_quality_runs(quality_root)
            self.assertEqual([run["run_id"] for run in runs], ["new", "old"])
            with patch.object(quality_selection, "TREE_RUNS_DIR", quality_root / "_tree_runs"):
                selection = quality_selection._selection_for_run(
                    runs[0], snapshot, source_path=source
                )
            self.assertEqual(selection["run_id"], "new")
            self.assertEqual(selection["selected_trajectories"], {"TASK-A": "TASK-A-1"})
            self.assertFalse(selection["tasks"][0]["passed_threshold"])
            self.assertEqual(selection["tasks"][0]["global_score"], 4.0)

            filtered = quality_selection.filter_snapshot(snapshot, selection)
            self.assertEqual(len(filtered["groups"]), 1)
            self.assertEqual(filtered["groups"][0]["meta_task"], "TASK-A-1")
            self.assertEqual(filtered["row_count"], 1)

            result = export_session_workbook(
                workbook_path=source,
                snapshot=filtered,
                session={"group_exports": {"group_0": True}, "row_edits": {}},
                output_dir=root / "exports",
                export_id="top1only123456789",
            )
            self.assertEqual(result["summary"], {"rows": 1, "groups": 1})
            self.assertEqual(result["sheets"], {"原生_异常待处理": 1})
            exported = load_workbook(root / "exports" / result["filename"], read_only=True, data_only=True)
            try:
                sheet = exported["原生_异常待处理"]
                self.assertEqual(sheet.max_row, 2)
                self.assertEqual(
                    [sheet.cell(1, column).value for column in range(1, 8)],
                    ["文件夹名", "image", "xml", "action", "summary", "actions_box", "is_refined"],
                )
                self.assertEqual(sheet["A2"].value, "TASK-A-1")
            finally:
                exported.close()

    def test_quality_batches_use_tree_run_ids_and_allow_partial_review(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "annotated.xlsx"
            tree_root = root / "tree_runs"
            quality_root = root / "quality_results"
            write_multi_task_workbook(source)
            source_hash = quality_selection._source_sha256(source)
            write_tree_batch(tree_root, "first", "2026-08-27T10:00:00+08:00", source_hash)
            write_tree_batch(tree_root, "second", "2026-08-27T12:00:00+08:00", source_hash)
            write_tree_batch(tree_root, "third", "2026-08-27T14:00:00+08:00", source_hash)
            write_multi_task_quality(
                quality_root,
                "first",
                "2026-08-27T11:00:00+08:00",
                {"TASK-A": {"TASK-A-1": 4.0, "TASK-A-2": 3.0}},
            )
            write_multi_task_quality(
                quality_root,
                "second",
                "2026-08-27T13:00:00+08:00",
                {
                    "TASK-A": {"TASK-A-1": 4.0, "TASK-A-2": 4.0},
                    "TASK-B": {"TASK-B-1": 2.0, "TASK-B-2": 3.0},
                },
            )
            batches = quality_selection.correction_batches(quality_root, tree_root)

            self.assertEqual(
                [item["tree_run_id"] for item in batches["batches"]],
                ["second", "first"],
            )
            self.assertEqual(batches["default_tree_run_id"], "second")
            self.assertEqual(batches["batches"][1]["reviewed_task_count"], 1)
            self.assertEqual(batches["batches"][1]["total_task_count"], 2)

            snapshot = load_snapshot(source, asset_root=root / "assets")
            runs = quality_selection._completed_quality_runs(quality_root, tree_root)
            with patch.object(quality_selection, "TREE_RUNS_DIR", tree_root):
                selection = quality_selection._selection_for_run(
                    runs[1], snapshot, source_path=source
                )
            self.assertEqual(selection["tree_run_id"], "first")
            self.assertEqual(selection["reviewed_task_count"], 1)
            self.assertEqual(selection["selected_trajectories"], {"TASK-A": "TASK-A-1"})

    def test_selected_old_batch_does_not_depend_on_latest_batch(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "annotated.xlsx"
            tree_root = root / "tree_runs"
            quality_root = root / "quality_results"
            write_ranked_workbook(source)
            write_ranked_assets(root / "assets")
            source_hash = quality_selection._source_sha256(source)
            write_tree_batch(tree_root, "old", "2026-08-27T10:00:00+08:00", source_hash, ("TASK-A",), 3)
            write_tree_batch(tree_root, "new", "2026-08-27T12:00:00+08:00", source_hash, ("TASK-A",), 3)
            write_quality_run(quality_root, "old", "2026-08-27T11:00:00+08:00", source_hash)
            write_quality_run(quality_root, "new", "2026-08-27T13:00:00+08:00", source_hash)
            snapshot = load_snapshot(source, asset_root=root / "assets")
            selection = quality_selection.top1_recommendation(
                "old",
                root=quality_root,
                tree_root=tree_root,
                source_path=source,
                asset_root=root / "assets",
            )
            self.assertEqual(selection["tree_run_id"], "old")
            self.assertEqual(selection["tasks"][0]["trajectory_id"], "TASK-A-1")

    def test_quality_selection_rejects_trajectory_version_mismatch(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "annotated.xlsx"
            assets = root / "rollout_trajectories"
            quality_root = root / "quality_results"
            write_ranked_workbook(source)
            write_ranked_assets(assets)
            write_quality_run(quality_root, "new", "2026-08-27T12:00:00+08:00", "wrong-hash")
            snapshot = load_snapshot(source, asset_root=assets)
            runs = quality_selection._completed_quality_runs(quality_root)

            with patch.object(quality_selection, "TREE_RUNS_DIR", quality_root / "_tree_runs"):
                with self.assertRaises(quality_selection.QualitySourceMismatch):
                    quality_selection._selection_for_run(
                        runs[0], snapshot, source_path=source
                    )

    def test_fixed_source_rejects_legacy_upload_source_ids(self):
        from backend.trajectory_correction.assets import source_from_id

        with self.assertRaises(ValueError):
            source_from_id("backend_workspace/trajectory_correction/inputs/old.xlsx")

    def test_quality_selection_rejects_incomplete_run(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            write_quality_run(root, "incomplete", "2026-08-27T12:00:00+08:00", "unused", valid=False)
            self.assertEqual(quality_selection._completed_quality_runs(root), [])

    def test_recommendation_blocks_when_no_completed_quality_run_exists(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            with patch.object(quality_selection, "QUALITY_RESULTS_DIR", Path(temp_dir) / "missing"):
                recommendation = quality_selection.top1_recommendation()
            self.assertEqual(recommendation["status"], "blocked")
            self.assertEqual(recommendation["tasks"], [])


if __name__ == "__main__":
    unittest.main()
