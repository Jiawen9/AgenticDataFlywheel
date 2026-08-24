from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from backend.bounding_box.build_annotations import resolve_action_box
from backend.export_vla_trajectories import collect_rows, write_xlsx
from backend.trajectories_preprocessing import (
    annotate_trajectory_workbook,
    format_actions_box,
    swipe_direction,
)


UI_XML = """<?xml version="1.0" encoding="UTF-8"?>
<hierarchy>
  <node bounds="[0,0][200,300]" enabled="true" visible-to-user="true"
        clickable="false" scrollable="true" class="android.widget.FrameLayout">
    <node bounds="[20,20][120,100]" enabled="true" visible-to-user="true"
          clickable="true" scrollable="false" class="android.widget.Button" />
  </node>
</hierarchy>
"""


class FakeReviewResult:
    def __init__(self, decision, bbox, confidence=0.9, reason="test"):
        self.decision = decision
        self.bbox = tuple(bbox)
        self.confidence = confidence
        self.reason = reason

    def to_dict(self):
        return {
            "decision": self.decision,
            "bbox": list(self.bbox),
            "confidence": self.confidence,
            "reason": self.reason,
            "raw_response": "{}",
            "cached": False,
        }


class AcceptingReviewer:
    def review(self, **kwargs):
        return FakeReviewResult("accept", kwargs["candidate_bbox"])


class ReplacingThenAcceptingReviewer:
    def __init__(self):
        self.calls = 0

    def review(self, **kwargs):
        self.calls += 1
        if self.calls == 1:
            return FakeReviewResult("replace", (0, 0, 10, 10))
        return FakeReviewResult("accept", kwargs["candidate_bbox"])


class FailingReviewer:
    def review(self, **kwargs):
        raise RuntimeError("review failed")


def create_step(run_dir: Path, step: int, action: dict, summary: str = "test action") -> list[str]:
    prefix = f"step{step:03d}"
    image_path = run_dir / f"{prefix}_vla_input.jpg"
    stability_path = run_dir / f"{prefix}_vla_input_stability.jpg"
    xml_path = run_dir / f"{prefix}_vla_input_ui.xml"
    response_path = run_dir / f"{prefix}_vla_model_response.json"
    Image.new("RGB", (200, 300), "white").save(image_path)
    Image.new("RGB", (200, 300), "white").save(stability_path)
    xml_path.write_text(UI_XML, encoding="utf-8")
    response_path.write_text(
        json.dumps(
            {
                "content": (
                    "<thought>test</thought><tool_call>"
                    + json.dumps(action)
                    + f"</tool_call><summary>{summary}</summary>"
                )
            }
        ),
        encoding="utf-8",
    )
    return [run_dir.name, str(image_path.resolve()), str(xml_path.resolve()), json.dumps(action), summary]


class ExportDiscoveryTests(unittest.TestCase):
    def test_collect_rows_recurses_and_ignores_prefetch_staging(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            run_dir = root / "group" / "run-1"
            run_dir.mkdir(parents=True)
            create_step(run_dir, 1, {"action": "click", "coordinate": [50, 50]})
            (run_dir / "_trajectory_for_evaluate.json").write_text(
                json.dumps({"actions_flat": []}), encoding="utf-8"
            )
            staging = run_dir / "_prefetch_staging" / "candidate"
            staging.mkdir(parents=True)
            create_step(staging, 2, {"action": "click", "coordinate": [50, 50]})
            rejected = run_dir / "_stability_rejected" / "candidate"
            rejected.mkdir(parents=True)
            create_step(rejected, 3, {"action": "click", "coordinate": [50, 50]})

            rows, warnings = collect_rows(root)

            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0][0], "run-1")
            self.assertEqual(rows[0][1], str(Path("group") / "run-1" / "step001_vla_input.jpg"))
            self.assertEqual(rows[0][2], str(Path("group") / "run-1" / "step001_vla_input_ui.xml"))
            self.assertEqual(warnings, [])


class ActionFormattingTests(unittest.TestCase):
    def test_action_box_formats(self):
        bbox = (43, 426, 265, 570)
        self.assertEqual(
            format_actions_box({"action": "click"}, bbox),
            "click(bbox=<bbox>[43,426,265,570]</bbox>)",
        )
        self.assertEqual(
            format_actions_box({"action": "long_press"}, bbox),
            "long_press(bbox=<bbox>[43,426,265,570]</bbox>)",
        )

    def test_all_swipe_directions(self):
        cases = {
            "left": ([10, 5], [0, 5]),
            "right": ([0, 5], [10, 5]),
            "up": ([5, 10], [5, 0]),
            "down": ([5, 0], [5, 10]),
        }
        for expected, (start, end) in cases.items():
            action = {"action": "swipe", "start_coordinate": start, "end_coordinate": end}
            self.assertEqual(swipe_direction(action), expected)
        action = {"action": "swipe", "start_coordinate": [10, 5], "end_coordinate": [0, 5]}
        self.assertEqual(
            format_actions_box(action, (1, 2, 30, 40)),
            "swipe_screen(bbox=<bbox>[1,2,30,40]</bbox>, direction=left)",
        )


class BoxResolutionTests(unittest.TestCase):
    def test_qwen_accepts_rule_box(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            image_path = Path(temp_dir) / "image.jpg"
            Image.new("RGB", (200, 300), "white").save(image_path)
            resolution = resolve_action_box(
                image_path=image_path,
                xml_text=UI_XML,
                action={"action": "click", "coordinate": [50, 50]},
                action_summary="click button",
                reviewer=AcceptingReviewer(),
            )
            self.assertTrue(resolution.verified)
            self.assertEqual(resolution.result.source, "qwen_verified")
            x1, y1, x2, y2 = resolution.result.bbox
            self.assertTrue(x1 <= 50 <= x2 and y1 <= 50 <= y2)

    def test_replacement_excluding_click_is_corrected_and_reviewed_again(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            image_path = Path(temp_dir) / "image.jpg"
            Image.new("RGB", (200, 300), "white").save(image_path)
            reviewer = ReplacingThenAcceptingReviewer()
            resolution = resolve_action_box(
                image_path=image_path,
                xml_text=UI_XML,
                action={"action": "long_press", "coordinate": [100, 150]},
                action_summary="long press item",
                reviewer=reviewer,
            )
            self.assertEqual(reviewer.calls, 2)
            x1, y1, x2, y2 = resolution.result.bbox
            self.assertTrue(x1 <= 100 <= x2 and y1 <= 150 <= y2)


class WorkbookAnnotationTests(unittest.TestCase):
    def _create_workbook_fixture(self, root: Path) -> tuple[Path, Path]:
        trajectory_root = root / "rollout_trajectories"
        run_dir = trajectory_root / "group" / "run-1"
        run_dir.mkdir(parents=True)
        actions = [
            {"action": "click", "coordinate": [50, 50]},
            {"action": "swipe", "start_coordinate": [100, 200], "end_coordinate": [100, 50]},
            {"action": "long_press", "coordinate": [60, 60]},
            {"action": "wait"},
        ]
        for index, action in enumerate(actions, 1):
            create_step(run_dir, index, action)
        (run_dir / "_trajectory_for_evaluate.json").write_text(
            json.dumps(
                {
                    "actions_flat": [
                        {"global_step": index, "action": action}
                        for index, action in enumerate(actions, 1)
                    ]
                }
            ),
            encoding="utf-8",
        )
        rows, warnings = collect_rows(trajectory_root)
        self.assertEqual(warnings, [])
        source_path = root / "source.xlsx"
        write_xlsx(rows, source_path)
        return source_path, trajectory_root

    def test_workbook_adds_only_target_action_boxes(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source_path, trajectory_root = self._create_workbook_fixture(root)
            output_path = root / "annotated.xlsx"

            counts = annotate_trajectory_workbook(
                source_path,
                output_path,
                reviewer=AcceptingReviewer(),
                trajectory_root=trajectory_root,
            )

            self.assertEqual(counts, {"rows": 4, "annotated": 3, "blank": 1})
            workbook = load_workbook(output_path)
            sheet = workbook.active
            self.assertEqual(sheet.cell(1, 6).value, "actions_box")
            self.assertEqual(
                sheet.cell(2, 2).value,
                str(Path("group") / "run-1" / "step001_vla_input.jpg"),
            )
            self.assertEqual(
                sheet.cell(2, 3).value,
                str(Path("group") / "run-1" / "step001_vla_input_ui.xml"),
            )
            self.assertTrue(sheet.cell(2, 6).value.startswith("click(bbox=<bbox>"))
            self.assertIn("direction=up", sheet.cell(3, 6).value)
            self.assertTrue(sheet.cell(4, 6).value.startswith("long_press(bbox=<bbox>"))
            self.assertIsNone(sheet.cell(5, 6).value)
            self.assertEqual(sheet.auto_filter.ref, "A1:F5")
            self.assertEqual(sheet.freeze_panes, "A2")

    def test_failure_does_not_publish_annotated_workbook(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source_path, trajectory_root = self._create_workbook_fixture(root)
            output_path = root / "annotated.xlsx"

            with self.assertRaisesRegex(RuntimeError, "run-1/step001"):
                annotate_trajectory_workbook(
                    source_path,
                    output_path,
                    reviewer=FailingReviewer(),
                    trajectory_root=trajectory_root,
                )
            self.assertFalse(output_path.exists())


if __name__ == "__main__":
    unittest.main()
