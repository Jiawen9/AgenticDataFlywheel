from __future__ import annotations

import copy
import io
import json
import re
import tempfile
import unittest
import uuid
from concurrent.futures import Future, ThreadPoolExecutor
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.task_generation.constants import KNOWLEDGE_BASE_FILES
from backend.task_generation.jobs import TaskGenerationJobManager
from backend.task_generation.knowledge_base import merged_nodes, replace_knowledge_base, scene_tree_text, snapshot_knowledge_base, tree_payload
from backend.task_generation.router import router
from backend.task_generation.service import run_augmentation, run_initial_generation
from backend.task_generation.tree_store import (
    META_SHEET, VersionConflict, current_root, flatten, import_scene_workbook,
    read_tree, save_tree, validate_tree,
)
from backend.tests.test_task_generation import _write_knowledge_base


class SimulatedModel:
    """Deterministic local model; API tests never contact a model endpoint."""
    def __init__(self):
        self.prompts = []

    def complete(self, prompt, **kwargs):
        self.prompts.append(prompt)
        if '"dependency_relationships"' in prompt:
            return '{"dependency_relationships":"zero"}'
        if "# Scene Mapping Table" in prompt:
            match = re.search(r"(?m)^(.+) > (.+) > (.+) \| 涵盖App", prompt)
            return json.dumps(dict(zip(("scene", "capability", "sub_capability"), match.groups())), ensure_ascii=False)
        count = int(re.search(r"生成 (\d+) 条", prompt).group(1))
        return json.dumps([{"task": f"模拟任务 {index + 1}"} for index in range(count)], ensure_ascii=False)


class ImmediateExecutor:
    def submit(self, fn, *args, **kwargs):
        future = Future()
        future.set_result(fn(*args, **kwargs))
        return future


class DeferredExecutor:
    def __init__(self):
        self.pending = []

    def submit(self, fn, *args, **kwargs):
        self.pending.append((fn, args, kwargs))
        return Future()

    def run_all(self):
        for fn, args, kwargs in self.pending:
            fn(*args, **kwargs)
        self.pending.clear()


class EditableTreeTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp.cleanup)
        self.base = Path(self.temp.name)
        self.kb = self.base / "kb"
        _write_knowledge_base(self.kb)

    def test_import_merges_apps_retains_distinct_app_configs_and_stable_ids(self):
        source = self.kb / KNOWLEDGE_BASE_FILES["scene_tree"]
        frame = pd.read_excel(source)
        first = frame.iloc[0].to_dict()
        pd.DataFrame([{**first, "target_app": "AppA", "reference_example": "A示例"},
                      {**first, "target_app": "AppB", "reference_example": "B示例", "use_resource_prior": False}]).to_excel(source, index=False)
        original = source.read_bytes()
        tree = tree_payload(self.kb)
        self.assertEqual(tree["leaf_count"], 1)
        self.assertEqual(tree["execution_unit_count"], 2)
        leaf = flatten(tree["scenes"])[0][0]
        self.assertEqual([config["reference_example"] for config in leaf["app_configs"]], ["A示例", "B示例"])
        self.assertEqual(leaf["id"], flatten(tree_payload(self.kb)["scenes"])[0][0]["id"])
        self.assertEqual(original, source.read_bytes())

    def test_rename_preserves_ids_and_control_associations_and_old_snapshot(self):
        before = tree_payload(self.kb)
        old_root = current_root(self.kb)
        snapshot = snapshot_knowledge_base(self.base / "snapshot", root=self.kb)
        scenes = copy.deepcopy(before["scenes"])
        scenes[0]["label"] = "新场景"
        scenes[0]["children"][0]["label"] = "新能力"
        leaf = scenes[0]["children"][0]["children"][0]
        leaf["label"] = "新任务类型"
        leaf["app_configs"][0]["reference_example"] = "新示例"
        after = save_tree(scenes, before["version"], root=self.kb)
        self.assertNotEqual(after["version"], before["version"])
        self.assertEqual(flatten(after["scenes"])[0][0]["id"], leaf["id"])
        self.assertTrue(flatten(after["scenes"])[0][0]["app_configs"][0]["control_prior_available"])
        current_nodes = merged_nodes(self.kb)
        self.assertEqual(current_nodes[0]["sub_capability_desc"], "操控描述")
        self.assertEqual(current_nodes[0]["reference_example"], "新示例")
        self.assertEqual(merged_nodes(Path(snapshot["directory"]))[0]["scene"], "场景A")
        self.assertEqual(read_tree(old_root)["scenes"][0]["label"], "场景A")
        self.assertIn("新场景 > 新能力 > 新任务类型", scene_tree_text(self.kb))

    def test_empty_branches_and_no_app_leaf_round_trip_through_excel(self):
        tree = tree_payload(self.kb)
        scenes = copy.deepcopy(tree["scenes"])
        leaf = flatten(scenes)[0][0]
        leaf["app_configs"] = []
        scenes.append({"id": str(uuid.uuid4()), "kind": "scene", "label": "空场景", "children": []})
        scenes[0]["children"].append({"id": str(uuid.uuid4()), "kind": "capability", "label": "空能力", "children": []})
        after = save_tree(scenes, tree["version"], root=self.kb)
        self.assertEqual(after["leaf_count"], 1)
        self.assertEqual(after["execution_unit_count"], 0)
        self.assertFalse(flatten(after["scenes"])[0][0]["generatable"])
        exported = current_root(self.kb) / KNOWLEDGE_BASE_FILES["scene_tree"]
        imported = import_scene_workbook(exported)
        self.assertEqual(imported, validate_tree(scenes))
        book = load_workbook(exported)
        self.assertEqual(book[META_SHEET].sheet_state, "hidden")
        book.close()
        self.assertEqual(len(pd.read_excel(current_root(self.kb) / KNOWLEDGE_BASE_FILES["control_prior"])), 1)

    def test_delete_all_then_upload_does_not_resurrect_tree(self):
        initial = tree_payload(self.kb)
        saved = save_tree([], initial["version"], root=self.kb)
        exported = current_root(self.kb) / KNOWLEDGE_BASE_FILES["scene_tree"]
        replace_knowledge_base("scene_tree", exported, root=self.kb, base_version=saved["version"])
        self.assertEqual(tree_payload(self.kb)["scenes"], [])

    def test_validation_failures_and_stale_save_do_not_publish(self):
        tree = tree_payload(self.kb)
        invalid = []
        blank = copy.deepcopy(tree["scenes"]); blank[0]["label"] = " "
        invalid.append(blank)
        duplicate = copy.deepcopy(tree["scenes"]); duplicate.append(copy.deepcopy(duplicate[0]))
        invalid.append(duplicate)
        wrong_depth = copy.deepcopy(tree["scenes"]); wrong_depth[0]["kind"] = "capability"
        invalid.append(wrong_depth)
        duplicate_apps = copy.deepcopy(tree["scenes"]); leaf = flatten(duplicate_apps)[0][0]; leaf["app_configs"].append(copy.deepcopy(leaf["app_configs"][0]))
        invalid.append(duplicate_apps)
        bad_id = copy.deepcopy(tree["scenes"]); bad_id[0]["id"] = "../../escape"
        invalid.append(bad_id)
        bad_bool = copy.deepcopy(tree["scenes"]); flatten(bad_bool)[0][0]["app_configs"][0]["use_resource_prior"] = "false"
        invalid.append(bad_bool)
        for scenes in invalid:
            with self.subTest(scenes=scenes), self.assertRaises(ValueError):
                save_tree(scenes, tree["version"], root=self.kb)
        self.assertEqual(read_tree(self.kb)["version"], tree["version"])
        save_tree(tree["scenes"], tree["version"], root=self.kb)
        with self.assertRaises(VersionConflict):
            save_tree(tree["scenes"], tree["version"], root=self.kb)

    def test_conflicting_excel_rows_are_rejected_without_overwrite(self):
        tree = tree_payload(self.kb)
        frame = pd.read_excel(self.kb / KNOWLEDGE_BASE_FILES["scene_tree"])
        bad = self.base / "bad.xlsx"
        pd.concat([frame, frame.assign(reference_example="不同示例")]).to_excel(bad, index=False)
        with self.assertRaisesRegex(ValueError, "配置冲突"):
            replace_knowledge_base("scene_tree", bad, root=self.kb)
        self.assertEqual(read_tree(self.kb)["version"], tree["version"])

    def test_failed_pointer_publication_keeps_old_bundle(self):
        tree = tree_payload(self.kb)
        old_root = current_root(self.kb)
        original_replace = Path.replace

        def fail_pointer(path, target):
            if Path(target).name == "current.json":
                raise OSError("simulated interrupted publication")
            return original_replace(path, target)

        with patch.object(Path, "replace", fail_pointer), self.assertRaises(OSError):
            save_tree([], tree["version"], root=self.kb)
        self.assertEqual(current_root(self.kb), old_root)
        self.assertEqual(tree_payload(self.kb)["leaf_count"], 1)

    def test_concurrent_saves_only_one_base_version_can_win(self):
        tree = tree_payload(self.kb)
        def save_once(label):
            scenes = copy.deepcopy(tree["scenes"]); scenes[0]["label"] = label
            try:
                save_tree(scenes, tree["version"], root=self.kb)
                return "saved"
            except VersionConflict:
                return "conflict"
        with ThreadPoolExecutor(max_workers=2) as executor:
            results = list(executor.map(save_once, ["甲场景", "乙场景"]))
        self.assertCountEqual(results, ["saved", "conflict"])

    def test_examples_are_not_written_as_excel_formulas(self):
        tree = tree_payload(self.kb)
        scenes = copy.deepcopy(tree["scenes"])
        flatten(scenes)[0][0]["app_configs"][0]["reference_example"] = '=HYPERLINK("https://example.invalid")'
        save_tree(scenes, tree["version"], root=self.kb)
        book = load_workbook(current_root(self.kb) / KNOWLEDGE_BASE_FILES["scene_tree"])
        self.assertEqual(book.worksheets[0]["F2"].data_type, "s")
        book.close()

    def test_new_prior_upload_changes_version_and_preserves_tree_identity(self):
        tree = tree_payload(self.kb)
        resource = self.base / "resource.xlsx"
        with pd.ExcelWriter(resource) as writer:
            pd.DataFrame([{"实体": "新资源"}]).to_excel(writer, sheet_name="AppB", index=False)
        replace_knowledge_base("resource_prior", resource, root=self.kb, base_version=tree["version"])
        after = tree_payload(self.kb)
        self.assertEqual(flatten(after["scenes"])[0][0]["id"], flatten(tree["scenes"])[0][0]["id"])
        self.assertEqual(flatten(after["scenes"])[0][0]["app_configs"][1]["resource_count"], 1)
        with self.assertRaises(VersionConflict):
            replace_knowledge_base("resource_prior", resource, root=self.kb, base_version=tree["version"])

    def make_manager(self, executor=None, model=None):
        model = model or SimulatedModel()
        manager = TaskGenerationJobManager(
            self.base / "jobs", self.base / "runs", self.base / "exports", self.kb, self.base / "logs",
            executor=executor or ImmediateExecutor(),
            initial_runner=lambda ids, count, **kwargs: run_initial_generation(ids, count, **kwargs, model=model),
            augmentation_runner=lambda path, count, **kwargs: run_augmentation(path, count, **kwargs, model=model),
        )
        self.addCleanup(manager.shutdown)
        return manager

    def test_queued_job_keeps_snapshot_and_selected_app_config_after_edit(self):
        executor = DeferredExecutor()
        manager = self.make_manager(executor)
        tree = tree_payload(self.kb)
        leaf = flatten(tree["scenes"])[0][0]
        job = manager.submit_initial([{"node_id": leaf["id"], "apps": ["AppA", "AppB"]}], 2, version=tree["version"])
        self.assertEqual(job["total_items"], 2)
        self.assertEqual(job["expected_main_tasks"], 4)
        scenes = copy.deepcopy(tree["scenes"]); scenes[0]["label"] = "改名后"
        flatten(scenes)[0][0]["app_configs"] = [leaf["app_configs"][1]]
        after = save_tree(scenes, tree["version"], root=self.kb)
        executor.run_all()
        rows = manager.results(job["job_id"])
        self.assertEqual([row["app"] for row in rows], ["AppA", "AppA", "AppB", "AppB"])
        self.assertTrue(all(row["scene"] == "场景A" and row["source_node_id"] == leaf["id"] for row in rows))
        self.assertEqual(manager.get(job["job_id"])["completed_items"], 2)
        self.assertNotIn("execution_units", manager.list_jobs()[0])
        with self.assertRaises(VersionConflict):
            manager.submit_initial([{"node_id": leaf["id"], "apps": ["AppA"]}], 1, version=tree["version"])
        with self.assertRaises(ValueError):
            manager.submit_initial([{"node_id": leaf["id"], "apps": ["AppA"]}], 1, version=after["version"])
        newer = manager.submit_initial([{"node_id": leaf["id"], "apps": ["AppB"]}], 1, version=after["version"])
        executor.run_all()
        self.assertEqual(manager.results(newer["job_id"])[0]["scene"], "改名后")

    def test_api_edit_generate_review_export_and_augmentation(self):
        model = SimulatedModel()
        manager = self.make_manager(model=model)
        app = FastAPI(); app.include_router(router)
        with patch("backend.task_generation.router.manager", manager), TestClient(app) as client:
            tree = client.get("/api/task-generation/tree").json()
            scenes = copy.deepcopy(tree["scenes"]); scenes[0]["label"] = "网页新场景"
            saved = client.put("/api/task-generation/tree", json={"base_version": tree["version"], "scenes": scenes})
            self.assertEqual(saved.status_code, 200)
            self.assertEqual(client.put("/api/task-generation/tree", json={"base_version": tree["version"], "scenes": []}).status_code, 409)
            current = saved.json(); leaf = flatten(current["scenes"])[0][0]
            body = {"version": current["version"], "selections": [{"node_id": leaf["id"], "apps": ["AppA", "AppB"]}], "generate_n": 2}
            self.assertEqual(client.post("/api/task-generation/jobs", json={**body, "version": tree["version"]}).status_code, 409)
            created = client.post("/api/task-generation/jobs", json=body)
            self.assertEqual(created.status_code, 202)
            job_id = created.json()["job_id"]
            self.assertEqual(client.get(f"/api/task-generation/jobs/{job_id}").json()["status"], "succeeded")
            rows = client.get(f"/api/task-generation/jobs/{job_id}/results").json()["results"]
            self.assertEqual(len(rows), 4)
            result_id = rows[0]["result_id"]
            self.assertEqual(client.patch(f"/api/task-generation/jobs/{job_id}/results/{result_id}", json={"task": "人工修改"}).status_code, 200)
            exported = client.post(f"/api/task-generation/jobs/{job_id}/export").json()
            download = client.get(exported["download_url"])
            self.assertEqual(pd.read_excel(io.BytesIO(download.content)).iloc[0]["task"], "人工修改")
            tree_download = client.get("/api/task-generation/tree/export")
            self.assertEqual(tree_download.status_code, 200)
            self.assertEqual(pd.read_excel(io.BytesIO(tree_download.content)).iloc[0]["scene"], "网页新场景")
            seed = io.BytesIO()
            pd.DataFrame([{"任务": "失败任务", "涉及APP": "AppA", "任务结果": False}]).to_excel(seed, index=False)
            augmented = client.post("/api/task-generation/augmentation-jobs", data={"generate_n": 1}, files={"file": ("seeds.xlsx", seed.getvalue())})
            self.assertEqual(augmented.status_code, 202)
            aug_rows = manager.results(augmented.json()["job_id"])
            self.assertEqual(aug_rows[0]["scene"], "网页新场景")
            self.assertTrue(any("操控描述" in prompt for prompt in model.prompts))


if __name__ == "__main__":
    unittest.main()
