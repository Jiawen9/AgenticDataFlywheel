from __future__ import annotations

import json
import tempfile
import unittest
from functools import partial
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from PIL import Image

from backend.trajectories_tree.intermediate_state_classifier import IntermediateStateResult
from backend.tree_build_service import build_tree_run
from backend.quality_input_builder import build_quality_workbook


class FakeClassifier:
    model = "fake-model"

    def __init__(self, _model: str, _cache: Path) -> None:
        pass

    def classify(self, **_kwargs):
        return IntermediateStateResult(False, "none", 1.0, "normal", "{}", True, "页面保持稳定")


class FailingSecondTaskClassifier(FakeClassifier):
    def classify(self, **kwargs):
        if kwargs["trajectory"].startswith("TASK-B"):
            raise RuntimeError("synthetic classifier failure")
        return super().classify(**kwargs)


class FakeAlignmentReviewer:
    model = "fake-model"

    def __init__(self, _model: str, _cache: Path) -> None:
        pass

    def review(self, **_kwargs):
        raise AssertionError("stable one-step tasks must not request alignment review")


class FakeSummarizer:
    model = "fake-model"

    def summarize_trajectory(self, task, trajectory_id, steps):
        return f"{task}: {trajectory_id} 共 {len(steps)} 步"


def prepare_source(root: Path) -> tuple[Path, Path]:
    trajectory_root = root / "rollout"
    workbook_path = root / "annotated.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["文件夹名", "image", "xml", "action", "summary", "actions_box"])
    for task_id in ["TASK-A", "TASK-B"]:
        trajectory_id = f"{task_id}-1"
        directory = trajectory_root / task_id / trajectory_id
        directory.mkdir(parents=True)
        image_path = directory / "step001_vla_input.jpg"
        Image.new("RGB", (100, 200), "white").save(image_path)
        (directory / "step001_vla_input_ui.xml").write_text("<hierarchy />", encoding="utf-8")
        (directory / "turn001_orch_model_request.json").write_text(
            json.dumps(
                {
                    "messages": [
                        {
                            "role": "user",
                            "content": f"**原始目标**: {task_id} goal\n\nnext",
                        }
                    ]
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        relative_image = image_path.relative_to(trajectory_root)
        relative_xml = (directory / "step001_vla_input_ui.xml").relative_to(trajectory_root)
        sheet.append(
            [
                trajectory_id,
                str(relative_image),
                str(relative_xml),
                '{"action":"wait"}',
                f"{task_id} summary",
                "",
            ]
        )
    workbook.save(workbook_path)
    return trajectory_root, workbook_path


class TreeBuildServiceTests(unittest.TestCase):
    def test_batch_publishes_separate_task_trees_and_manifest(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            trajectory_root, workbook_path = prepare_source(root)
            runs_dir = root / "runs"
            updates = []
            with (
                patch("backend.tree_build_service.configure_reviewer_environment", return_value="fake-model"),
                patch("backend.tree_build_service.QwenIntermediateStateClassifier", FakeClassifier),
                patch("backend.tree_build_service.QwenStateAlignmentReviewer", FakeAlignmentReviewer),
            ):
                run_id, manifest = build_tree_run(
                    ["TASK-A", "TASK-B"],
                    job_id="job-success",
                    progress=updates.append,
                    xlsx_path=workbook_path,
                    trajectory_root=trajectory_root,
                    runs_dir=runs_dir,
                    env_path=root / ".env",
                    classification_cache=root / "classification.json",
                    alignment_cache=root / "alignment.json",
                    quality_builder=partial(build_quality_workbook, summarizer=FakeSummarizer()),
                )

            published = runs_dir / run_id
            self.assertTrue((published / "manifest.json").is_file())
            self.assertTrue((published / "TASK-A.json").is_file())
            self.assertTrue((published / "TASK-B.json").is_file())
            self.assertTrue((published / "rubric_trajectories.xlsx").is_file())
            self.assertEqual(manifest["task_count"], 2)
            self.assertEqual(manifest["quality_input_file"], "rubric_trajectories.xlsx")
            workbook = load_workbook(published / "rubric_trajectories.xlsx", read_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["Tasks", "Trajectories", "Steps"])
                self.assertEqual(workbook["Trajectories"].max_row, 3)
                self.assertEqual(workbook["Steps"].max_row, 3)
                headers = [cell.value for cell in workbook["Steps"][1]]
                self.assertNotIn("thought", headers)
            finally:
                workbook.close()
            self.assertEqual([item["task_id"] for item in manifest["tasks"]], ["TASK-A", "TASK-B"])
            first_tree = json.loads((published / "TASK-A.json").read_text(encoding="utf-8"))
            self.assertEqual(first_tree["task_id"], "TASK-A")
            self.assertEqual(first_tree["original_step_count"], 1)
            self.assertEqual(first_tree["children"][0]["occurrences"][0]["summary"], "TASK-A summary")
            self.assertEqual(updates[-1]["stage"], "publishing")

    def test_failure_removes_temporary_batch_and_publishes_nothing(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            trajectory_root, workbook_path = prepare_source(root)
            runs_dir = root / "runs"
            with (
                patch("backend.tree_build_service.configure_reviewer_environment", return_value="fake-model"),
                patch(
                    "backend.tree_build_service.QwenIntermediateStateClassifier",
                    FailingSecondTaskClassifier,
                ),
                patch("backend.tree_build_service.QwenStateAlignmentReviewer", FakeAlignmentReviewer),
            ):
                with self.assertRaisesRegex(RuntimeError, "synthetic classifier failure"):
                    build_tree_run(
                        ["TASK-A", "TASK-B"],
                        job_id="job-failure",
                        progress=lambda _changes: None,
                        xlsx_path=workbook_path,
                        trajectory_root=trajectory_root,
                        runs_dir=runs_dir,
                        env_path=root / ".env",
                        classification_cache=root / "classification.json",
                        alignment_cache=root / "alignment.json",
                    )

            self.assertEqual(list(runs_dir.iterdir()), [])


if __name__ == "__main__":
    unittest.main()
