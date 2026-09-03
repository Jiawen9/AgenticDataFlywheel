"""Reproduce the SFT/RL/native routing rules of ``human8.0.py``."""

from __future__ import annotations

from copy import copy
from datetime import datetime
import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .constants import WORKBOOK_SUFFIXES


def _edit(session: dict[str, Any], row_number: int) -> dict[str, Any]:
    return session.get("row_edits", {}).get(str(row_number), {})


def _active_rows(session: dict[str, Any], group: dict[str, Any]) -> list[dict[str, Any]]:
    return [row for row in group["rows"] if not _edit(session, int(row["excel_row"])).get("deleted", False)]


def _is_action_edited(session: dict[str, Any], row_number: int) -> bool:
    return "actions" in _edit(session, row_number)


def _is_sop_edited(session: dict[str, Any], row_number: int) -> bool:
    return "sop" in _edit(session, row_number)


def _row_values(
    sheet: Any,
    headers: list[str],
    row_number: int,
    session: dict[str, Any],
    *,
    restored_action: str | None = None,
) -> list[Any]:
    edits = _edit(session, row_number)
    values = [sheet.cell(row_number, index + 1).value for index in range(len(headers))]
    action_name = "actions" if "actions" in headers else "action"
    if action_name in headers:
        values[headers.index(action_name)] = (
            restored_action if restored_action is not None else edits.get("actions", values[headers.index(action_name)])
        )
    if "sop" in headers and "sop" in edits:
        values[headers.index("sop")] = edits["sop"]
    if "actions_box" in headers and "actions_box" in edits:
        values[headers.index("actions_box")] = edits["actions_box"]
    if "summary" in headers and "summary" in edits:
        values[headers.index("summary")] = edits["summary"]
    if "thought" in headers and "thought" in edits:
        values[headers.index("thought")] = edits["thought"]
    cot = session.get("cot", {}).get(str(row_number), {})
    if isinstance(cot, dict):
        action_value = values[headers.index(action_name)] if action_name in headers else ""
        try:
            action_object = json.loads(str(action_value))
        except (TypeError, ValueError):
            action_object = None
        action_hash = hashlib.sha256(
            json.dumps(action_object, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode()
        ).hexdigest() if isinstance(action_object, dict) else ""
        if (
            cot.get("content_tag") == "thought_summary"
            and action_hash
            and cot.get("action_hash") == action_hash
        ):
            if "summary" in headers and cot.get("summary"):
                values[headers.index("summary")] = cot["summary"]
            if "thought" in headers and cot.get("thought"):
                values[headers.index("thought")] = cot["thought"]
    return values


def _append_sheet(workbook: Workbook, sheet_name: str, headers: list[str], rows: list[tuple[list[Any], int]]) -> int:
    sheet = workbook.create_sheet(sheet_name)
    sheet.append(headers)
    refined_index = headers.index("is_refined")
    for values, refined_row in rows:
        values = list(values)
        while len(values) < len(headers):
            values.append(None)
        values[refined_index] = refined_row
        sheet.append(values)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    return len(rows)


def export_session_workbook(
    *,
    workbook_path: Path,
    snapshot: dict[str, Any],
    session: dict[str, Any],
    output_dir: Path,
    export_id: str,
) -> dict[str, Any]:
    if workbook_path.suffix.lower() not in WORKBOOK_SUFFIXES:
        raise ValueError("仅支持 .xlsx 或 .xlsm 文件")
    source = load_workbook(workbook_path, data_only=False, keep_vba=workbook_path.suffix.lower() == ".xlsm")
    try:
        sheet = source.active
        source_headers = [
            str(sheet.cell(1, column).value).strip()
            for column in range(1, sheet.max_column + 1)
            if str(sheet.cell(1, column).value).strip()
        ]
        headers = [*source_headers]
        if "is_refined" not in headers:
            headers.append("is_refined")
        if "thought" not in headers:
            headers.append("thought")

        routed: dict[str, list[tuple[list[Any], int]]] = {
            "SFT_人工精修": [],
            "RL_负向反思": [],
            "原生_完美通过": [],
            "原生_异常待处理": [],
        }
        group_exports = session.get("group_exports", {})
        for group in snapshot["groups"]:
            if not bool(group_exports.get(group["group_id"], group["export"])):
                continue
            rows = _active_rows(session, group)
            action_rows = [row for row in rows if _is_action_edited(session, int(row["excel_row"]))]
            if not action_rows:
                if any(_is_sop_edited(session, int(row["excel_row"])) for row in rows):
                    routed["SFT_人工精修"].extend(
                        (_row_values(sheet, headers, int(row["excel_row"]), session), 1 if _is_sop_edited(session, int(row["excel_row"])) else 0)
                        for row in rows
                    )
                else:
                    sheet_name = "原生_完美通过" if group["quality"] == "完成且过程正常" else "原生_异常待处理"
                    routed[sheet_name].extend(
                        (_row_values(sheet, headers, int(row["excel_row"]), session), 0)
                        for row in rows
                    )
                continue

            first_action = int(action_rows[0]["excel_row"])
            sft_rows = rows[: next(index for index, row in enumerate(rows) if int(row["excel_row"]) == first_action) + 1]
            routed["SFT_人工精修"].extend(
                (_row_values(sheet, headers, int(row["excel_row"]), session), 1 if int(row["excel_row"]) == first_action else 0)
                for row in sft_rows
            )
            if len(action_rows) >= 2:
                second_action = int(action_rows[1]["excel_row"])
                rl_rows = rows[: next(index for index, row in enumerate(rows) if int(row["excel_row"]) == second_action) + 1]
                original_first = _edit(session, first_action).get("original_actions")
                routed["RL_负向反思"].extend(
                    (
                        _row_values(
                            sheet,
                            headers,
                            int(row["excel_row"]),
                            session,
                            restored_action=original_first if int(row["excel_row"]) == first_action else None,
                        ),
                        1 if int(row["excel_row"]) == second_action else 0,
                    )
                    for row in rl_rows
                )

        if not any(routed.values()):
            raise ValueError("当前没有符合导出条件的数据；请至少勾选一个任务")

        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{workbook_path.stem}_精修与筛查_{timestamp}_{export_id[:6]}.xlsx"
        output_path = output_dir / filename
        exported = Workbook()
        default = exported.active
        exported.remove(default)
        counts: dict[str, int] = {}
        for name, rows in routed.items():
            if rows:
                counts[name] = _append_sheet(exported, name, headers, rows)
        temporary = output_path.with_name(f".{output_path.name}.tmp")
        exported.save(temporary)
        temporary.replace(output_path)
        return {
            "filename": filename,
            "sheets": counts,
            "summary": {"rows": sum(counts.values()), "groups": len(snapshot["groups"])},
        }
    finally:
        source.close()


def export_full_dataset_workbook(
    *,
    workbook_path: Path,
    snapshot: dict[str, Any],
    session: dict[str, Any],
    output_dir: Path,
    export_id: str,
) -> dict[str, Any]:
    """Copy the source workbook and overlay the correction session's final values."""
    if workbook_path.suffix.lower() not in WORKBOOK_SUFFIXES:
        raise ValueError("仅支持 .xlsx 或 .xlsm 文件")
    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(workbook_path, data_only=False, keep_vba=keep_vba)
    try:
        sheet = workbook.active
        columns = {
            str(sheet.cell(1, column).value).strip(): column
            for column in range(1, sheet.max_column + 1)
            if sheet.cell(1, column).value is not None
            and str(sheet.cell(1, column).value).strip()
        }
        action_name = "actions" if "actions" in columns else "action"
        if action_name not in columns:
            raise ValueError("源工作簿缺少 action/actions 列")

        def ensure_column(name: str, style_from: str | None = None) -> int:
            existing = columns.get(name)
            if existing is not None:
                return existing
            column = sheet.max_column + 1
            target = sheet.cell(1, column, name)
            source_column = columns.get(style_from or "")
            if source_column is not None:
                source = sheet.cell(1, source_column)
                target._style = copy(source._style)
                target.font = copy(source.font)
                target.fill = copy(source.fill)
                target.border = copy(source.border)
                target.alignment = copy(source.alignment)
                target.number_format = source.number_format
                target.protection = copy(source.protection)
            columns[name] = column
            return column

        bbox_column = ensure_column("actions_box")
        summary_column = ensure_column("summary")
        thought_column = ensure_column("thought", "summary")
        action_column = columns[action_name]

        changed_rows = 0
        all_rows = [row for group in snapshot["groups"] for row in group["rows"]]
        for row in all_rows:
            row_number = int(row["excel_row"])
            edits = _edit(session, row_number)
            cot = session.get("cot", {}).get(str(row_number), {})
            relevant_edit = any(
                name in edits for name in ("actions", "actions_box", "summary", "thought")
            )
            if not relevant_edit and not (isinstance(cot, dict) and cot):
                continue

            action_value = edits.get("actions", sheet.cell(row_number, action_column).value)
            bbox_value = edits.get("actions_box", sheet.cell(row_number, bbox_column).value)
            summary_value = edits.get(
                "summary",
                sheet.cell(row_number, summary_column).value
                or row.get("original_summary")
                or row.get("summary")
                or "",
            )
            thought_value = edits.get(
                "thought",
                sheet.cell(row_number, thought_column).value
                or row.get("original_thought")
                or row.get("thought")
                or "",
            )

            try:
                action_object = json.loads(str(action_value))
            except (TypeError, ValueError):
                action_object = None
            action_hash = (
                hashlib.sha256(
                    json.dumps(
                        action_object,
                        ensure_ascii=False,
                        sort_keys=True,
                        separators=(",", ":"),
                    ).encode()
                ).hexdigest()
                if isinstance(action_object, dict)
                else ""
            )
            if (
                isinstance(cot, dict)
                and cot.get("content_tag") == "thought_summary"
                and action_hash
                and cot.get("action_hash") == action_hash
            ):
                summary_value = cot.get("summary") or summary_value
                thought_value = cot.get("thought") or thought_value

            sheet.cell(row_number, action_column).value = action_value
            sheet.cell(row_number, bbox_column).value = bbox_value
            sheet.cell(row_number, summary_column).value = summary_value
            sheet.cell(row_number, thought_column).value = thought_value
            changed_rows += 1

        output_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        suffix = workbook_path.suffix.lower()
        filename = f"{workbook_path.stem}_专家纠偏完整数据集_{timestamp}_{export_id[:6]}{suffix}"
        output_path = output_dir / filename
        temporary = output_path.with_name(f".{output_path.name}.tmp")
        workbook.save(temporary)
        temporary.replace(output_path)
        return {
            "filename": filename,
            "sheets": {sheet.title: max(0, sheet.max_row - 1)},
            "summary": {
                "rows": max(0, sheet.max_row - 1),
                "groups": len(snapshot["groups"]),
                "changed_rows": changed_rows,
            },
        }
    finally:
        workbook.close()
