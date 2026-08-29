"""Read and interpret the Excel contract used by ``human8.0.py``.

The desktop tool uses pandas and keeps its edits in memory.  The web module
uses openpyxl instead, so the original workbook stays untouched while a
small JSON draft stores only the user's edits.
"""

from __future__ import annotations

import json
import re
import threading
from copy import deepcopy
from datetime import date, datetime, time
from pathlib import Path, PurePosixPath
from typing import Any, Iterable

from openpyxl import load_workbook

from .constants import ACTION_TYPES, WORKBOOK_SUFFIXES


REQUIRED_COLUMNS = {"task", "meta_task", "image"}
ACTION_COLUMN_NAMES = ("actions", "action")
ANNOTATED_REQUIRED_COLUMNS = {"文件夹名", "image", "action"}


# Reading the annotated workbook is comparatively expensive because the
# adapter also resolves trajectory metadata and image assets for every row.
# Keep the parsed result in memory for the lifetime of the backend process and
# invalidate it when the source workbook changes.  The lock deliberately
# covers the first load so concurrent API requests do not all parse the same
# workbook at once.
_SNAPSHOT_CACHE: dict[tuple[str, str, str | None, int, int], dict[str, Any]] = {}
_SNAPSHOT_CACHE_LOCK = threading.RLock()


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def parse_action(value: Any) -> dict[str, Any]:
    raw = text(value)
    try:
        parsed = json.loads(raw)
    except (TypeError, ValueError):
        parsed = None
    if isinstance(parsed, list) and len(parsed) == 1 and isinstance(parsed[0], dict):
        parsed = parsed[0]
    if isinstance(parsed, dict):
        action = text(parsed.get("action")).lower()
        return {**parsed, "action": action or "unknown"}
    return {"action": "unknown", "raw": raw}


def action_type(action: dict[str, Any]) -> str:
    value = text(action.get("action")).lower()
    return value if value in ACTION_TYPES else "unknown"


def _step_number(image: str, fallback: int) -> int:
    name = Path(image.replace("\\", "/")).name
    match = re.search(r"step(\d+)", name, re.IGNORECASE)
    return int(match.group(1)) if match else fallback


def _headers(sheet: Any) -> dict[str, int]:
    return {
        text(cell.value): cell.column
        for cell in sheet[1]
        if text(cell.value)
    }


def _column(headers: dict[str, int], *names: str) -> int | None:
    for name in names:
        if name in headers:
            return headers[name]
    return None


def _cell_value(sheet: Any, row: int, headers: dict[str, int], *names: str) -> Any:
    column = _column(headers, *names)
    return sheet.cell(row, column).value if column else None


def _quality(rows: list[dict[str, Any]]) -> str:
    if not rows:
        return "未知"
    result = text(rows[0].get("task_manual_result"))
    completion = "完成" if result in {"1", "1.0"} else "未完成" if result in {"0", "0.0"} else "未知"
    abnormal = False
    for row in rows[:-1]:
        micro = text(row.get("micro_manual"))
        macro = text(row.get("macro_manual"))
        if micro in {"0", "0.0"} or macro in {"0", "0.0"}:
            abnormal = True
            break
        if not micro and not macro and "bad_interval" in text(row.get("Bad_Interval")).lower():
            abnormal = True
            break
    if completion == "完成":
        return "完成但过程有异常" if abnormal else "完成且过程正常"
    if completion == "未完成":
        return "未完成且过程有异常" if abnormal else "未完成但过程未发现明显异常"
    return "未知"


def _group_prefix(quality: str) -> str:
    return {
        "完成且过程正常": "[🟢 完/正]",
        "完成但过程有异常": "[🔴 完/异常]",
        "未完成且过程有异常": "[🔴 未/异常]",
        "未完成但过程未发现明显异常": "[🟡 未/正]",
    }.get(quality, "[⚪ 未知]")


def _row_payload(sheet: Any, row_number: int, headers: dict[str, int]) -> dict[str, Any]:
    values = {
        name: json_value(sheet.cell(row_number, column).value)
        for name, column in headers.items()
    }
    image = text(_cell_value(sheet, row_number, headers, "image"))
    actions = text(_cell_value(sheet, row_number, headers, *ACTION_COLUMN_NAMES))
    action = parse_action(actions)
    return {
        "excel_row": row_number,
        "step": _step_number(image, row_number - 1),
        "task": text(_cell_value(sheet, row_number, headers, "task")),
        "meta_task": text(_cell_value(sheet, row_number, headers, "meta_task")),
        "image": image,
        "xml": text(_cell_value(sheet, row_number, headers, "xml")),
        "actions": actions,
        "action": action,
        "sop": text(_cell_value(sheet, row_number, headers, "sop")),
        "summary": text(_cell_value(sheet, row_number, headers, "summary", "action_summary")),
        "task_manual_result": text(_cell_value(sheet, row_number, headers, "task_manual_result")),
        "micro_manual": text(_cell_value(sheet, row_number, headers, "micro_manual")),
        "macro_manual": text(_cell_value(sheet, row_number, headers, "macro_manual")),
        "micro_pred": text(_cell_value(sheet, row_number, headers, "micro_pred")),
        "macro_pred": text(_cell_value(sheet, row_number, headers, "macro_pred")),
        "Bad_Interval": text(_cell_value(sheet, row_number, headers, "Bad_Interval")),
        "trajectory_quality_type": text(
            _cell_value(sheet, row_number, headers, "trajectory_quality_type")
        ),
        "actions_box": text(_cell_value(sheet, row_number, headers, "actions_box")),
        "values": values,
    }


def _worksheet_max_row(sheet: Any) -> int:
    max_row = sheet.max_row
    if max_row is None:
        try:
            sheet.calculate_dimension(force=True)
        except (AttributeError, ValueError):
            pass
        max_row = sheet.max_row
    return max_row or 1


def _safe_asset_path(value: str, root: Path) -> Path | None:
    """Resolve a workbook resource below ``root`` without path traversal."""
    normalized = text(value).replace("\\", "/").strip("/")
    if not normalized or normalized.startswith("..") or "/../" in f"/{normalized}/":
        return None
    candidate = (root / PurePosixPath(normalized)).resolve()
    try:
        candidate.relative_to(root.resolve())
    except ValueError:
        return None
    return candidate


def _trajectory_json_path(
    asset_root: Path | None,
    trajectory_id: str,
    image: str,
) -> Path | None:
    if asset_root is None or not asset_root.is_dir():
        return None

    image_path = _safe_asset_path(image, asset_root)
    candidates: list[Path] = []
    if image_path is not None:
        candidates.extend(
            [
                image_path.parent / "_trajectory_for_evaluate.json",
                image_path.parent.parent / "_trajectory_for_evaluate.json",
            ]
        )
    if trajectory_id:
        direct = asset_root / trajectory_id / "_trajectory_for_evaluate.json"
        if direct not in candidates:
            candidates.append(direct)
        # The built-in rollout layout has one task directory above the
        # trajectory directory.  Match the exact directory name only.
        candidates.extend(
            path / "_trajectory_for_evaluate.json"
            for path in asset_root.rglob(trajectory_id)
            if path.is_dir()
        )

    seen: set[Path] = set()
    for candidate in candidates:
        resolved = candidate.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        try:
            resolved.relative_to(asset_root.resolve())
        except ValueError:
            continue
        if resolved.is_file():
            return resolved
    return None


def _trajectory_task(
    asset_root: Path | None,
    trajectory_id: str,
    image: str,
    cache: dict[str, str],
) -> str:
    if trajectory_id in cache:
        return cache[trajectory_id]
    task = ""
    json_path = _trajectory_json_path(asset_root, trajectory_id, image)
    if json_path is not None:
        try:
            payload = json.loads(json_path.read_text(encoding="utf-8-sig"))
            if isinstance(payload, dict):
                task = text(payload.get("task"))
        except (OSError, TypeError, ValueError):
            task = ""
    cache[trajectory_id] = task
    return task


def _annotated_row_payload(
    sheet: Any,
    row_number: int,
    headers: dict[str, int],
    *,
    asset_root: Path | None,
    task_cache: dict[str, str],
) -> dict[str, Any]:
    values = {
        name: json_value(sheet.cell(row_number, column).value)
        for name, column in headers.items()
    }
    image = text(_cell_value(sheet, row_number, headers, "image"))
    trajectory_id = text(_cell_value(sheet, row_number, headers, "文件夹名"))
    if not trajectory_id and image:
        trajectory_id = PurePosixPath(image.replace("\\", "/")).parent.name
    actions = text(_cell_value(sheet, row_number, headers, "action"))
    task = _trajectory_task(asset_root, trajectory_id, image, task_cache) or trajectory_id
    return {
        "excel_row": row_number,
        "step": _step_number(image, row_number - 1),
        "task": task,
        "meta_task": trajectory_id,
        "trajectory_id": trajectory_id,
        "image": image,
        "xml": text(_cell_value(sheet, row_number, headers, "xml")),
        # Keep the internal plural name used by the correction editor while
        # retaining the source workbook's singular ``action`` header.
        "actions": actions,
        "action": parse_action(actions),
        "sop": "",
        "summary": text(_cell_value(sheet, row_number, headers, "summary")),
        "task_manual_result": "",
        "micro_manual": "",
        "macro_manual": "",
        "micro_pred": "",
        "macro_pred": "",
        "Bad_Interval": "",
        "trajectory_quality_type": "",
        "actions_box": text(_cell_value(sheet, row_number, headers, "actions_box")),
        "values": values,
    }


def _load_annotated_snapshot(
    sheet: Any,
    headers: dict[str, int],
    *,
    asset_root: Path | None,
    workbook_path: Path,
) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    task_cache: dict[str, str] = {}
    for row_number in range(2, _worksheet_max_row(sheet) + 1):
        payload = _annotated_row_payload(
            sheet,
            row_number,
            headers,
            asset_root=asset_root,
            task_cache=task_cache,
        )
        if not payload["meta_task"] and not payload["image"] and not payload["actions"]:
            continue
        if not payload["meta_task"]:
            raise ValueError(f"Excel 第 {row_number} 行缺少 文件夹名")
        rows.append(payload)

    groups_by_id: dict[str, dict[str, Any]] = {}
    groups: list[dict[str, Any]] = []
    for row in rows:
        trajectory_id = row["meta_task"]
        group = groups_by_id.get(trajectory_id)
        if group is None:
            group = {
                "group_id": f"group_{len(groups)}",
                "task": row["task"] or trajectory_id,
                "meta_task": trajectory_id,
                "quality": "未知",
                "prefix": "[⚪ 未知]",
                "export": False,
                "rows": [],
            }
            groups_by_id[trajectory_id] = group
            groups.append(group)
        group["rows"].append(row)

    for group in groups:
        group["initial_indices"] = [row["excel_row"] for row in group["rows"]]

    return {
        "workbook": str(workbook_path),
        "headers": list(headers.keys()),
        "groups": groups,
        "row_count": len(rows),
        "source_kind": "annotated_workbook",
    }


def _load_snapshot_uncached(
    workbook_path: Path,
    *,
    asset_root: Path | None = None,
    source_kind: str | None = None,
) -> dict[str, Any]:
    """Load the first sheet and return a JSON-safe immutable-ish snapshot."""
    if workbook_path.suffix.lower() not in WORKBOOK_SUFFIXES:
        raise ValueError("仅支持 .xlsx 或 .xlsm 文件；旧版 .xls 请先另存为 .xlsx")
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Excel 文件不存在：{workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = _headers(sheet)
        if ANNOTATED_REQUIRED_COLUMNS.issubset(headers):
            return _load_annotated_snapshot(
                sheet,
                headers,
                asset_root=asset_root,
                workbook_path=workbook_path,
            )
        missing = sorted(REQUIRED_COLUMNS - headers.keys())
        if not _column(headers, *ACTION_COLUMN_NAMES):
            missing.append("actions")
        if missing:
            raise ValueError(f"Excel 缺少必要列：{', '.join(missing)}")

        rows: list[dict[str, Any]] = []
        max_row = _worksheet_max_row(sheet)
        # Some Excel writers omit the worksheet dimension when the sheet only
        # has headers.  Treat that as an empty dataset instead of attempting
        # ``None + 1`` and leaking a server traceback to the browser.
        for row_number in range(2, max_row + 1):
            payload = _row_payload(sheet, row_number, headers)
            if not payload["task"] and not payload["meta_task"] and not payload["image"]:
                continue
            rows.append(payload)
    finally:
        workbook.close()

    groups: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for row in rows:
        key = (row["task"], row["meta_task"])
        if current is None or current["task"] != key[0] or current["meta_task"] != key[1]:
            current = {
                "group_id": f"group_{len(groups)}",
                "task": key[0],
                "meta_task": key[1],
                "rows": [],
            }
            groups.append(current)
        current["rows"].append(row)

    for group in groups:
        quality = _quality(group["rows"])
        group["quality"] = quality
        group["prefix"] = _group_prefix(quality)
        group["export"] = quality == "完成且过程正常"
        group["initial_indices"] = [row["excel_row"] for row in group["rows"]]

    return {
        "workbook": str(workbook_path),
        "headers": list(headers.keys()),
        "groups": groups,
        "row_count": len(rows),
    }


def _snapshot_cache_key(
    workbook_path: Path,
    *,
    asset_root: Path | None,
    source_kind: str | None,
) -> tuple[str, str, str | None, int, int]:
    resolved_workbook = workbook_path.resolve()
    stat = resolved_workbook.stat()
    return (
        str(resolved_workbook),
        str(asset_root.resolve()) if asset_root is not None else "",
        source_kind,
        int(stat.st_mtime_ns),
        int(stat.st_size),
    )


def clear_snapshot_cache() -> None:
    """Clear parsed workbook snapshots, primarily for tests and reloads."""
    with _SNAPSHOT_CACHE_LOCK:
        _SNAPSHOT_CACHE.clear()


def load_snapshot(
    workbook_path: Path,
    *,
    asset_root: Path | None = None,
    source_kind: str | None = None,
) -> dict[str, Any]:
    """Load a workbook snapshot with process-local, change-aware caching.

    A deep copy is returned at the boundary so callers can safely filter or
    annotate their result without mutating the cached source snapshot.
    """
    workbook_path = Path(workbook_path)
    if workbook_path.suffix.lower() not in WORKBOOK_SUFFIXES:
        raise ValueError("仅支持 .xlsx 或 .xlsm 文件；旧版 .xls 请先另存为 .xlsx")
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Excel 文件不存在：{workbook_path}")

    key = _snapshot_cache_key(
        workbook_path,
        asset_root=asset_root,
        source_kind=source_kind,
    )
    with _SNAPSHOT_CACHE_LOCK:
        cached = _SNAPSHOT_CACHE.get(key)
        if cached is not None:
            return deepcopy(cached)

        # Remove stale versions for the same source.  This keeps repeated
        # edits to a workbook from growing the process-local cache forever.
        resolved_workbook = key[0]
        for stale_key in tuple(_SNAPSHOT_CACHE):
            if stale_key[0] == resolved_workbook and stale_key != key:
                _SNAPSHOT_CACHE.pop(stale_key, None)

        # Keep parsing under the lock: the first request pays the parse cost,
        # while concurrent requests wait for and then reuse its result.
        snapshot = _load_snapshot_uncached(
            workbook_path,
            asset_root=asset_root,
            source_kind=source_kind,
        )
        _SNAPSHOT_CACHE[key] = deepcopy(snapshot)
        return deepcopy(snapshot)


def iter_group_rows(snapshot: dict[str, Any], group_id: str) -> Iterable[dict[str, Any]]:
    for group in snapshot.get("groups", []):
        if group.get("group_id") == group_id:
            return group.get("rows", [])
    raise KeyError(group_id)
