"""Export rollout trajectories to Excel and append reviewed action boxes."""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    from .model_config import load_env_values, load_model_config
    from .bounding_box.build_annotations import resolve_action_box
    from .bounding_box.qwen_reviewer import QwenBoxReviewer
    from .export_vla_trajectories import collect_rows, write_xlsx
except ImportError:  # Keep direct `python backend/trajectories_preprocessing.py` usage working.
    from model_config import load_env_values, load_model_config
    from bounding_box.build_annotations import resolve_action_box
    from bounding_box.qwen_reviewer import QwenBoxReviewer
    from export_vla_trajectories import collect_rows, write_xlsx


BACKEND_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = BACKEND_DIR.parent
WORKSPACE_DIR = PROJECT_ROOT / "backend_workspace"
DEFAULT_SOURCE = WORKSPACE_DIR / "rollout_trajectories"
DEFAULT_EXPORT_OUTPUT = WORKSPACE_DIR / "trajectories_to_excel.xlsx"
DEFAULT_ANNOTATED_OUTPUT = WORKSPACE_DIR / "annotated_trajectories.xlsx"
DEFAULT_ENV_FILE = BACKEND_DIR / ".env"
DEFAULT_CACHE_FILE = BACKEND_DIR / "bounding_box" / "qwen_review_cache.json"
TARGET_ACTIONS = {"click", "swipe", "long_press"}
STEP_IMAGE_RE = re.compile(r"^step(?P<step>\d+)_vla_input\.jpg$", re.IGNORECASE)
REQUIRED_COLUMNS = ("文件夹名", "image", "xml", "action", "summary")


def read_env_file(path: Path) -> dict[str, str]:
    """Read the small KEY=VALUE configuration used by this backend."""
    return load_env_values(path)


def configure_reviewer_environment(env_file: Path, module: str = "bbox") -> str:
    """Resolve a module profile and export compatibility variables for old clients."""
    config = load_model_config(env_file, module=module)
    os.environ["MODEL_CONFIG_PATH"] = str(Path(env_file).expanduser().resolve())
    os.environ["TRAJECTORY_API_KEY"] = config.api_key
    os.environ["TRAJECTORY_API_BASE_URL"] = config.base_url
    os.environ["TRAJECTORY_MODEL"] = config.model
    os.environ["TRAJECTORY_HTTP_TIMEOUT"] = str(config.timeout)
    os.environ["TRAJECTORY_HTTP_VERIFY"] = str(config.verify).lower()
    os.environ["TRAJECTORY_HTTP_TRUST_ENV"] = str(config.trust_env).lower()
    if config.proxy:
        os.environ["TRAJECTORY_HTTP_PROXY_URL"] = config.proxy
    return config.model


def export_trajectories(source_root: Path, output_path: Path) -> tuple[int, list[str]]:
    """Run the repository exporter against the complete rollout tree."""
    source_root = source_root.expanduser().resolve()
    if not source_root.is_dir():
        raise FileNotFoundError(f"trajectory source directory does not exist: {source_root}")
    rows, warnings = collect_rows(source_root)
    if not rows:
        raise ValueError(f"no final VLA trajectory steps found under {source_root}")
    write_xlsx(rows, output_path.expanduser().resolve())
    return len(rows), warnings


def parse_action_cell(value: Any) -> dict[str, Any]:
    if not isinstance(value, str) or not value.strip():
        raise ValueError("action cell is empty")
    parsed = json.loads(value)
    if isinstance(parsed, list):
        if len(parsed) != 1:
            raise ValueError(f"expected one action, found {len(parsed)}")
        parsed = parsed[0]
    if not isinstance(parsed, dict):
        raise ValueError(f"action must be a JSON object, got {type(parsed).__name__}")
    return parsed


def step_from_image_path(image_path: Path) -> int:
    match = STEP_IMAGE_RE.match(image_path.name)
    if not match:
        raise ValueError(f"cannot determine step from image filename: {image_path.name}")
    return int(match.group("step"))


def load_executed_actions(run_dir: Path) -> dict[int, dict[str, Any]]:
    evaluation_path = run_dir / "_trajectory_for_evaluate.json"
    if not evaluation_path.is_file():
        raise FileNotFoundError(f"evaluation trajectory does not exist: {evaluation_path}")
    payload = json.loads(evaluation_path.read_text(encoding="utf-8-sig"))
    actions = payload.get("actions_flat")
    if not isinstance(actions, list):
        raise ValueError(f"actions_flat is missing or invalid: {evaluation_path}")

    result: dict[int, dict[str, Any]] = {}
    for item in actions:
        if not isinstance(item, dict) or "global_step" not in item:
            continue
        step = int(item["global_step"])
        action = item.get("action")
        if not isinstance(action, dict):
            raise ValueError(f"invalid executed action at {run_dir.name}/step{step:03d}")
        if step in result:
            raise ValueError(f"duplicate executed action at {run_dir.name}/step{step:03d}")
        result[step] = action
    return result


def swipe_direction(action: dict[str, Any]) -> str:
    start = action.get("start_coordinate")
    end = action.get("end_coordinate")
    if not isinstance(start, list) or len(start) < 2 or not isinstance(end, list) or len(end) < 2:
        raise ValueError("swipe action is missing start/end coordinates")
    dx = float(end[0]) - float(start[0])
    dy = float(end[1]) - float(start[1])
    if dx == 0 and dy == 0:
        raise ValueError("swipe start and end coordinates are identical")
    if abs(dx) > abs(dy):
        return "left" if dx < 0 else "right"
    return "up" if dy < 0 else "down"


def format_actions_box(action: dict[str, Any], bbox: tuple[int, int, int, int]) -> str:
    kind = str(action.get("action", "")).lower()
    bbox_text = "[" + ",".join(str(int(value)) for value in bbox) + "]"
    tagged_bbox = f"<bbox>{bbox_text}</bbox>"
    if kind == "click":
        return f"click(bbox={tagged_bbox})"
    if kind == "long_press":
        return f"long_press(bbox={tagged_bbox})"
    if kind == "swipe":
        return f"swipe_screen(bbox={tagged_bbox}, direction={swipe_direction(action)})"
    raise ValueError(f"unsupported boxed action: {kind!r}")


def _header_map(sheet: Any) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in sheet[1]
        if cell.value is not None and str(cell.value).strip()
    }


def resolve_artifact_path(value: Any, trajectory_root: Path | None) -> Path:
    path = Path(str(value)).expanduser()
    if path.is_absolute():
        return path.resolve()
    if trajectory_root is None:
        raise ValueError(f"relative artifact path requires trajectory_root: {path}")
    return (trajectory_root / path).resolve()


def annotate_trajectory_workbook(
    source_path: Path,
    output_path: Path,
    *,
    reviewer: Any,
    max_review_rounds: int = 4,
    trajectory_root: Path | None = None,
) -> dict[str, int]:
    """Append actions_box values, publishing the output only after all rows succeed."""
    source_path = source_path.expanduser().resolve()
    output_path = output_path.expanduser().resolve()
    if trajectory_root is not None:
        trajectory_root = trajectory_root.expanduser().resolve()
    workbook = load_workbook(source_path)
    sheet = workbook.active
    headers = _header_map(sheet)
    missing_columns = [name for name in REQUIRED_COLUMNS if name not in headers]
    if missing_columns:
        raise ValueError(f"workbook is missing columns: {', '.join(missing_columns)}")

    box_column = headers.get("actions_box")
    if box_column is None:
        box_column = sheet.max_column + 1
        box_header = sheet.cell(row=1, column=box_column, value="actions_box")
        source_header = sheet.cell(row=1, column=headers["action"])
        box_header._style = copy(source_header._style)
        box_header.font = copy(source_header.font)
        box_header.fill = copy(source_header.fill)
        box_header.border = copy(source_header.border)
        box_header.alignment = copy(source_header.alignment)
        box_header.number_format = source_header.number_format
        box_header.protection = copy(source_header.protection)
    sheet.column_dimensions[get_column_letter(box_column)].width = 70

    action_cache: dict[Path, dict[int, dict[str, Any]]] = {}
    counts = {"rows": 0, "annotated": 0, "blank": 0}
    for row_number in range(2, sheet.max_row + 1):
        counts["rows"] += 1
        row_label = f"Excel row {row_number}"
        try:
            raw_action = parse_action_cell(sheet.cell(row_number, headers["action"]).value)
            raw_kind = str(raw_action.get("action", "")).lower()
            box_cell = sheet.cell(row_number, box_column)
            if raw_kind not in TARGET_ACTIONS:
                box_cell.value = None
                counts["blank"] += 1
                continue

            image_path = resolve_artifact_path(
                sheet.cell(row_number, headers["image"]).value,
                trajectory_root,
            )
            run_dir = image_path.parent
            step = step_from_image_path(image_path)
            row_label = f"{run_dir.name}/step{step:03d} (Excel row {row_number})"
            if run_dir not in action_cache:
                action_cache[run_dir] = load_executed_actions(run_dir)
            executed_action = action_cache[run_dir].get(step)
            if executed_action is None:
                raise ValueError("executed action is missing from _trajectory_for_evaluate.json")
            executed_kind = str(executed_action.get("action", "")).lower()
            if executed_kind != raw_kind:
                raise ValueError(
                    f"action mismatch: Excel has {raw_kind!r}, evaluation has {executed_kind!r}"
                )

            stability_path = run_dir / f"step{step:03d}_vla_input_stability.jpg"
            if not stability_path.is_file():
                raise FileNotFoundError(f"stability screenshot does not exist: {stability_path}")
            xml_path = resolve_artifact_path(
                sheet.cell(row_number, headers["xml"]).value,
                trajectory_root,
            )
            if not xml_path.is_file():
                raise FileNotFoundError(f"UI XML does not exist: {xml_path}")
            xml_text = xml_path.read_text(encoding="utf-8", errors="replace")
            summary_value = sheet.cell(row_number, headers["summary"]).value
            resolution = resolve_action_box(
                image_path=stability_path,
                xml_text=xml_text,
                action=executed_action,
                action_summary=str(summary_value or ""),
                reviewer=reviewer,
                max_review_rounds=max_review_rounds,
            )
            box_cell.value = format_actions_box(executed_action, resolution.result.bbox)
            counts["annotated"] += 1
            print(f"Annotated {row_label}: {box_cell.value}", flush=True)
        except Exception as exc:
            raise RuntimeError(f"failed to annotate {row_label}: {exc}") from exc

    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(f".{output_path.stem}.tmp{output_path.suffix}")
    try:
        workbook.save(temporary_path)
        temporary_path.replace(output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    return counts


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export rollout trajectories and add Qwen-reviewed action boxes."
    )
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--export-output", type=Path, default=DEFAULT_EXPORT_OUTPUT)
    parser.add_argument("--annotated-output", type=Path, default=DEFAULT_ANNOTATED_OUTPUT)
    parser.add_argument("--env-file", type=Path, default=DEFAULT_ENV_FILE)
    parser.add_argument("--max-review-rounds", type=int, default=4)
    return parser.parse_args()


def run_pipeline(
    *,
    source: Path,
    export_output: Path,
    annotated_output: Path,
    env_file: Path,
    max_review_rounds: int,
) -> dict[str, Any]:
    row_count, warnings = export_trajectories(source, export_output)
    print(f"Exported {row_count} steps to: {export_output.expanduser().resolve()}", flush=True)
    for warning in warnings:
        print(f"Warning: {warning}", file=sys.stderr)

    model = configure_reviewer_environment(env_file.expanduser().resolve(), module="bbox")
    reviewer = QwenBoxReviewer(model=model, cache_path=DEFAULT_CACHE_FILE)
    counts = annotate_trajectory_workbook(
        export_output,
        annotated_output,
        reviewer=reviewer,
        max_review_rounds=max(1, max_review_rounds),
        trajectory_root=source,
    )
    print(f"Annotated workbook: {annotated_output.expanduser().resolve()}", flush=True)
    print(
        f"Rows={counts['rows']} annotated={counts['annotated']} blank={counts['blank']}",
        flush=True,
    )
    return {"exported_rows": row_count, "warnings": warnings, **counts}


def main() -> int:
    args = parse_args()
    try:
        run_pipeline(
            source=args.source,
            export_output=args.export_output,
            annotated_output=args.annotated_output,
            env_file=args.env_file,
            max_review_rounds=args.max_review_rounds,
        )
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
