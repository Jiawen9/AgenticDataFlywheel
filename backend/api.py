"""FastAPI application for trajectory collection, tree building, and inspection."""

from __future__ import annotations

import json
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any, Optional

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from openpyxl import load_workbook

from .trajectory_data import (
    ANNOTATED_XLSX,
    PROJECT_ROOT,
    TREE_RUNS_DIR,
    WORKSPACE_DIR,
    discover_tasks,
    find_tree_run,
    list_tree_runs,
    load_annotated_trajectory,
    load_annotated_trajectories,
    resolve_image_asset,
    task_summaries,
    trajectory_summaries,
    update_action_bbox,
)
from .quality_data import quality_manifest, quality_task, rubric_ready
from .quality_jobs import QualityJobManager
from .tree_build_jobs import TreeBuildJobManager
from .trajectory_correction.router import configure_cot_job_manager, router as correction_router
from .trajectory_correction.cot_jobs import CotJobManager
from .task_generation.jobs import TaskGenerationJobManager
from .task_generation.router import configure_job_manager, router as task_generation_router
from .data_publishing.router import router as data_publishing_router


FRONTEND_DIST = PROJECT_ROOT / "frontend" / "dist"


def _observation_index(workbook_path: Path, task_id: str) -> dict[tuple[str, int], str]:
    if not workbook_path.is_file():
        return {}
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if "Steps" not in workbook.sheetnames:
            return {}
        rows = workbook["Steps"].iter_rows(values_only=True)
        headers = {str(value): index for index, value in enumerate(next(rows, ())) if value is not None}
        required = {"trajectory_id", "task_id", "step_id", "observation"}
        if not required.issubset(headers):
            return {}
        result = {}
        for row in rows:
            if str(row[headers["task_id"]]) != task_id:
                continue
            observation = str(row[headers["observation"]] or "").strip()
            if observation:
                result[(str(row[headers["trajectory_id"]]), int(row[headers["step_id"]]))] = observation
        return result
    finally:
        workbook.close()


def _attach_tree_observations(tree: dict[str, Any], observations: dict[tuple[str, int], str]) -> None:
    stack = [tree]
    while stack:
        node = stack.pop()
        for occurrence in node.get("occurrences", []):
            key = (str(occurrence.get("trajectory", "")), int(occurrence.get("step", 0)))
            if key in observations:
                occurrence["observation"] = observations[key]
        stack.extend(node.get("children", []))
    for trajectory in tree.get("source_trajectories", []):
        trajectory_id = str(trajectory.get("trajectory", ""))
        for step in trajectory.get("steps", []):
            key = (trajectory_id, int(step.get("step", 0)))
            if key in observations:
                step["observation"] = observations[key]


class TreeBuildRequest(BaseModel):
    task_ids: list[str]


class QualityJobRequest(BaseModel):
    run_id: str
    task_ids: list[str]


class BBoxUpdateRequest(BaseModel):
    excel_row: int
    bbox: list[int]
    action: Optional[dict[str, Any]] = None


app = FastAPI(title="Agentic Data Flywheel", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=False,
    allow_methods=["GET", "POST", "PATCH", "OPTIONS"],
    allow_headers=["*"],
)
app.include_router(correction_router)
app.include_router(task_generation_router)
app.include_router(data_publishing_router)
model_job_executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="model-job")
job_manager = TreeBuildJobManager(executor=model_job_executor)
quality_job_manager = QualityJobManager(executor=model_job_executor)
cot_job_manager = CotJobManager(executor=model_job_executor)
task_generation_job_manager = TaskGenerationJobManager(executor=model_job_executor)
configure_job_manager(task_generation_job_manager)
configure_cot_job_manager(cot_job_manager)


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/tasks")
def get_tasks() -> dict[str, Any]:
    return {"tasks": task_summaries()}


@app.get("/api/tasks/{task_id}/trajectories")
def get_task_trajectories(task_id: str) -> dict[str, Any]:
    tasks = {item["task_id"]: item for item in task_summaries()}
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="任务不存在")
    trajectories = trajectory_summaries(task_id)
    return {"task": tasks[task_id], "trajectories": trajectories}


@app.get("/api/tasks/{task_id}/trajectories/{trajectory_id}")
def get_task_trajectory(task_id: str, trajectory_id: str) -> dict[str, Any]:
    tasks = {item["task_id"]: item for item in task_summaries()}
    if task_id not in tasks:
        raise HTTPException(status_code=404, detail="任务不存在")
    trajectory = load_annotated_trajectory(task_id, trajectory_id)
    if trajectory is None:
        raise HTTPException(status_code=404, detail="轨迹不存在")
    return {"trajectory": trajectory}


@app.patch("/api/tasks/{task_id}/trajectories/{trajectory_id}/steps/{step}/bbox")
def patch_step_bbox(
    task_id: str,
    trajectory_id: str,
    step: int,
    request: BBoxUpdateRequest,
) -> dict[str, Any]:
    if len(request.bbox) != 4:
        raise HTTPException(status_code=422, detail="bbox 必须包含四个整数")
    try:
        actions_box = update_action_bbox(
            task_id,
            trajectory_id,
            step,
            request.excel_row,
            tuple(request.bbox),
            action_override=request.action,
        )
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except (OSError, PermissionError) as exc:
        raise HTTPException(status_code=409, detail=f"无法更新 Excel，请确认文件未被占用：{exc}") from exc
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return {"actions_box": actions_box}


@app.post("/api/tree-builds", status_code=202)
def create_tree_build(request: TreeBuildRequest) -> dict[str, Any]:
    unique = list(dict.fromkeys(value.strip() for value in request.task_ids if value.strip()))
    if not unique:
        raise HTTPException(status_code=422, detail="至少选择一个任务")
    tasks = {item["task_id"]: item for item in task_summaries()}
    unknown = [task_id for task_id in unique if task_id not in tasks]
    if unknown:
        raise HTTPException(status_code=404, detail=f"任务不存在：{', '.join(unknown)}")
    unprocessed = [task_id for task_id in unique if not tasks[task_id]["annotated"]]
    if unprocessed:
        raise HTTPException(
            status_code=409,
            detail=f"任务尚未完成轨迹预处理：{', '.join(unprocessed)}",
        )
    return job_manager.submit(unique)


@app.get("/api/tree-builds/{job_id}")
def get_tree_build(job_id: str) -> dict[str, Any]:
    payload = job_manager.get(job_id)
    if payload is None:
        raise HTTPException(status_code=404, detail="建树作业不存在")
    return payload


@app.post("/api/quality-jobs", status_code=202)
def create_quality_job(request: QualityJobRequest) -> dict[str, Any]:
    run_id = request.run_id.strip()
    manifest = find_tree_run(run_id)
    if manifest is None:
        raise HTTPException(status_code=404, detail="任务集不存在")
    unique = list(dict.fromkeys(value.strip() for value in request.task_ids if value.strip()))
    if not unique:
        raise HTTPException(status_code=422, detail="至少选择一个任务")
    available = {str(item.get("task_id")) for item in manifest.get("tasks", [])}
    unknown = [task_id for task_id in unique if task_id not in available]
    if unknown:
        raise HTTPException(status_code=404, detail=f"任务不在该任务集中：{', '.join(unknown)}")
    return quality_job_manager.submit(run_id, unique)


@app.get("/api/quality-jobs")
def get_quality_jobs() -> dict[str, Any]:
    return {"jobs": quality_job_manager.list_jobs()}


@app.get("/api/quality-jobs/{job_id}")
def get_quality_job(job_id: str) -> dict[str, Any]:
    payload = quality_job_manager.get(job_id)
    if payload is None:
        raise HTTPException(status_code=404, detail="质检作业不存在")
    return payload


@app.get("/api/tree-runs")
def get_tree_runs() -> dict[str, Any]:
    return {"runs": list_tree_runs()}


@app.get("/api/tree-runs/{run_id}")
def get_tree_run(run_id: str) -> dict[str, Any]:
    manifest = find_tree_run(run_id)
    if manifest is None:
        raise HTTPException(status_code=404, detail="任务集不存在")
    return manifest


@app.get("/api/tree-runs/{run_id}/quality")
def get_run_quality(run_id: str) -> dict[str, Any]:
    tree_manifest = find_tree_run(run_id)
    if tree_manifest is None:
        raise HTTPException(status_code=404, detail="任务集不存在")
    published = quality_manifest(run_id)
    completed = {str(item.get("task_id")): item for item in published.get("tasks", [])}
    tasks = []
    for task in tree_manifest.get("tasks", []):
        task_id = str(task.get("task_id", ""))
        tasks.append({
            "task_id": task_id,
            "status": "succeeded" if task_id in completed else "unreviewed",
            "rubric_ready": rubric_ready(task_id),
            **completed.get(task_id, {}),
        })
    return {"run_id": run_id, "updated_at": published.get("updated_at"), "tasks": tasks}


@app.get("/api/tree-runs/{run_id}/tasks/{task_id}/quality")
def get_task_quality(run_id: str, task_id: str) -> dict[str, Any]:
    manifest = find_tree_run(run_id)
    if manifest is None:
        raise HTTPException(status_code=404, detail="任务集不存在")
    available = {str(item.get("task_id")) for item in manifest.get("tasks", [])}
    if task_id not in available:
        raise HTTPException(status_code=404, detail="任务不在该任务集中")
    result = quality_task(run_id, task_id)
    if result is None:
        raise HTTPException(status_code=404, detail="该任务还没有质检结果")
    return result


@app.get("/api/tree-runs/{run_id}/tasks/{task_id}/tree")
def get_task_tree(run_id: str, task_id: str) -> JSONResponse:
    manifest = find_tree_run(run_id)
    if manifest is None:
        raise HTTPException(status_code=404, detail="任务集不存在")
    task = next((item for item in manifest.get("tasks", []) if item.get("task_id") == task_id), None)
    if task is None:
        raise HTTPException(status_code=404, detail="任务不在该任务集中")
    tree_path = (TREE_RUNS_DIR / run_id / str(task.get("tree_file", ""))).resolve()
    run_dir = (TREE_RUNS_DIR / run_id).resolve()
    try:
        tree_path.relative_to(run_dir)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="无效的树文件路径") from exc
    if not tree_path.is_file():
        raise HTTPException(status_code=404, detail="轨迹树文件缺失")
    tree = json.loads(tree_path.read_text(encoding="utf-8"))
    run_workbook = run_dir / str(manifest.get("quality_input_file", "rubric_trajectories.xlsx"))
    workbook = run_workbook if run_workbook.is_file() else WORKSPACE_DIR / "rubric_trajectories.xlsx"
    _attach_tree_observations(tree, _observation_index(workbook, task_id))
    return JSONResponse(tree)


@app.get("/api/assets/{relative_path:path}")
def get_asset(relative_path: str) -> FileResponse:
    try:
        path = resolve_image_asset(relative_path)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail="图片不存在") from exc
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return FileResponse(path, headers={"Cache-Control": "public, max-age=3600"})


@app.api_route(
    "/api/{unmatched_path:path}",
    methods=["GET", "POST", "PUT", "PATCH", "DELETE", "OPTIONS"],
    include_in_schema=False,
)
def unmatched_api(unmatched_path: str) -> None:
    raise HTTPException(status_code=404, detail=f"API 不存在：/api/{unmatched_path}")


if FRONTEND_DIST.is_dir():
    assets = FRONTEND_DIST / "assets"
    if assets.is_dir():
        app.mount("/assets", StaticFiles(directory=assets), name="frontend-assets")

    @app.get("/{full_path:path}", include_in_schema=False)
    def frontend(full_path: str) -> FileResponse:
        candidate = (FRONTEND_DIST / full_path).resolve()
        try:
            candidate.relative_to(FRONTEND_DIST.resolve())
        except ValueError:
            candidate = FRONTEND_DIST / "index.html"
        if candidate.is_file():
            return FileResponse(candidate)
        return FileResponse(FRONTEND_DIST / "index.html")
else:
    @app.get("/", include_in_schema=False)
    def api_root() -> dict[str, str]:
        return {"message": "Frontend has not been built. Run npm run build --prefix frontend."}
