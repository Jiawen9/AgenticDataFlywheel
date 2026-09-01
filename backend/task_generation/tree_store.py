"""Editable scenario tree and immutable, atomically published knowledge bundles.

Only current.json is replaced during publication. Readers and job snapshots pin
one immutable directory; legacy workbooks at the root are never overwritten.
"""
from __future__ import annotations

import copy
import hashlib
import json
import os
import shutil
import threading
import uuid
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

import pandas as pd
from openpyxl import Workbook, load_workbook

from .constants import KNOWLEDGE_BASE_DIR, KNOWLEDGE_BASE_FILES
from .knowledge_base import _clean, _truthy, parse_app_list, validate_workbook


TREE_FILE = "scene_tree.json"
META_SHEET = "_scene_tree_nodes"
SCENE_COLUMNS = ["scene", "capability", "sub_capability", "target_app", "use_resource_prior", "reference_example"]
KINDS = ("scene", "capability", "sub_capability")
_LOCK = threading.RLock()


class VersionConflict(ValueError):
    pass


@contextmanager
def write_lock(root: Path) -> Iterator[None]:
    """Serialize publication across threads and local server processes."""
    root.mkdir(parents=True, exist_ok=True)
    with _LOCK, (root / ".publication.lock").open("a+b") as stream:
        stream.seek(0, 2)
        if stream.tell() == 0:
            stream.write(b"0")
            stream.flush()
        stream.seek(0)
        if os.name == "nt":
            import msvcrt
            msvcrt.locking(stream.fileno(), msvcrt.LK_LOCK, 1)
        else:
            import fcntl
            fcntl.flock(stream.fileno(), fcntl.LOCK_EX)
        try:
            yield
        finally:
            stream.seek(0)
            if os.name == "nt":
                msvcrt.locking(stream.fileno(), msvcrt.LK_UNLCK, 1)
            else:
                fcntl.flock(stream.fileno(), fcntl.LOCK_UN)


def _json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _write_json(path: Path, value: Any) -> None:
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


def _pointed_root(root: Path) -> Path | None:
    pointer = root / "current.json"
    if not pointer.exists():
        return None
    version = _json(pointer)["version"]
    if not isinstance(version, str) or uuid.UUID(version).hex != version:
        raise ValueError("知识库当前版本指针无效")
    target = root / "versions" / version
    if not (target / TREE_FILE).is_file():
        raise ValueError("知识库当前版本不完整")
    return target


def flatten(scenes: list[dict[str, Any]]) -> list[tuple[dict[str, Any], tuple[str, str, str]]]:
    return [(leaf, (scene["label"], capability["label"], leaf["label"]))
            for scene in scenes for capability in scene.get("children", [])
            for leaf in capability.get("children", [])]


def validate_tree(scenes: Any) -> list[dict[str, Any]]:
    seen_ids: set[str] = set()

    def walk(nodes: Any, depth: int) -> list[dict[str, Any]]:
        if not isinstance(nodes, list):
            raise ValueError("场景树节点必须是数组")
        result = []
        names: set[str] = set()
        for value in nodes:
            if not isinstance(value, dict) or value.get("kind") != KINDS[depth]:
                raise ValueError("场景树必须按场景、一级能力、任务类型三级组织")
            name = value.get("label")
            if not isinstance(name, str) or not name.strip() or len(name.strip()) > 200:
                raise ValueError("节点名称不能为空且不能超过 200 字")
            name = name.strip()
            if name in names:
                raise ValueError(f"同级节点名称重复：{name}")
            names.add(name)
            try:
                identifier = str(uuid.UUID(value.get("id", "")))
            except (ValueError, TypeError, AttributeError) as exc:
                raise ValueError("节点 ID 必须是 UUID") from exc
            if identifier in seen_ids:
                raise ValueError("节点 ID 不能重复")
            seen_ids.add(identifier)
            node = {"id": identifier, "kind": KINDS[depth], "label": name}
            if depth < 2:
                if value.get("app_configs"):
                    raise ValueError("只能在任务类型节点配置 App")
                node["children"] = walk(value.get("children", []), depth + 1)
            else:
                if value.get("children"):
                    raise ValueError("任务类型不能包含子节点")
                configs = value.get("app_configs", [])
                if not isinstance(configs, list):
                    raise ValueError("App 配置必须是数组")
                apps: set[str] = set()
                node["app_configs"] = []
                for config in configs:
                    if not isinstance(config, dict):
                        raise ValueError("App 配置格式错误")
                    app = config.get("app")
                    example = config.get("reference_example", "")
                    resource = config.get("use_resource_prior", False)
                    if not isinstance(app, str) or not app.strip() or len(app.strip()) > 200:
                        raise ValueError("App 名称不能为空且不能超过 200 字")
                    app = app.strip()
                    if app in apps:
                        raise ValueError(f"任务类型 {name} 的 App 重复：{app}")
                    if not isinstance(example, str) or len(example) > 20000 or not isinstance(resource, bool):
                        raise ValueError("参考示例必须是文本（最多 20000 字），资源开关必须是布尔值")
                    apps.add(app)
                    node["app_configs"].append({"app": app, "reference_example": example, "use_resource_prior": resource})
            result.append(node)
        return result

    return walk(scenes, 0)


def _paths(scenes: list[dict[str, Any]]) -> dict[tuple[str, ...], dict[str, Any]]:
    found: dict[tuple[str, ...], dict[str, Any]] = {}

    def visit(nodes: list[dict[str, Any]], prefix: tuple[str, ...]) -> None:
        for node in nodes:
            path = (*prefix, node["label"])
            found[path] = node
            visit(node.get("children", []), path)
    visit(scenes, ())
    return found


def import_scene_workbook(path: Path, previous: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    validate_workbook(path, "scene_tree")
    previous_paths = _paths(previous or [])
    scenes: list[dict[str, Any]] = []
    # The optional metadata sheet makes empty branches and UUIDs round-trip.
    with pd.ExcelFile(path) as book:
        frame = pd.read_excel(book, sheet_name=0)
        if META_SHEET in book.sheet_names:
            metadata = pd.read_excel(book, sheet_name=META_SHEET).fillna("")
            required = {"id", "parent_id", "kind", "label"}
            if not required.issubset(metadata.columns):
                raise ValueError("场景树身份元数据缺少必要列")
            by_id = {}
            for _, row in metadata.iterrows():
                identifier = _clean(row["id"])
                if identifier in by_id:
                    raise ValueError("场景树身份元数据 ID 重复")
                by_id[identifier] = {"id": identifier, "label": _clean(row["label"]), "kind": row["kind"], "children": [], "app_configs": []}
            for _, row in metadata.iterrows():
                node = by_id[_clean(row["id"])]
                parent = _clean(row["parent_id"])
                if parent:
                    if parent not in by_id:
                        raise ValueError("场景树身份元数据父节点不存在")
                    by_id[parent]["children"].append(node)
                else:
                    scenes.append(node)
            scenes = validate_tree(scenes)
            if len(_paths(scenes)) != len(by_id):
                raise ValueError("场景树身份元数据含有孤立或循环节点")
    path_index = _paths(scenes)
    for row_number, (_, row) in enumerate(frame.iterrows(), start=2):
        if all(not _clean(row.get(key)) for key in SCENE_COLUMNS):
            continue
        parts = tuple(_clean(row.get(key)) for key in SCENE_COLUMNS[:3])
        if not all(parts):
            raise ValueError(f"场景树第 {row_number} 行的场景、能力和任务类型不能为空")
        parent_children = scenes
        for depth in range(3):
            key = parts[:depth + 1]
            node = path_index.get(key)
            if node is None:
                node = {"id": previous_paths.get(key, {}).get("id", str(uuid.uuid4())), "label": parts[depth], "kind": KINDS[depth]}
                node["app_configs" if depth == 2 else "children"] = []
                parent_children.append(node)
                path_index[key] = node
            if depth < 2:
                parent_children = node["children"]
        for app in dict.fromkeys(parse_app_list(row.get("target_app"))):
            config = {"app": app, "reference_example": _clean(row.get("reference_example")), "use_resource_prior": _truthy(row.get("use_resource_prior"))}
            existing = next((item for item in node["app_configs"] if item["app"] == app), None)
            if existing and existing != config:
                raise ValueError(f"场景树第 {row_number} 行与同任务类型/App 的配置冲突：{app}；请先合并冲突行")
            if not existing:
                node["app_configs"].append(config)
    return validate_tree(scenes)


def _string_cell(sheet: Any, row: int, column: int, value: Any) -> None:
    cell = sheet.cell(row, column, value)
    if isinstance(value, str):
        cell.data_type = "s"  # User examples and labels must not become Excel formulas.


def write_scene_workbook(path: Path, scenes: list[dict[str, Any]]) -> None:
    book = load_workbook(path) if path.is_file() else Workbook()
    try:
        old = book.worksheets[0]
        title = old.title
        book.remove(old)
        sheet = book.create_sheet(title, 0)
        for column, label in enumerate(SCENE_COLUMNS, 1):
            _string_cell(sheet, 1, column, label)
        row_number = 2
        for leaf, path_labels in flatten(scenes):
            for config in leaf["app_configs"]:
                values = [*path_labels, json.dumps([config["app"]], ensure_ascii=False), config["use_resource_prior"], config["reference_example"]]
                for column, value in enumerate(values, 1):
                    _string_cell(sheet, row_number, column, value)
                row_number += 1
        if META_SHEET in book.sheetnames:
            book.remove(book[META_SHEET])
        meta = book.create_sheet(META_SHEET)
        meta.append(["id", "parent_id", "kind", "label"])

        def visit(nodes: list[dict[str, Any]], parent: str = "") -> None:
            for node in nodes:
                index = meta.max_row + 1
                for column, value in enumerate([node["id"], parent, node["kind"], node["label"]], 1):
                    _string_cell(meta, index, column, value)
                visit(node.get("children", []), node["id"])
        visit(scenes)
        meta.sheet_state = "hidden"
        book.save(path)
    finally:
        book.close()


def _rename_priors(path: Path, before: list[dict[str, Any]], after: list[dict[str, Any]]) -> None:
    if not path.is_file():
        return
    old = {leaf["id"]: labels for leaf, labels in flatten(before)}
    mapping = {old[leaf["id"]]: labels for leaf, labels in flatten(after) if leaf["id"] in old and old[leaf["id"]] != labels}
    if not mapping:
        return
    book = load_workbook(path)
    try:
        sheet = book.worksheets[0]
        columns = {cell.value: cell.column for cell in sheet[1]}
        if not set(SCENE_COLUMNS[:3]).issubset(columns):
            raise ValueError("操控先验列不完整，无法同步重命名")
        for row in range(2, sheet.max_row + 1):
            original = tuple(_clean(sheet.cell(row, columns[key]).value) for key in SCENE_COLUMNS[:3])
            if original in mapping:
                for key, value in zip(SCENE_COLUMNS[:3], mapping[original]):
                    _string_cell(sheet, row, columns[key], value)
        book.save(path)
    finally:
        book.close()


def _publish(root: Path, source: Path, scenes: list[dict[str, Any]] | None = None,
             replacement: tuple[str, Path] | None = None) -> Path:
    """Caller holds write_lock. No live file changes before the last replace."""
    version = uuid.uuid4().hex
    versions = root / "versions"
    versions.mkdir(exist_ok=True)
    staging = versions / f".staging-{version}"
    staging.mkdir()
    try:
        for filename in KNOWLEDGE_BASE_FILES.values():
            if (source / filename).is_file():
                shutil.copy2(source / filename, staging / filename)
        previous = _json(source / TREE_FILE)["scenes"] if (source / TREE_FILE).exists() else []
        if replacement:
            kind, uploaded = replacement
            shutil.copy2(uploaded, staging / KNOWLEDGE_BASE_FILES[kind])
        scene_path = staging / KNOWLEDGE_BASE_FILES["scene_tree"]
        if scenes is None:
            if previous and not (replacement and replacement[0] == "scene_tree"):
                scenes = previous
            else:
                scenes = import_scene_workbook(scene_path, previous) if scene_path.exists() else []
        scenes = validate_tree(scenes)
        _rename_priors(staging / KNOWLEDGE_BASE_FILES["control_prior"], previous, scenes)
        if scene_path.exists() or not replacement:
            write_scene_workbook(scene_path, scenes)
        _write_json(staging / TREE_FILE, {"version": version, "created_at": datetime.now(timezone.utc).isoformat(), "scenes": scenes})
        published = versions / version
        staging.rename(published)
        pointer = root / f".current-{version}.tmp"
        _write_json(pointer, {"version": version})
        pointer.replace(root / "current.json")
        return published
    finally:
        if staging.is_dir():
            shutil.rmtree(staging)


def current_root(root: Path = KNOWLEDGE_BASE_DIR) -> Path:
    root = Path(root)
    if (root / TREE_FILE).is_file():  # Immutable version or a job snapshot.
        return root
    current = _pointed_root(root)
    if current:
        return current
    with write_lock(root):
        return _pointed_root(root) or _publish(root, root)


def read_tree(root: Path = KNOWLEDGE_BASE_DIR) -> dict[str, Any]:
    return _json(current_root(root) / TREE_FILE)


def save_tree(scenes: Any, base_version: str, *, root: Path = KNOWLEDGE_BASE_DIR) -> dict[str, Any]:
    clean = validate_tree(scenes)
    current_root(root)
    with write_lock(root):
        source = _pointed_root(root)
        if source is None or source.name != base_version:
            raise VersionConflict("知识库已更新，请保留草稿并刷新最新版本后重试")
        previous = _json(source / TREE_FILE)["scenes"]
        old_locations = {node["id"]: (node["kind"], tuple(p["id"] for p in _ancestors(previous, path))) for path, node in _paths(previous).items()}
        for path, node in _paths(clean).items():
            if node["id"] in old_locations:
                location = (node["kind"], tuple(p["id"] for p in _ancestors(clean, path)))
                if location != old_locations[node["id"]]:
                    raise ValueError("第一版不支持移动现有节点或改变节点层级")
        published = _publish(root, source, clean)
    return tree_payload(published)


def _ancestors(scenes: list[dict[str, Any]], path: tuple[str, ...]) -> list[dict[str, Any]]:
    index = _paths(scenes)
    return [index[path[:depth]] for depth in range(1, len(path))]


def replace_workbook(kind: str, source: Path, *, root: Path = KNOWLEDGE_BASE_DIR, base_version: str | None = None) -> Path:
    if kind not in KNOWLEDGE_BASE_FILES:
        raise ValueError(f"未知知识库类型：{kind}")
    validate_workbook(source, kind)
    with write_lock(root):
        current = _pointed_root(root)
        if base_version is not None and (current is None or current.name != base_version):
            raise VersionConflict("知识库已更新，请刷新后重新上传")
        return _publish(root, current or root, replacement=(kind, source))


def prior_status(root: Path) -> tuple[dict[tuple[str, str, str, str], str], dict[str, int], list[str]]:
    controls: dict[tuple[str, str, str, str], list[str]] = {}
    resources: dict[str, int] = {}
    warnings: list[str] = []
    try:
        frame = pd.read_excel(root / KNOWLEDGE_BASE_FILES["control_prior"])
        for _, row in frame.iterrows():
            for app in parse_app_list(row.get("target_app")):
                key = (*(_clean(row.get(part)) for part in SCENE_COLUMNS[:3]), app)
                desc = _clean(row.get("sub_capability_desc"))
                if desc and desc not in controls.setdefault(key, []):
                    controls[key].append(desc)
    except (OSError, ValueError) as exc:
        warnings.append(f"操控先验不可读：{exc}")
    try:
        with pd.ExcelFile(root / KNOWLEDGE_BASE_FILES["resource_prior"]) as book:
            resources = {sheet: len(pd.read_excel(book, sheet_name=sheet)) for sheet in book.sheet_names}
    except (OSError, ValueError) as exc:
        warnings.append(f"资源先验不可读：{exc}")
    return {key: "\n".join(values) for key, values in controls.items()}, resources, warnings


def tree_payload(root: Path = KNOWLEDGE_BASE_DIR) -> dict[str, Any]:
    source = current_root(root)
    value = copy.deepcopy(_json(source / TREE_FILE))
    controls, resources, warnings = prior_status(source)
    leaves = flatten(value["scenes"])
    for leaf, labels in leaves:
        leaf.update(dict(zip(SCENE_COLUMNS[:3], labels)))
        leaf["generatable"] = bool(leaf["app_configs"])
        for config in leaf["app_configs"]:
            config["control_prior_available"] = bool(controls.get((*labels, config["app"])))
            config["resource_count"] = resources.get(config["app"], 0)
    value.update({"leaf_count": len(leaves), "execution_unit_count": sum(len(leaf["app_configs"]) for leaf, _ in leaves), "warnings": warnings})
    return value


def snapshot(destination: Path, *, root: Path = KNOWLEDGE_BASE_DIR, version: str | None = None) -> dict[str, Any]:
    source = current_root(root)
    value = _json(source / TREE_FILE)
    if version is not None and value["version"] != version:
        raise VersionConflict("知识库已更新，请刷新场景树后重新提交")
    # A pinned immutable source remains valid even if another edit publishes now.
    for kind, filename in KNOWLEDGE_BASE_FILES.items():
        validate_workbook(source / filename, kind)
    destination.mkdir(parents=True, exist_ok=True)
    files = []
    for filename in [*KNOWLEDGE_BASE_FILES.values(), TREE_FILE]:
        target = destination / filename
        shutil.copy2(source / filename, target)
        files.append({"filename": filename, "sha256": hashlib.sha256(target.read_bytes()).hexdigest()})
    return {"directory": str(destination), "version": value["version"], "files": files}
