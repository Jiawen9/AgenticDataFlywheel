# -*- coding: utf-8 -*-
"""Build an action-prefix tree while ignoring only proven short insertions.

Every trajectory step is classified, but only one/two-step transient runs with
two high-confidence normal recovery steps are ignored. Low-confidence steps
may additionally use action/bbox realignment and Qwen screenshot confirmation.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Protocol

from openpyxl import load_workbook

try:
    from ..trajectories_preprocessing import configure_reviewer_environment
    from .intermediate_state_classifier import (
        IntermediateStateResult,
        QwenIntermediateStateClassifier,
    )
    from .state_alignment_reviewer import (
        QwenStateAlignmentReviewer,
        StateAlignmentResult,
    )
except ImportError:  # Support direct execution during local debugging.
    BACKEND_IMPORT_ROOT = Path(__file__).resolve().parents[1]
    if str(BACKEND_IMPORT_ROOT) not in sys.path:
        sys.path.insert(0, str(BACKEND_IMPORT_ROOT))
    from trajectories_preprocessing import configure_reviewer_environment
    from trajectories_tree.intermediate_state_classifier import (
        IntermediateStateResult,
        QwenIntermediateStateClassifier,
    )
    from trajectories_tree.state_alignment_reviewer import (
        QwenStateAlignmentReviewer,
        StateAlignmentResult,
    )


FULL_SCORE = 3.5
BOX_OVERLAP_THRESHOLD = 0.5
MAX_INCIDENTAL_SKIP = 2
RESYNC_CONFIRM_STEPS = 2
BBOX_RE = re.compile(
    r"<bbox>\[\s*(-?\d+)\s*,\s*(-?\d+)\s*,\s*(-?\d+)\s*,\s*(-?\d+)\s*]\s*</bbox>"
)
BOX_DIRECTION_RE = re.compile(r"\bdirection\s*=\s*(left|right|up|down)\b", re.IGNORECASE)
STEP_NUMBER_RE = re.compile(r"step(\d+)", re.IGNORECASE)

BACKEND_DIR = Path(__file__).resolve().parents[1]
PROJECT_ROOT = BACKEND_DIR.parent
WORKSPACE_DIR = PROJECT_ROOT / "backend_workspace"
DEFAULT_XLSX = WORKSPACE_DIR / "annotated_trajectories.xlsx"
DEFAULT_TRAJECTORY_ROOT = WORKSPACE_DIR / "rollout_trajectories"
DEFAULT_OUTPUT = WORKSPACE_DIR / "trajectory_tree.json"
DEFAULT_ENV = BACKEND_DIR / ".env"
DEFAULT_CLASSIFICATION_CACHE = (
    BACKEND_DIR / "trajectories_tree" / "qwen_intermediate_state_cache.json"
)
DEFAULT_ALIGNMENT_CACHE = (
    BACKEND_DIR / "trajectories_tree" / "qwen_state_alignment_cache.json"
)


class StepClassifier(Protocol):
    model: str

    def classify(self, **kwargs: Any) -> IntermediateStateResult: ...


class AlignmentReviewer(Protocol):
    model: str

    def review(self, **kwargs: Any) -> StateAlignmentResult: ...


def parse_action(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if not isinstance(value, str):
        return {}
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        return {}


def parse_bbox(value: Any) -> tuple[int, int, int, int] | None:
    if not isinstance(value, str):
        return None
    match = BBOX_RE.search(value)
    if not match:
        return None
    xa, ya, xb, yb = map(int, match.groups())
    x0, x1 = sorted((xa, xb))
    y0, y1 = sorted((ya, yb))
    if x0 == x1 or y0 == y1:
        return None
    return x0, y0, x1, y1


def direction(action: dict[str, Any]) -> str | None:
    start, end = action.get("start_coordinate"), action.get("end_coordinate")
    if (
        not isinstance(start, list)
        or not isinstance(end, list)
        or len(start) < 2
        or len(end) < 2
    ):
        return None
    try:
        dx, dy = float(end[0]) - float(start[0]), float(end[1]) - float(start[1])
    except (TypeError, ValueError):
        return None
    if abs(dx) < abs(dy):
        return "up" if dy < 0 else "down"
    return "left" if dx < 0 else "right"


def box_direction(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    match = BOX_DIRECTION_RE.search(value)
    return match.group(1).lower() if match else None


def point_in_bbox(point: Any, bbox: tuple[int, int, int, int] | None) -> bool:
    if bbox is None or not isinstance(point, list) or len(point) < 2:
        return False
    try:
        x, y = float(point[0]), float(point[1])
    except (TypeError, ValueError):
        return False
    x0, y0, x1, y1 = bbox
    return x0 <= x <= x1 and y0 <= y <= y1


def bbox_overlap(
    first: tuple[int, int, int, int],
    second: tuple[int, int, int, int],
) -> float:
    """Return intersection over the smaller box area."""
    ax0, ay0, ax1, ay1 = first
    bx0, by0, bx1, by1 = second
    width = max(0, min(ax1, bx1) - max(ax0, bx0))
    height = max(0, min(ay1, by1) - max(ay0, by0))
    intersection = width * height
    smaller_area = min((ax1 - ax0) * (ay1 - ay0), (bx1 - bx0) * (by1 - by0))
    return intersection / smaller_area if smaller_area > 0 else 0.0


@dataclass
class Step:
    trajectory: str
    step_index: int
    excel_row: int
    image: str
    xml: str
    action_text: str
    action: dict[str, Any]
    summary: str
    actions_box: str
    classification: IntermediateStateResult | None = None
    classification_candidate: bool = False
    effective_intermediate: bool = False
    uncertain: bool = False
    counted_in_tree: bool = True
    tree_decision: str = ""
    decision_source: str = ""
    tree_node_id: int | None = None
    tree_score: float = 0.0
    skip_count: int = 0
    confirmation_node_ids: list[int] = field(default_factory=list)
    confirmation_step_indices: list[int] = field(default_factory=list)
    alignment_review: dict[str, Any] | None = None
    observation: str = ""
    excluded_intermediate_terminate: bool = False

    def apply_classification(
        self,
        result: IntermediateStateResult,
        confidence_threshold: float,
    ) -> None:
        self.classification = result
        self.observation = result.observation
        self.classification_candidate = bool(
            result.is_intermediate and result.confidence >= confidence_threshold
        )
        self.uncertain = result.confidence < confidence_threshold

    def audit_dict(self, confidence_threshold: float) -> dict[str, Any]:
        classification_value = None
        if self.classification is not None:
            classification_value = self.classification.to_dict(confidence_threshold)
            # The classifier identifies a candidate; the bounded seed policy
            # decides whether that candidate is actually ignored.
            classification_value["effective_intermediate"] = self.effective_intermediate
            classification_value["policy_candidate"] = self.classification_candidate
        return {
            "step": self.step_index,
            "excel_row": self.excel_row,
            "image": self.image,
            "xml": self.xml,
            "action_text": self.action_text,
            "action": self.action,
            "summary": self.summary,
            "actions_box": self.actions_box,
            "observation": self.observation,
            "excluded_intermediate_terminate": self.excluded_intermediate_terminate,
            "classification": classification_value,
            "classification_candidate": self.classification_candidate,
            "alignment_review": self.alignment_review,
            "counted_in_tree": self.counted_in_tree,
            "decision": self.tree_decision,
            "decision_source": self.decision_source,
            "node_id": self.tree_node_id,
            "score": self.tree_score,
            "skip_count": self.skip_count,
            "confirmation_node_ids": self.confirmation_node_ids,
            "confirmation_step_indices": self.confirmation_step_indices,
        }


def score_step(candidate: Step, reference: Step) -> float:
    """Score one action against a node reference using 3.5-point semantics."""
    kind = candidate.action.get("action")
    if not kind or kind != reference.action.get("action"):
        return 0.0

    candidate_box = parse_bbox(candidate.actions_box)
    reference_box = parse_bbox(reference.actions_box)
    if kind in {"click", "long_press"}:
        if candidate_box is not None and reference_box is not None:
            return (
                FULL_SCORE
                if bbox_overlap(candidate_box, reference_box) >= BOX_OVERLAP_THRESHOLD
                else 0.5
            )
        return (
            FULL_SCORE
            if point_in_bbox(candidate.action.get("coordinate"), reference_box)
            else 0.5
        )

    if kind == "swipe":
        candidate_direction = box_direction(candidate.actions_box) or direction(candidate.action)
        reference_direction = box_direction(reference.actions_box) or direction(reference.action)
        same_direction = bool(candidate_direction and candidate_direction == reference_direction)
        if candidate_box is not None and reference_box is not None:
            same_box = bbox_overlap(candidate_box, reference_box) >= BOX_OVERLAP_THRESHOLD
            return FULL_SCORE if same_box and same_direction else 0.5
        hit = point_in_bbox(candidate.action.get("start_coordinate"), reference_box)
        return FULL_SCORE if hit and same_direction else 0.5

    if kind == "type":
        return FULL_SCORE if candidate.action.get("text") == reference.action.get("text") else 2.5
    if kind == "open":
        return FULL_SCORE if candidate.action.get("text") == reference.action.get("text") else 3.0
    if kind == "system_button":
        return FULL_SCORE if candidate.action.get("button") == reference.action.get("button") else 0.0
    if kind == "terminate":
        return FULL_SCORE if candidate.action.get("status") == reference.action.get("status") else 3.0
    if kind == "wait":
        return FULL_SCORE
    return FULL_SCORE if candidate.action == reference.action else 0.0


@dataclass
class Node:
    id: int
    depth: int
    reference: Step | None = None
    children: list["Node"] = field(default_factory=list)
    occurrences: list[dict[str, Any]] = field(default_factory=list)
    terminal_trajectories: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        if self.reference is None:
            first_occurrence = self.occurrences[0] if self.occurrences else {}
            label, action, summary, box, image, xml = (
                "桌面",
                {"action": "desktop"},
                "所有轨迹的共同起点：设备桌面",
                "",
                str(first_occurrence.get("image", "")),
                str(first_occurrence.get("xml", "")),
            )
        else:
            label = self.reference.action.get("action", "unknown")
            action = self.reference.action
            summary = self.reference.summary
            box = self.reference.actions_box
            image = self.reference.image
            xml = self.reference.xml
        return {
            "id": self.id,
            "depth": self.depth,
            "label": label,
            "action": action,
            "summary": summary,
            "observation": self.reference.observation if self.reference else "",
            "actions_box": box,
            "image": image,
            "xml": xml,
            "reference_trajectory": (
                self.reference.trajectory
                if self.reference
                else str(first_occurrence.get("trajectory", ""))
            ),
            "reference_step": self.reference.step_index if self.reference else 0,
            "occurrence_count": len(self.occurrences),
            "occurrences": self.occurrences,
            "terminal_trajectories": self.terminal_trajectories,
            "children": [child.to_dict() for child in self.children],
        }


@dataclass
class ResyncCandidate:
    skip_count: int
    confirmation_steps: list[Step]
    confirmation_nodes: list[Node]


@dataclass
class BuildStatistics:
    resync_candidate_count: int = 0
    resync_review_count: int = 0
    resync_review_cache_hit_count: int = 0
    resync_accepted_count: int = 0
    resync_rejected_count: int = 0
    resync_ignored_step_count: int = 0

    def to_dict(self) -> dict[str, int]:
        return {
            "resync_candidate_count": self.resync_candidate_count,
            "resync_review_count": self.resync_review_count,
            "resync_review_cache_hit_count": self.resync_review_cache_hit_count,
            "resync_accepted_count": self.resync_accepted_count,
            "resync_rejected_count": self.resync_rejected_count,
            "resync_ignored_step_count": self.resync_ignored_step_count,
        }


def _step_number(image: str, fallback: int) -> int:
    match = STEP_NUMBER_RE.search(Path(image).name)
    return int(match.group(1)) if match else fallback


def load_trajectories(path: Path, sheet_name: str | None) -> list[tuple[str, list[Step]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        headers = {
            str(cell.value).strip(): cell.column
            for cell in sheet[1]
            if cell.value is not None
        }
        required = {"文件夹名", "image", "xml", "action", "summary", "actions_box"}
        missing = sorted(required - headers.keys())
        if missing:
            raise ValueError(f"Excel 缺少必要列: {', '.join(missing)}")

        grouped: dict[str, list[Step]] = {}
        for row in range(2, sheet.max_row + 1):
            raw_trajectory = sheet.cell(row, headers["文件夹名"]).value
            if raw_trajectory is None or not str(raw_trajectory).strip():
                continue
            trajectory = str(raw_trajectory).strip()
            steps = grouped.setdefault(trajectory, [])
            image = str(sheet.cell(row, headers["image"]).value or "").strip()
            action_text = str(sheet.cell(row, headers["action"]).value or "")
            steps.append(
                Step(
                    trajectory=trajectory,
                    step_index=_step_number(image, len(steps) + 1),
                    excel_row=row,
                    image=image,
                    xml=str(sheet.cell(row, headers["xml"]).value or "").strip(),
                    action_text=action_text,
                    action=parse_action(action_text),
                    summary=str(sheet.cell(row, headers["summary"]).value or ""),
                    actions_box=str(sheet.cell(row, headers["actions_box"]).value or ""),
                )
            )
    finally:
        workbook.close()

    for trajectory, steps in grouped.items():
        steps.sort(key=lambda step: (step.step_index, step.excel_row))
        indices = [step.step_index for step in steps]
        if len(indices) != len(set(indices)):
            raise ValueError(f"轨迹 {trajectory} 包含重复 step 编号")
    return sorted(grouped.items(), key=lambda item: item[0])


def _artifact_path(root: Path, value: str, *, require_file: bool = True) -> Path:
    path = Path(value)
    resolved = path.resolve() if path.is_absolute() else (root / path).resolve()
    if require_file and not resolved.is_file():
        raise FileNotFoundError(f"轨迹资源不存在: {value} -> {resolved}")
    return resolved


def classify_trajectories(
    trajectories: list[tuple[str, list[Step]]],
    classifier: StepClassifier,
    trajectory_root: Path,
    confidence_threshold: float,
    task: str = "",
) -> None:
    total = sum(
        1
        for _, steps in trajectories
        for position, step in enumerate(steps)
        if not (step.action.get("action") == "terminate" and position < len(steps) - 1)
    )
    completed = 0
    for trajectory, steps in trajectories:
        if steps and steps[-1].action.get("action") == "terminate":
            status = steps[-1].action.get("status")
            if status not in {"success", "failure"}:
                raise ValueError(
                    f"final terminate must use status success or failure: "
                    f"trajectory={trajectory}, step={steps[-1].step_index}, status={status!r}"
                )
        for position, step in enumerate(steps):
            if step.action.get("action") == "terminate" and position < len(steps) - 1:
                step.excluded_intermediate_terminate = True
                step.counted_in_tree = False
                step.tree_decision = "ignore_intermediate"
                step.decision_source = "intermediate_terminate"
                continue
            previous_summary = steps[position - 1].summary if position > 0 else ""
            next_step = steps[position + 1] if position + 1 < len(steps) else None
            current_path = _artifact_path(trajectory_root, step.image)
            done_name = re.sub(r"_input(?:_stability)?\.jpg$", "_done.jpg", current_path.name, flags=re.IGNORECASE)
            done_path = current_path.with_name(done_name)
            try:
                result = classifier.classify(
                    trajectory=trajectory,
                    step_index=step.step_index,
                    current_image_path=current_path,
                    next_image_path=(
                        _artifact_path(trajectory_root, next_step.image)
                        if next_step is not None
                        else None
                    ),
                    action=step.action,
                    summary=step.summary,
                    previous_summary=previous_summary,
                    next_summary=next_step.summary if next_step is not None else "",
                    after_image_path=done_path if done_path.is_file() else None,
                    task=task,
                )
            except Exception as exc:
                raise RuntimeError(
                    f"中间状态分类失败: trajectory={trajectory}, "
                    f"step={step.step_index}, error={exc}"
                ) from exc
            step.apply_classification(result, confidence_threshold)
            completed += 1
            print(
                f"Classified {completed}/{total}: {trajectory} "
                f"step{step.step_index:03d}; cached={result.cached}; "
                f"intermediate={result.is_intermediate}; confidence={result.confidence:.3f}"
            )


def apply_bounded_skip_policy(
    steps: list[Step],
    *,
    confidence_threshold: float = 0.8,
    max_skip: int = MAX_INCIDENTAL_SKIP,
    confirmation_steps: int = RESYNC_CONFIRM_STEPS,
) -> None:
    """Ignore only short classified runs followed by stable recovery steps."""
    for step in steps:
        step.effective_intermediate = False
        step.counted_in_tree = not step.excluded_intermediate_terminate
        step.skip_count = 0
        step.confirmation_step_indices = []

    def protected_task_action(position: int, step: Step) -> bool:
        """Never erase actions that establish the trajectory's common task start."""
        kind = str(step.action.get("action", "")).lower()
        summary = step.summary.lower()
        launch_words = ("打开", "启动", "open", "launch")
        app_words = ("应用", "app", "爱奇艺")
        return (
            kind == "open"
            or (any(word in summary for word in launch_words) and any(word in summary for word in app_words))
        )

    position = 0
    while position < len(steps):
        if steps[position].excluded_intermediate_terminate:
            position += 1
            continue
        if not steps[position].classification_candidate:
            position += 1
            continue
        end = position
        while end < len(steps) and steps[end].classification_candidate:
            end += 1
        run_length = end - position
        confirmation = steps[end : end + confirmation_steps]
        has_stable_recovery = bool(
            len(confirmation) == confirmation_steps
            and all(
                item.classification is not None
                and not item.classification.is_intermediate
                and item.classification.confidence >= confidence_threshold
                for item in confirmation
            )
        )
        if run_length <= max_skip and has_stable_recovery:
            confirmation_indices = [item.step_index for item in confirmation]
            ignored_items = [
                item for offset, item in enumerate(steps[position:end], position)
                if not protected_task_action(offset, item)
            ]
            for item in ignored_items:
                item.effective_intermediate = True
                item.counted_in_tree = False
                item.skip_count = len(ignored_items)
                item.confirmation_step_indices = confirmation_indices
        position = end


def best_matching_child(parent: Node, step: Step) -> tuple[float, Node | None]:
    scored = [
        (score_step(step, child.reference), child)
        for child in parent.children
        if child.reference is not None
    ]
    return max(scored, key=lambda item: (item[0], -item[1].id), default=(0.0, None))


def find_resync_candidate(
    parent: Node,
    steps: list[Step],
    position: int,
    *,
    max_skip: int = MAX_INCIDENTAL_SKIP,
    confirmation_count: int = RESYNC_CONFIRM_STEPS,
) -> ResyncCandidate | None:
    """Find an uncertain one/two-step gap followed by two full tree matches."""
    for skip_count in range(1, max_skip + 1):
        confirmation_start = position + skip_count
        confirmation_end = confirmation_start + confirmation_count
        if confirmation_end > len(steps):
            continue
        skipped_steps = steps[position:confirmation_start]
        if not skipped_steps or not all(step.uncertain for step in skipped_steps):
            continue
        cursor = parent
        nodes: list[Node] = []
        candidates = steps[confirmation_start:confirmation_end]
        for candidate_step in candidates:
            score, node = best_matching_child(cursor, candidate_step)
            if node is None or score < FULL_SCORE:
                nodes = []
                break
            nodes.append(node)
            cursor = node
        if len(nodes) == confirmation_count:
            return ResyncCandidate(
                skip_count=skip_count,
                confirmation_steps=candidates,
                confirmation_nodes=nodes,
            )
    return None


def _classification_dict(step: Step, confidence_threshold: float) -> dict[str, Any] | None:
    if step.classification is None:
        return None
    value = step.classification.to_dict(confidence_threshold)
    value["effective_intermediate"] = step.effective_intermediate
    value["policy_candidate"] = step.classification_candidate
    return value


def _step_context(step: Step) -> dict[str, Any]:
    return {
        "trajectory": step.trajectory,
        "step": step.step_index,
        "action": step.action,
        "summary": step.summary,
        "actions_box": step.actions_box,
    }


def _record_ignored_step(
    decisions: list[dict[str, Any]],
    step: Step,
    *,
    parent_node: int,
    decision_source: str,
    reason: str,
    confidence_threshold: float,
) -> None:
    step.counted_in_tree = False
    step.tree_decision = "ignore_intermediate"
    step.decision_source = decision_source
    step.tree_node_id = None
    step.tree_score = 0.0
    decisions.append(
        {
            "trajectory": step.trajectory,
            "step": step.step_index,
            "excel_row": step.excel_row,
            "parent_node": parent_node,
            "node": None,
            "score": 0.0,
            "decision": "ignore_intermediate",
            "decision_source": decision_source,
            "reason": reason,
            "skip_count": step.skip_count,
            "confirmation_node_ids": step.confirmation_node_ids,
            "confirmation_step_indices": step.confirmation_step_indices,
            "classification": _classification_dict(step, confidence_threshold),
            "alignment_review": step.alignment_review,
            "action": step.action,
            "summary": step.summary,
            "image": step.image,
            "xml": step.xml,
        }
    )


def build_tree(
    trajectories: list[tuple[str, list[Step]]],
    *,
    confidence_threshold: float,
    trajectory_root: Path,
    alignment_reviewer: AlignmentReviewer | None,
    max_incidental_skip: int = MAX_INCIDENTAL_SKIP,
) -> tuple[Node, list[dict[str, Any]], BuildStatistics]:
    root = Node(id=0, depth=0)
    next_id = 1
    decisions: list[dict[str, Any]] = []
    statistics = BuildStatistics()

    for trajectory, steps in trajectories:
        if steps:
            initial_image = str(Path(steps[0].image).with_name("initial_orch.jpg"))
            initial_xml = str(Path(steps[0].xml).with_name("initial_orch_ui.xml"))
            if not _artifact_path(
                trajectory_root, initial_image, require_file=False
            ).is_file():
                initial_image = steps[0].image
            if not _artifact_path(
                trajectory_root, initial_xml, require_file=False
            ).is_file():
                initial_xml = steps[0].xml
            root.occurrences.append(
                {
                    "trajectory": trajectory,
                    "step": 0,
                    "excel_row": steps[0].excel_row,
                    "image": initial_image,
                    "xml": initial_xml,
                    "action": {"action": "desktop"},
                    "action_text": "",
                    "summary": "轨迹起始桌面",
                    "observation": "",
                    "actions_box": "",
                    "score": FULL_SCORE,
                    "reused": len(root.occurrences) > 0,
                    "classification": None,
                    "alignment_review": None,
                }
            )
        current = root
        position = 0
        while position < len(steps):
            step = steps[position]
            parent_id = current.id

            if step.excluded_intermediate_terminate:
                _record_ignored_step(
                    decisions,
                    step,
                    parent_node=parent_id,
                    decision_source="intermediate_terminate",
                    reason="non-final terminate is excluded before model classification",
                    confidence_threshold=confidence_threshold,
                )
                position += 1
                continue

            if step.effective_intermediate:
                _record_ignored_step(
                    decisions,
                    step,
                    parent_node=parent_id,
                    decision_source="bounded_classification",
                    reason=(
                        "Qwen classification found a one/two-step transient run "
                        "followed by two high-confidence normal recovery steps"
                    ),
                    confidence_threshold=confidence_threshold,
                )
                position += 1
                continue

            best_score, matched = best_matching_child(current, step)
            reused = matched is not None and best_score >= FULL_SCORE

            if (
                not reused
                and alignment_reviewer is not None
            ):
                candidate = find_resync_candidate(
                    current,
                    steps,
                    position,
                    max_skip=max_incidental_skip,
                )
                if candidate is not None:
                    statistics.resync_candidate_count += 1
                    skipped = steps[position : position + candidate.skip_count]
                    try:
                        review = alignment_reviewer.review(
                            trajectory=trajectory,
                            skipped_steps=[_step_context(item) for item in skipped],
                            candidate_image_paths=[
                                _artifact_path(
                                    trajectory_root, item.image, require_file=False
                                )
                                for item in candidate.confirmation_steps
                            ],
                            reference_image_paths=[
                                _artifact_path(
                                    trajectory_root,
                                    node.reference.image,
                                    require_file=False,
                                )
                                for node in candidate.confirmation_nodes
                                if node.reference is not None
                            ],
                            candidate_steps=[
                                _step_context(item)
                                for item in candidate.confirmation_steps
                            ],
                            reference_steps=[
                                _step_context(node.reference)
                                for node in candidate.confirmation_nodes
                                if node.reference is not None
                            ],
                        )
                    except Exception as exc:
                        raise RuntimeError(
                            f"截图一致性复核失败: trajectory={trajectory}, "
                            f"step={step.step_index}, skip_count={candidate.skip_count}, "
                            f"error={exc}"
                        ) from exc
                    statistics.resync_review_count += 1
                    if review.cached:
                        statistics.resync_review_cache_hit_count += 1
                    review_dict = review.to_dict(confidence_threshold)
                    accepted = bool(review_dict["accepted"])
                    if accepted:
                        statistics.resync_accepted_count += 1
                        statistics.resync_ignored_step_count += candidate.skip_count
                        confirmation_ids = [node.id for node in candidate.confirmation_nodes]
                        confirmation_steps = [
                            item.step_index for item in candidate.confirmation_steps
                        ]
                        for skipped_step in skipped:
                            skipped_step.effective_intermediate = True
                            skipped_step.skip_count = candidate.skip_count
                            skipped_step.confirmation_node_ids = confirmation_ids
                            skipped_step.confirmation_step_indices = confirmation_steps
                            skipped_step.alignment_review = review_dict
                            _record_ignored_step(
                                decisions,
                                skipped_step,
                                parent_node=parent_id,
                                decision_source="uncertain_structural_resync",
                                reason=(
                                    "low-confidence one/two-step gap followed by two full "
                                    "action/bbox matches and accepted Qwen screenshot-state review"
                                ),
                                confidence_threshold=confidence_threshold,
                            )
                        position += candidate.skip_count
                        continue
                    statistics.resync_rejected_count += 1
                    step.alignment_review = review_dict

            if not reused:
                matched = Node(id=next_id, depth=current.depth + 1, reference=step)
                next_id += 1
                current.children.append(matched)
            assert matched is not None
            matched.occurrences.append(
                {
                    "trajectory": trajectory,
                    "step": step.step_index,
                    "excel_row": step.excel_row,
                    "image": step.image,
                    "xml": step.xml,
                    "action": step.action,
                    "action_text": step.action_text,
                    "summary": step.summary,
                    "observation": step.observation,
                    "actions_box": step.actions_box,
                    "score": best_score,
                    "reused": reused,
                    "classification": _classification_dict(
                        step, confidence_threshold
                    ),
                    "alignment_review": step.alignment_review,
                }
            )
            step.counted_in_tree = True
            step.tree_decision = "reuse" if reused else "branch"
            step.decision_source = "prefix_match" if reused else "tree_branch"
            step.tree_node_id = matched.id
            step.tree_score = best_score
            decisions.append(
                {
                    "trajectory": trajectory,
                    "step": step.step_index,
                    "excel_row": step.excel_row,
                    "parent_node": parent_id,
                    "node": matched.id,
                    "score": best_score,
                    "decision": step.tree_decision,
                    "decision_source": step.decision_source,
                    "classification": _classification_dict(
                        step, confidence_threshold
                    ),
                    "alignment_review": step.alignment_review,
                    "action": step.action,
                    "summary": step.summary,
                    "image": step.image,
                    "xml": step.xml,
                }
            )
            current = matched
            position += 1
        current.terminal_trajectories.append(trajectory)
    return root, decisions, statistics


def source_trajectory_audit(
    trajectories: list[tuple[str, list[Step]]],
    confidence_threshold: float,
) -> list[dict[str, Any]]:
    return [
        {
            "trajectory": trajectory,
            "original_step_count": len(steps),
            "tree_step_count": sum(step.counted_in_tree for step in steps),
            "ignored_incidental_step_count": sum(
                not step.counted_in_tree for step in steps
            ),
            "ignored_intermediate_terminate_count": sum(
                step.excluded_intermediate_terminate for step in steps
            ),
            "steps": [step.audit_dict(confidence_threshold) for step in steps],
        }
        for trajectory, steps in trajectories
    ]


def write_output(
    root: Node,
    decisions: list[dict[str, Any]],
    statistics: BuildStatistics,
    trajectories: list[tuple[str, list[Step]]],
    *,
    model_name: str,
    confidence_threshold: float,
    max_incidental_skip: int,
    json_path: Path,
    extra_metadata: dict[str, Any] | None = None,
) -> None:
    all_steps = [step for _, steps in trajectories for step in steps]
    category_counts = Counter(
        step.classification.category
        for step in all_steps
        if step.classification is not None
    )
    tree = root.to_dict()
    tree.update(
        {
            "trajectory_count": len(trajectories),
            "matching_threshold": FULL_SCORE,
            "box_overlap_threshold": BOX_OVERLAP_THRESHOLD,
            "model_name": model_name,
            "confidence_threshold": confidence_threshold,
            "max_incidental_skip": max_incidental_skip,
            "resync_confirmation_steps": RESYNC_CONFIRM_STEPS,
            "original_step_count": len(all_steps),
            "tree_step_count": sum(step.counted_in_tree for step in all_steps),
            "ignored_incidental_step_count": sum(
                not step.counted_in_tree for step in all_steps
            ),
            "ignored_intermediate_terminate_count": sum(
                step.excluded_intermediate_terminate for step in all_steps
            ),
            "classification_count": sum(
                step.classification is not None for step in all_steps
            ),
            "classified_intermediate_count": sum(
                bool(step.classification and step.classification.is_intermediate)
                for step in all_steps
            ),
            "policy_candidate_count": sum(
                step.classification_candidate for step in all_steps
            ),
            "bounded_ignored_step_count": sum(
                step.decision_source == "bounded_classification"
                for step in all_steps
            ),
            "retained_long_or_terminal_intermediate_count": sum(
                step.classification_candidate and step.counted_in_tree
                for step in all_steps
            ),
            "low_confidence_count": sum(step.uncertain for step in all_steps),
            "classification_category_counts": dict(sorted(category_counts.items())),
            **statistics.to_dict(),
            "matching_policy": (
                "Every source step except non-final terminate actions is classified. "
                "Non-final terminate actions remain in the audit but are excluded from "
                "the tree and quality workbook before model calls. High-confidence one/two-step "
                "transient runs are ignored only when followed by two high-confidence "
                "normal steps, including recovery into a new branch. Low-confidence "
                "steps may be ignored only after a mismatch, two full action/bbox "
                "matches, and an accepted Qwen screenshot-state review. Confidently "
                "normal, long, and terminal states remain tree nodes."
            ),
            "source_trajectories": source_trajectory_audit(
                trajectories, confidence_threshold
            ),
            "decisions": decisions,
            **(extra_metadata or {}),
        }
    )
    json_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = json_path.with_name(f".{json_path.name}.tmp")
    temporary_path.write_text(
        json.dumps(tree, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    temporary_path.replace(json_path)


def count_nodes(node: Node) -> int:
    return 1 + sum(count_nodes(child) for child in node.children)


def _resolved(value: Path) -> Path:
    return value.expanduser().resolve()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Build an action-prefix tree with Qwen-confirmed short-gap resync."
    )
    parser.add_argument("xlsx", nargs="?", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--sheet", help="工作表名称；默认读取活动工作表")
    parser.add_argument(
        "--trajectory-root",
        type=Path,
        default=DEFAULT_TRAJECTORY_ROOT,
        help="image/xml 相对路径的资源根目录",
    )
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--env-file", type=Path, default=DEFAULT_ENV)
    parser.add_argument(
        "--cache",
        type=Path,
        default=DEFAULT_CLASSIFICATION_CACHE,
        help="全轨迹中间状态分类缓存",
    )
    parser.add_argument(
        "--alignment-cache",
        type=Path,
        default=DEFAULT_ALIGNMENT_CACHE,
        help="短跳步截图一致性复核缓存",
    )
    parser.add_argument(
        "--confidence-threshold",
        type=float,
        default=0.8,
        help="Qwen 判断生效所需的最低置信度（默认 0.8）",
    )
    parser.add_argument(
        "--max-incidental-skip",
        type=int,
        choices=(1, 2),
        default=MAX_INCIDENTAL_SKIP,
        help="mismatch 后最多尝试跳过的步骤数（默认 2）",
    )
    args = parser.parse_args(argv)
    if not 0.0 <= args.confidence_threshold <= 1.0:
        parser.error("--confidence-threshold must be between 0 and 1")

    xlsx_path = _resolved(args.xlsx)
    trajectory_root = _resolved(args.trajectory_root)
    output_path = _resolved(args.output)
    env_path = _resolved(args.env_file)
    classification_cache = _resolved(args.cache)
    alignment_cache = _resolved(args.alignment_cache)
    if not xlsx_path.is_file():
        parser.error(f"xlsx not found: {xlsx_path}")
    if not trajectory_root.is_dir():
        parser.error(f"trajectory root not found: {trajectory_root}")

    trajectories = load_trajectories(xlsx_path, args.sheet)
    model_name = configure_reviewer_environment(env_path, module="tree")
    classifier = QwenIntermediateStateClassifier(model_name, classification_cache)
    classify_trajectories(
        trajectories, classifier, trajectory_root, args.confidence_threshold
    )
    for _, steps in trajectories:
        apply_bounded_skip_policy(
            steps,
            confidence_threshold=args.confidence_threshold,
            max_skip=args.max_incidental_skip,
        )
    alignment_reviewer = QwenStateAlignmentReviewer(model_name, alignment_cache)
    # All model work happens before output writing. API/response failures leave
    # the previous complete JSON untouched while successful reviews stay cached.
    root, decisions, statistics = build_tree(
        trajectories,
        confidence_threshold=args.confidence_threshold,
        trajectory_root=trajectory_root,
        alignment_reviewer=alignment_reviewer,
        max_incidental_skip=args.max_incidental_skip,
    )
    write_output(
        root,
        decisions,
        statistics,
        trajectories,
        model_name=model_name,
        confidence_threshold=args.confidence_threshold,
        max_incidental_skip=args.max_incidental_skip,
        json_path=output_path,
    )

    ignored = sum(item["decision"] == "ignore_intermediate" for item in decisions)
    branches = sum(item["decision"] == "branch" for item in decisions)
    reuses = sum(item["decision"] == "reuse" for item in decisions)
    print(
        f"Trajectories: {len(trajectories)}; original steps: {len(decisions)}; "
        f"tree steps: {len(decisions) - ignored}; action nodes: {count_nodes(root) - 1}; "
        f"branches: {branches}; reuses: {reuses}; ignored short insertions: {ignored}; "
        f"alignment reviews: {statistics.resync_review_count}; "
        f"accepted: {statistics.resync_accepted_count}; "
        f"rejected: {statistics.resync_rejected_count}"
    )
    print(f"JSON: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
